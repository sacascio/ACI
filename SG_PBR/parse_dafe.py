#!/usr/bin/env python

import openpyxl
import operator
import ipaddress
from netaddr import *
import shutil
import sys
import getopt
import os.path
import fnmatch
import re
from IPy import IP
# from xlrd import open_workbook, XLRDError
import json
# from fileinput import filename
import warnings
from ciscoconfparse import CiscoConfParse

warnings.filterwarnings("ignore")

def append_l3out_no_bd(vrf_no_epg,all_l3_contracts,dafe_file):
	worksheets = []
	l3out_to_add = []

	wb = openpyxl.load_workbook(dafe_file, data_only=True)

    	for sheet in wb:
        	worksheets.append(sheet.title)
    	wb.close()

    	wb.active = worksheets.index('l3out')
    	ws = wb.active

	for v in vrf_no_epg:
    		row_start = ws.min_row
    		row_end = ws.max_row
    		for x in range(row_start + 1, row_end + 1):
			cell = 'C' + str(x)
                	vrf_cell = ws[cell].value
			if vrf_cell == v:
				cell = 'A' + str(x)
                		l3out_cell = ws[cell].value
				l3out_to_add.append([l3out_cell,v])

	wb.active = worksheets.index('external_epg')
        ws = wb.active

	for l3s in l3out_to_add:
		row_start = ws.min_row
                row_end = ws.max_row

		for x in range(row_start + 1, row_end + 1):
                        cell = 'B' + str(x)
                        l3out_cell = ws[cell].value
			l3s_found = l3s[0]	
			vrf_found = l3s[1]
	
			if l3out_cell == l3s_found:
                        	cell = 'A' + str(x)
                        	tenant_cell = ws[cell].value

                        	cell = 'D' + str(x)
                        	extepg_cell = ws[cell].value
                        	
				cell = 'G' + str(x)
                        	ccontract_cell = ws[cell].value

				cell = 'I' + str(x)
                        	pcontract_cell = ws[cell].value

				if tenant_cell not in all_l3_contracts:
					all_l3_contracts[tenant_cell] = {}
				if vrf_found not in all_l3_contracts[tenant_cell]:
					all_l3_contracts[tenant_cell][vrf_found] = {}
				if l3out_cell not in all_l3_contracts[tenant_cell][vrf_found]:
					all_l3_contracts[tenant_cell][vrf_found][l3out_cell] = {}
				if extepg_cell not in all_l3_contracts[tenant_cell][vrf_found][l3out_cell]:
					all_l3_contracts[tenant_cell][vrf_found][l3out_cell][extepg_cell] = {}
					all_l3_contracts[tenant_cell][vrf_found][l3out_cell][extepg_cell]['consumer'] = []
					all_l3_contracts[tenant_cell][vrf_found][l3out_cell][extepg_cell]['provider'] = []

				ccontracts = ccontract_cell.split(",")
				for cc in ccontracts:
					all_l3_contracts[tenant_cell][vrf_found][l3out_cell][extepg_cell]['consumer'].append(cc)
						
				pcontracts = pcontract_cell.split(",")
				for pc in pcontracts:
					all_l3_contracts[tenant_cell][vrf_found][l3out_cell][extepg_cell]['provider'].append(pc)
					
	return all_l3_contracts

def load_aci_json(dc,district):

	pattern = 'ce2_' + dc.upper() + '_' + district.upper() + '_FullBackup_to_*'
	files = os.listdir('.')
    	for name in files:
        	if fnmatch.fnmatch(name, pattern):
            		if not bool(re.search('^~', name, re.IGNORECASE)):
                		aci_json_file = name
                		break
    	try:
        	aci_json_file
    	except NameError:
        	print "ERROR: ACI JSON File not found"
        	sys.exit(9)

	with open(aci_json_file) as f:
		data = json.load(f)

	return data


def get_subnet_nhip(l3out,aci_json_cfg,tenant):
	routes = {}
	for tenant_id in aci_json_cfg['polUni']['children']:
		for tn in tenant_id:
			if tn == 'fvTenant' and tenant_id[tn]['attributes']['name'] == tenant:
				for l3extid in tenant_id[tn]['children']:
					for l3ext in l3extid:
						if l3ext == 'l3extOut' and l3extid[l3ext]['attributes']['name'] == l3out:
							for npid in l3extid[l3ext]['children']:
								for l_npid in npid:
									if l_npid == 'l3extLNodeP':
										for l3extLNodeP_children in npid[l_npid]['children']:
											for l3extLNodeP_children_keys in l3extLNodeP_children:
												if l3extLNodeP_children_keys == 'l3extRsNodeL3OutAtt':
													leafid = l3extLNodeP_children[l3extLNodeP_children_keys]['attributes']['tDn']
													if 'children' not in l3extLNodeP_children[l3extLNodeP_children_keys]:
														if leafid not in routes:
															routes[leafid] = []
														routes[leafid].append('NO STATIC ROUTES FOUND')
													else:
														for l3extRsNodeL3OutAtt in l3extLNodeP_children[l3extLNodeP_children_keys]['children']:
															for iproute in l3extRsNodeL3OutAtt:
																if iproute == 'ipRouteP':
																	snet = l3extRsNodeL3OutAtt[iproute]['attributes']['ip']
																	if leafid not in routes:
																		routes[leafid] = []
																		routes[leafid].append(snet)
																	else:
																		
																		routes[leafid].append(snet)
																	routes[leafid].sort()
	return routes

def add_other_l3outs_to_vrfl3out(total_l3_contracts,l3out,vrflist,save_ttype):
	delta_l3outs = {}
	#print json.dumps(l3out)
	for tenant in total_l3_contracts:
		if tenant == 'common':
			continue
		for vrf in total_l3_contracts[tenant]:
			if vrf is None:
				continue
			if vrf in save_ttype:
				vrftype = save_ttype[vrf]
			elif bool(re.search('DNS', vrf, re.IGNORECASE)) and vrf is not None:
				vrftype = 'B'
			else:
				vrftype = 'X'

			for l3outs in total_l3_contracts[tenant][vrf]:
				for extepg in total_l3_contracts[tenant][vrf][l3outs]:
					if l3outs not in l3out and vrf in vrflist:
						l3out[l3outs] = {}
						l3out[l3outs][extepg] = {}
						for c in total_l3_contracts[tenant][vrf][l3outs][extepg]['consumer']:
							l3out[l3outs][extepg][c] = {}
							if tenant not in delta_l3outs:
								delta_l3outs[tenant] = {}
							if vrf not in delta_l3outs[tenant]:
								delta_l3outs[tenant][vrf] = {}
								delta_l3outs[tenant][vrf][vrftype] = {}
							if l3outs not in delta_l3outs[tenant][vrf][vrftype]:
								delta_l3outs[tenant][vrf][vrftype][l3outs] = {}
							if extepg not in delta_l3outs[tenant][vrf][vrftype][l3outs]:
								delta_l3outs[tenant][vrf][vrftype][l3outs][extepg] = {}
							delta_l3outs[tenant][vrf][vrftype][l3outs][extepg][c] = {}
						for c in total_l3_contracts[tenant][vrf][l3outs][extepg]['provider']:
							l3out[l3outs][extepg][c] = {}
							if tenant not in delta_l3outs:
								delta_l3outs[tenant] = {}
							if vrf not in delta_l3outs[tenant]:
								delta_l3outs[tenant][vrf] = {}
								delta_l3outs[tenant][vrf][vrftype] = {}
							if l3outs not in delta_l3outs[tenant][vrf][vrftype]:
								delta_l3outs[tenant][vrf][vrftype][l3outs] = {}
							if extepg not in delta_l3outs[tenant][vrf][vrftype][l3outs]:
								delta_l3outs[tenant][vrf][vrftype][l3outs][extepg] = {}
							delta_l3outs[tenant][vrf][vrftype][l3outs][extepg][c] = {}
	return (l3out,delta_l3outs)

def check_epgs_for_contracts(write_to_aci_cfg,c):

	for t_epg in write_to_aci_cfg:
        	for v_epg in write_to_aci_cfg[t_epg]:
                	for e_epg in write_to_aci_cfg[t_epg][v_epg]:
                        	for e_data in write_to_aci_cfg[t_epg][v_epg][e_epg]:
                                	if c in e_data['curr_c_contracts'] or c in e_data['curr_p_contracts']:
                                        	return (True,e_epg,e_data['ap'])

	return (False,None,None)

def remove_dups(msg):
  return list(dict.fromkeys(msg))

def check_other_l3_outs(total_l3_contracts,c,l3outs,externalepg,delta_l3outs,output_messages):
	total_retval = 'no'
	delta_retval = 'no'
	msg = []
	pc = ['provider','consumer']
	
	for t_tenant in delta_l3outs:
       		for t_vrf in delta_l3outs[t_tenant]:
                	for ccctype in delta_l3outs[t_tenant][t_vrf]:
                        	for t_l3out in delta_l3outs[t_tenant][t_vrf][ccctype]:
                                	for t_extepg in delta_l3outs[t_tenant][t_vrf][ccctype][t_l3out]:
						for t_c in delta_l3outs[t_tenant][t_vrf][ccctype][t_l3out][t_extepg]:
							if t_c == c: 
								#print "WARNING: Contract %s on L3OUT: %s, EXT EPG %s, exists on external EPG %s as consumer SLATED FOR REMOVAL" % (c,l3outs,externalepg,extepg)
								delta_retval = 'yes'
	for t_pc in pc:	
		for tenant in total_l3_contracts:
			for vrf in total_l3_contracts[tenant]:
				for l3out in total_l3_contracts[tenant][vrf]:
					for extepg in total_l3_contracts[tenant][vrf][l3out]:
						if c in total_l3_contracts[tenant][vrf][l3out][extepg][t_pc]:
							if l3outs != l3out and extepg != externalepg:
								total_retval = 'yes'
								if delta_retval == 'no':
									# save for later
									#print "WARNING: Contract %s exists on L3OUT: %s, EXT EPG %s as %s - will not be removed" % (c,l3out,extepg,t_pc)
									output_messages[vrf].append("WARNING: Contract %s exists on L3OUT: %s, EXT EPG %s as %s - will not be removed" % (c,l3out,extepg,t_pc))
								else:
									# save for later
									#msg.append("OK: Contract %s does not exist on any other on L3OUT, ADDING TO L3OUT CLEANUP" % (c))
									output_messages[vrf].append("OK: Contract %s does not exist on any other on L3OUT, ADDING TO L3OUT CLEANUP" % (c))
	#msg = remove_dups(msg)
	#if len(msg) > 0:
	#	print '\n'.join(msg)	

	if total_retval == 'yes' and delta_retval == 'no':
		return 'yes'
	else:
		return 'no'

def write_contract_removal(tenant,vrf,c,l3out,extepg,ctype,dir_path,f_epg,f_ap,objtype):
	if tenant == 'Limited':
		tenant = 'LTD'
	fname = vrf + " 3 Type " + ctype + " - Remove contract from L3Out and EPG as provider_consumer.csv"
        if not os.path.isfile(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname):
        	f = open(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname, "a")
                f.write("TENANT,AP,EPG,OLD_EPG_CONTRACT,L3OUT,NETWORK,OLD_L3_CONTRACT" + '\n')
                f.close()
        f = open(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname, "a")
	if objtype == 'l3out':
        	f.write(tenant + ",,,," + l3out + "," + extepg + "," + c + '\n')
	else:
		f.write(tenant + "," + f_ap + "," + f_epg + "," + c + '\n')
        f.close()
	
def write_contract_addition(tenant,vrf,l3out,extepg,dir_path,ctype):
	fname = vrf + " 1 Type " + ctype + " - Associate contracts to L3out as provider.csv"
        if not os.path.isfile(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname):
        	f = open(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname, "a")
                f.write("TENANT,L3OUT,NETWORK,NEW_L3_CONTRACT" + '\n')
                f.close()
        f = open(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname, "a")
        f.write(tenant + "," + l3out + "," + extepg + "," + tenant + "-" + vrf + "-" + "SG-PBR-Permit_Any" + '\n')
        f.close()

def write_contract_addition_consumer(tenant,vrf,l3out,extepg,dir_path,ctype):
	fname = vrf + " 1 Type " + ctype + " - Associate contracts to L3out as consumer.csv"
        if not os.path.isfile(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname):
        	f = open(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname, "a")
                f.write("TENANT,L3OUT,NETWORK,NEW_L3_CONTRACT" + '\n')
                f.close()
        f = open(dir_path + "/ACI_" + tenant + "-" + vrf + "/" + fname, "a")
        f.write(tenant + "," + l3out + "," + extepg + "," + tenant + "-" + vrf + "-" + "SG-PBR-Permit_Any" + '\n')
        f.close()

def update_aci_contract_verification_file(tenant,vrf,dir_path,objtype,ap_or_l3out,epg_or_extepg):
	fname = tenant + "-" + vrf
        f = open(dir_path + "/ACI_CONTRACT_VERIFICATION" + "/" + fname, "a")
        f.write(objtype + "," + tenant + "," + ap_or_l3out + "," + epg_or_extepg + '\n')
        f.close()

def get_total_l3_contracts(dafe_file):
	worksheets = []
	d_l3out = {}

	wb = openpyxl.load_workbook(dafe_file, data_only=True)

    	for sheet in wb:
        	worksheets.append(sheet.title)
    	wb.close()

    	wb.active = worksheets.index('l3out')
    	ws = wb.active
    	row_start = ws.min_row
    	row_end = ws.max_row

    	for x in range(row_start + 1, row_end + 1):
        	cell = 'B' + str(x)
        	tenant = ws[cell].value
        	
		cell = 'C' + str(x)
        	vrf = ws[cell].value

		cell = 'A' + str(x)
        	l3out = ws[cell].value

		if tenant not in d_l3out:
			d_l3out[tenant] = {}
		if vrf not in d_l3out[tenant]:
			d_l3out[tenant][vrf] = {}
		if l3out not in d_l3out[tenant][vrf]:
			d_l3out[tenant][vrf][l3out] = {}

	wb.active = worksheets.index('external_epg')
    	ws = wb.active
    	row_start = ws.min_row
    	row_end = ws.max_row
	
	for x in range(row_start + 1, row_end + 1):
		cell = 'A' + str(x)
        	tenant = ws[cell].value
		
		cell = 'B' + str(x)
        	l3out = ws[cell].value

		cell = 'D' + str(x)
        	extepg = ws[cell].value
		
		cell = 'G' + str(x)
        	c_contracts = ws[cell].value

		cell = 'I' + str(x)
        	p_contracts = ws[cell].value

		for t in d_l3out:
			for v in d_l3out[t]:
				for l in d_l3out[t][v]:
					if l == l3out:
						d_l3out[t][v][l][extepg] = {}
						d_l3out[t][v][l][extepg]['provider'] = []
						d_l3out[t][v][l][extepg]['consumer'] = []

						if c_contracts is not None:
							x = c_contracts.split(",")
							for i in x:
								d_l3out[t][v][l][extepg]['consumer'].append(i)	

						if p_contracts is not None:
							x = p_contracts.split(",")
							for i in x:
								d_l3out[t][v][l][extepg]['provider'].append(i)	


	return d_l3out

def write_new_n7k_configs(vrfmember, p2psubnets, dc, district, n7k_data):
    district = district.lower()
    dc = dc.lower()
    dir_path = './output'
    inner_bgp_config = {}
    outer_bgp_config = {}
    bgp_shut_inner = {}
    bgp_shut_outer = {}
    bgp_rb_inner = {}
    bgp_rb_outer = {}
    bgp_rb_outer = {}
    svi_cleanup = {}
    bgp_cleanup_inner = {}
    bgp_cleanup_outer = {}
    svi_inner_cutover = {}
    subint_inner_cutover = {}
    subint_outer_cutover = {}
    cutover_dir = dir_path + "/" + "N7K_CUTOVER" + "/" + vrfmember
    rollback_dir = dir_path + "/" + "N7K_ROLLBACK" + "/" + vrfmember
    cleanup_dir = dir_path + "/" + "N7K_NEXT_CLEANUP" + "/" + vrfmember

    if not os.path.exists(dir_path + "/" + "N7K_PREWORK"):
        os.mkdir(dir_path + "/" + "N7K_PREWORK")
    if not os.path.exists(dir_path + "/" + "N7K_CUTOVER"):
        os.mkdir(dir_path + "/" + "N7K_CUTOVER")
    if not os.path.exists(dir_path + "/" + "N7K_ROLLBACK"):
        os.mkdir(dir_path + "/" + "N7K_ROLLBACK")
    if not os.path.exists(dir_path + "/" + "N7K_NEXT_CLEANUP"):
        os.mkdir(dir_path + "/" + "N7K_NEXT_CLEANUP")

    os.mkdir(cutover_dir)
    os.mkdir(rollback_dir)
    os.mkdir(cleanup_dir)

    k = 0

    # Get Inner BGP Details
    for n7k in n7k_data:
        # Create empty creds file - used for later
        # Modified so that there is 1 Creds file for N7K

        # open(dir_path + "/" + 'n7k_creds', 'a').close()
        open(dir_path + "/" + n7k + '_creds', 'a').close()

        if bool(re.search('outer', n7k, re.IGNORECASE)):
            continue
        bgp_rb_inner[n7k] = {}
        bgp_rb_inner[n7k]['neighbors'] = []
        bgp_rb_inner[n7k]['subint'] = []
        bgp_rb_inner[n7k]['svi'] = []
        bgp_cleanup_inner[n7k] = {}
        bgp_cleanup_inner[n7k]['neighbors'] = []

        neighbors = n7k_data[n7k][vrfmember]['neighbors']
        local_as = n7k_data[n7k][vrfmember]['local_as']
        remote_as = n7k_data[n7k][vrfmember]['remote_as']
        bgp_shut_inner[n7k] = {}
        bgp_shut_inner[n7k][vrfmember] = []
        bgp_shut_inner[n7k][vrfmember].append("! Shutdown the BGP adjacency to the outer N7K VDC connected to the FW")
        bgp_shut_inner[n7k][vrfmember].append("router bgp " + local_as)
        bgp_shut_inner[n7k][vrfmember].append(" vrf " + vrfmember)

        # Squeeze in rollback for new adjacencies
        bgp_rb_inner[n7k]['neighbors'].append("router bgp " + local_as)
        bgp_rb_inner[n7k]['neighbors'].append(" vrf " + vrfmember)

        # Squeeze in cleanup for new adjacencies
        bgp_cleanup_inner[n7k]['neighbors'].append("router bgp " + local_as)
        bgp_cleanup_inner[n7k]['neighbors'].append(" vrf " + vrfmember)

        for n in neighbors:
            bgp_shut_inner[n7k][vrfmember].append(" neighbor " + n + " remote-as " + remote_as)
            bgp_shut_inner[n7k][vrfmember].append("   shutdown")

            # Rollback
            bgp_rb_inner[n7k]['neighbors'].append("   neighbor " + n + " remote-as " + remote_as)
            bgp_rb_inner[n7k]['neighbors'].append("     no shutdown")

            # cleanup
            bgp_cleanup_inner[n7k]['neighbors'].append("   no neighbor " + n + " remote-as " + remote_as)

    # Get outer VDC BGP details
    for n7k in n7k_data:
        # Create empty creds file - used for later
        # Not needed - already created in the inner which will apply to all VDC's
        # open(dir_path + "/" + n7k + '_creds', 'a').close()

        if bool(re.search('inner', n7k, re.IGNORECASE)):
            continue
        bgp_rb_outer[n7k] = {}
        bgp_rb_outer[n7k]['neighbors'] = []
        bgp_rb_outer[n7k]['subint'] = []
        bgp_rb_outer[n7k]['svi'] = []
        bgp_cleanup_outer[n7k] = {}
        bgp_cleanup_outer[n7k]['neighbors'] = []

        for svi in n7k_data[n7k]:
            if svi == 'P2P':
                continue
            remote_as = n7k_data[n7k][svi]['remote_as']
            local_as = n7k_data[n7k][svi]['local_as']
            if 'vrf' not in n7k_data[n7k][svi]:
                continue
            for vrf in n7k_data[n7k][svi]['vrf']:
                if vrf == vrfmember:
                    outer_neighbors = n7k_data[n7k][svi]['vrf'][vrf]['neighbor_ip']
                    bgp_shut_outer[n7k] = []
                    bgp_shut_outer[n7k].append("router bgp " + local_as)

                    # Squeeze in rollback for BGP Neighbors and the SVI
                    bgp_rb_outer[n7k]['neighbors'].append("router bgp " + local_as)
                    bgp_rb_outer[n7k]['svi'].append("interface Vlan" + svi)
                    bgp_rb_outer[n7k]['svi'].append(" no shutdown")

                    # Squeeze in cleanup for BGP Neighbors
                    bgp_cleanup_outer[n7k]['neighbors'].append("router bgp " + local_as)

                    for n in outer_neighbors:
                        bgp_shut_outer[n7k].append(" neighbor " + n + " remote-as " + remote_as)
                        bgp_shut_outer[n7k].append("   shutdown")

                        # Rollback
                        bgp_rb_outer[n7k]['neighbors'].append(" neighbor " + n + " remote-as " + remote_as)
                        bgp_rb_outer[n7k]['neighbors'].append("    no shutdown")

                        # Cleanup
                        bgp_cleanup_outer[n7k]['neighbors'].append(" no neighbor " + n + " remote-as " + remote_as)

    if district.upper() == 'SDE':
        numn7k = ['1', '2']
    else:
        numn7k = ['1', '2', '3', '4']

    for i in numn7k:
        if district.upper() == 'SDE':
            n7kname = dc + district + 'nxc' + i + district + 'inner'

        else:
            n7kname = dc + 'dcinxc' + i + district.lower() + 'inner'

        f = open(dir_path + "/" + "N7K_CUTOVER" + "/" + "execute_cutover_" + vrfmember + ".sh", "a")
        f.write("../push_to_n7k.py -f " + vrfmember + "/" + n7kname + " -c ../" + n7kname + "_creds" + '\n')
        f.write("echo FINISHED UPDATING " + n7kname + '\n\n')
        f.close()
        fshowc = open(dir_path + "/" + "N7K_CUTOVER" + "/" + "execute_show_commands_" + vrfmember + ".sh", "a")
        fshowc.write(
            "../push_to_n7k.py -f " + vrfmember + "/" + n7kname + "_inner_show_commands" + " -c ../" + n7kname + "_creds" + ' > ' + n7kname + "_inner_output" + '\n\n')
        fshowc.close()
        encap = n7k_data[n7kname][vrfmember]['svi']
        inner_bgp_as = n7k_data[n7kname][vrfmember]['local_as']
        outer_bgp_as = n7k_data[n7kname][vrfmember]['remote_as']
        inner_bgp_config[n7kname] = {}
        inner_bgp_config[n7kname][vrfmember] = []
        subint_inner_cutover[n7kname] = []

        # Start preparing inner/outer BGP config
        inner_bgp_config[n7kname][vrfmember].append(
            "! Add new BGP neighbors to VRF " + vrfmember + " using the N7K Outer VDC IP addresses")
        inner_bgp_config[n7kname][vrfmember].append(
            "! These adjacencies will not come up until the N7K Outer VDCs are configured and enabled" + '\n')

        inner_bgp_config[n7kname][vrfmember].append("router bgp " + inner_bgp_as)
        inner_bgp_config[n7kname][vrfmember].append(" vrf " + vrfmember)
        inner_bgp_config[n7kname][vrfmember].append("  address-family ipv4 unicast")
        inner_bgp_config[n7kname][vrfmember].append("   maximum-paths 8")

        if not os.path.exists(dir_path + "/" + "N7K_PREWORK" + "/" + n7kname + "_" + vrfmember):
            f = open(dir_path + "/" + "N7K_PREWORK" + "/" + n7kname + "_" + vrfmember, "a")
            f.write("! Create sub interfaces to outer VDCs in VRF " + vrfmember + " in a shutdown state" + '\n')
            f.write("configure terminal" + '\n')
            f.close()

            f = open(dir_path + "/" + "N7K_PREWORK" + "/" + "execute_prework.sh", "a")
            f.write("../push_to_n7k.py -f " + n7kname + "_" + vrfmember + " -c ../" + n7kname + "_creds" + '\n')
            f.write("echo FINISHED UPDATING VRF " + vrfmember + " ON " + n7kname + '\n\n')
            f.close()
        else:
            f = open(dir_path + "/" + "N7K_PREWORK" + "/" + n7kname + "_" + vrfmember, "a")
            f.write("! Create sub interfaces to outer VDCs in VRF " + vrfmember + " in a shutdown state" + '\n')
            f.close()

        # if not os.path.exists(cutover_dir + "/" +  n7kname):
        #	f = open(cutover_dir + "/" +  n7kname, "a")
        #	f.write("configure terminal" + '\n')
        #	f.close()

        for inner_int in sorted(n7k_data[n7kname]['P2P']):
            for outer_int in n7k_data[n7kname]['P2P'][inner_int]:
                for outer_7k in n7k_data[n7kname]['P2P'][inner_int][outer_int]:
                    if outer_7k not in subint_outer_cutover:
                        f = open(dir_path + "/" + "N7K_CUTOVER" + "/" + "execute_cutover_" + vrfmember + ".sh", "a")
                        f.write(
                            "../push_to_n7k.py -f " + vrfmember + "/" + outer_7k + " -c ../" + outer_7k + "_creds" + '\n')
                        f.write("echo FINISHED UPDATING " + outer_7k + '\n\n')
                        f.close()
                        fshowc = open(
                            dir_path + "/" + "N7K_CUTOVER" + "/" + "execute_show_commands_" + vrfmember + ".sh", "a")
                        fshowc.write(
                            "../push_to_n7k.py -f " + vrfmember + "/" + outer_7k + "_outer_show_commands" + " -c ../" + outer_7k + "_creds" + ' > ' + outer_7k + "_outer_output" + '\n\n')
                        fshowc.close()
                        subint_outer_cutover[outer_7k] = []

                    if outer_7k not in outer_bgp_config:
                        outer_bgp_config[outer_7k] = {}
                    if vrfmember not in outer_bgp_config[outer_7k]:
                        outer_bgp_config[outer_7k][vrfmember] = []
                        outer_bgp_config[outer_7k][vrfmember].append(
                            "! Add new BGP neighbors to VRF " + vrfmember + " using the N7K Inner VDC IP addresses")
                        outer_bgp_config[outer_7k][vrfmember].append(
                            "! These adjacencies will not come up until the N7K Inner VDCs are configured and enabled" + '\n')
                        outer_bgp_config[outer_7k][vrfmember].append("router bgp " + outer_bgp_as)

                    # Write Inner config
                    f = open(dir_path + "/" + "N7K_PREWORK" + "/" + n7kname + "_" + vrfmember, "a")
                    f.write("interface Ethernet" + inner_int + "." + encap + '\n')
                    f.write(" description To_" + outer_7k + "_E" + outer_int + "." + encap + '\n')
                    f.write(" shutdown" + '\n')
                    f.write(" vrf member " + vrfmember + '\n')
                    f.write(" mtu 9192" + '\n')
                    f.write(" encapsulation dot1Q " + encap + '\n')

                    # Write cutover part to enable the new sub interfaces on inner
                    subint_inner_cutover[n7kname].append("interface Ethernet" + inner_int + "." + encap)
                    subint_inner_cutover[n7kname].append(" no shutdown")

                    # fsub = open(cutover_dir + "/" +  n7kname, "a")
                    # fsub.write("! Enable sub interfaces to outer VDCs for VRF " + vrfmember +  '\n')
                    # fsub.write("interface Ethernet" + inner_int + "." + encap + '\n')
                    # fsub.write(" no shutdown" + '\n')
                    # fsub.write('\n')
                    # fsub.close()

                    # Write Rollback part for inner
                    if bool(re.search('inner', n7kname, re.IGNORECASE)):
                        bgp_rb_inner[n7kname]['subint'].append("no interface Ethernet" + inner_int + "." + encap)

                    # Custom IP addressing for SDE because we already started this way
                    if k == 2 and district.upper() == 'SDE':
                        ipinner = p2psubnets[3]
                    elif k == 3 and district.upper() == 'SDE':
                        ipinner = p2psubnets[2]
                    else:
                        ipinner = p2psubnets[k]

                    ipinnerx = ipinner.split('/')
                    ipinner = IPAddress(ipinnerx[0])
                    mask = ipinnerx[1]
                    ipinner = ipinner + 1
                    ipouter = ipinner + 1
                    ipinner = str(IPAddress(ipinner))
                    ipouter = str(IPAddress(ipouter))
                    fcfg = open(dir_path + "/" + "N7K_IP_MAPPING_" + vrfmember + ".txt", "a")
                    fcfg.write(
                        n7kname + "," + inner_int + "." + encap + "," + ipinner + "---" + outer_7k + "," + outer_int + "." + encap + "," + ipouter + '\n')
                    fcfg.close()
                    inner_bgp_config[n7kname][vrfmember].append("  neighbor " + ipouter + " remote-as " + outer_bgp_as)
                    inner_bgp_config[n7kname][vrfmember].append("   description TO_" + outer_7k)
                    inner_bgp_config[n7kname][vrfmember].append("   address-family ipv4 unicast")
                    inner_bgp_config[n7kname][vrfmember].append("    send-community both")

                    # Write rollback part for inner BGP config and write it later
                    if bool(re.search('inner', n7kname, re.IGNORECASE)):
                        bgp_rb_inner[n7kname]['neighbors'].append(
                            "   no neighbor " + ipouter + " remote-as " + outer_bgp_as)

                    outer_bgp_config[outer_7k][vrfmember].append("  neighbor " + ipinner + " remote-as " + inner_bgp_as)
                    outer_bgp_config[outer_7k][vrfmember].append("   description TO_" + n7kname)
                    outer_bgp_config[outer_7k][vrfmember].append("   address-family ipv4 unicast")
                    outer_bgp_config[outer_7k][vrfmember].append("    send-community both")
                    outer_bgp_config[outer_7k][vrfmember].append("    route-map PERMIT_DEFAULT_ONLY out")
                    outer_bgp_config[outer_7k][vrfmember].append("    default-originate")

                    # Write rollback part for outer BGP config and write it later
                    if bool(re.search('outer', outer_7k, re.IGNORECASE)):
                        bgp_rb_outer[outer_7k]['neighbors'].append(
                            "   no neighbor " + ipinner + " remote-as " + inner_bgp_as)

                    f.write(" ip address " + ipinner + "/" + mask + '\n')
                    f.write('\n')
                    k = k + 1
                    f.close()

                    # Write outer config
                    if not os.path.exists(dir_path + "/" + "N7K_PREWORK" + "/" + outer_7k + "_" + vrfmember):
                        f = open(dir_path + "/" + "N7K_PREWORK" + "/" + outer_7k + "_" + vrfmember, "a")
                        f.write("configure terminal" + '\n')
                        f.close()

                        f = open(dir_path + "/" + "N7K_PREWORK" + "/" + "execute_prework.sh", "a")
                        f.write(
                            "../push_to_n7k.py -f " + outer_7k + "_" + vrfmember + " -c ../" + outer_7k + "_creds" + '\n')
                        f.write("echo FINISHED UPDATING VRF " + vrfmember + " ON " + outer_7k + '\n\n')
                        f.close()

                    f = open(dir_path + "/" + "N7K_PREWORK" + "/" + outer_7k + "_" + vrfmember, "a")
                    f.write(
                        "! Create sub interfaces to inner VDCs for VRF " + vrfmember + " in a shutdown state" + '\n')
                    f.write("interface Ethernet" + outer_int + "." + encap + '\n')
                    f.write(" description To_" + n7kname + "_E" + inner_int + "." + encap + '\n')
                    f.write(" shutdown" + '\n')
                    f.write(" mtu 9192" + '\n')
                    f.write(" encapsulation dot1Q " + encap + '\n')
                    f.write(" ip address " + ipouter + "/" + mask + '\n')
                    f.write('\n')
                    f.close()

                    # Write cutover part to enable the new sub interfaces on outer
                    subint_outer_cutover[outer_7k].append("interface Ethernet" + outer_int + "." + encap)
                    subint_outer_cutover[outer_7k].append(" no shutdown")

                    # if not os.path.exists(cutover_dir + "/" + outer_7k):
                    #	f = open(cutover_dir + "/" +  outer_7k, "a")
                    #	f.write("configure terminal" +  '\n')
                    #	f.close()
                    # f = open(cutover_dir + "/" +  outer_7k, "a")
                    # f.write("! Enable sub interfaces to inner VDCs for VRF " + vrfmember +  '\n')
                    # f.write("interface Ethernet" + outer_int + "." + encap + '\n')
                    # f.write(" no shutdown" + '\n')
                    # f.write('\n')
                    # f.close()

                    # Rollback new sub interfaces on outer
                    if bool(re.search('outer', outer_7k, re.IGNORECASE)):
                        bgp_rb_outer[outer_7k]['subint'].append("no interface Ethernet" + outer_int + "." + encap)

        # Shutdown existing SVI on inner
        # fsvi = open(cutover_dir + "/" +  n7kname, "a")
        svi_inner_cutover[n7kname] = []
        svi_inner_cutover[n7kname].append("interface Vlan" + encap)
        svi_inner_cutover[n7kname].append(" shutdown")

        # fsvi.write("! Shutdown inside SVI vrf " + vrfmember + " for firewall cluster" + '\n')
        # fsvi.write("interface Vlan" + encap + '\n')
        # fsvi.write(" shutdown" + '\n')
        # fsvi.write('\n')
        # fsvi.close()

        # Write rollback part
        if bool(re.search('inner', n7kname, re.IGNORECASE)):
            bgp_rb_inner[n7kname]['svi'].append("interface Vlan" + encap)
            bgp_rb_inner[n7kname]['svi'].append(" no shutdown")

        # Write cleanup part for next window
        if bool(re.search('inner', n7kname, re.IGNORECASE)):
            if n7kname not in svi_cleanup:
                svi_cleanup[n7kname] = {}
                svi_cleanup[n7kname]['svi'] = []
            svi_cleanup[n7kname]['svi'].append("no interface Vlan" + str(encap))

    for n7ks in inner_bgp_config:
        for vrfmember in inner_bgp_config[n7ks]:
            f = open(dir_path + "/" + "N7K_PREWORK" + "/" + n7ks + "_" + vrfmember, "a")
            f.write('\n')
            f.write(('\n'.join(inner_bgp_config[n7ks][vrfmember])))
            f.write('\n')
            f.write('\n')
            f.close()

    for n7ks in outer_bgp_config:
        for vrfmember in outer_bgp_config[n7ks]:
            f = open(dir_path + "/" + "N7K_PREWORK" + "/" + n7ks + "_" + vrfmember, "a")
            f.write(('\n'.join(outer_bgp_config[n7ks][vrfmember])))
            f.write('\n')
            f.write('\n')
            f.close()

    for n7ks in bgp_shut_outer:
        f = open(cutover_dir + "/" + n7ks, "a")
        f.write("configure terminal" + '\n')
        f.write("! Shutdown the BGP adjacency to the N7K Inner in VRF " + vrfmember + '\n')
        f.write(('\n'.join(bgp_shut_outer[n7ks])))
        f.write('\n')
        f.write('\n')
        f.close()

    for n7ks in subint_outer_cutover:
        f = open(cutover_dir + "/" + n7ks, "a")
        f.write("! Enable sub interfaces to inner VDCs for VRF " + vrfmember + '\n')
        f.write(('\n'.join(subint_outer_cutover[n7ks])))
        f.write('\n')
        f.write('\n')
        f.close()

    for n7ks in svi_inner_cutover:
        f = open(cutover_dir + "/" + n7ks, "a")
        f.write(" configure terminal" + '\n')
        f.write("! Shutdown Inside SVI VRF " + vrfmember + " for firewall-cluster" + '\n')
        f.write(('\n'.join(svi_inner_cutover[n7ks])))
        f.write('\n')
        f.write('\n')
        f.close()

    for n7ks in bgp_shut_inner:
        for vrfs in bgp_shut_inner[n7ks]:
            f = open(cutover_dir + "/" + n7ks, "a")
            f.write(('\n'.join(bgp_shut_inner[n7ks][vrfs])))
            f.write('\n')
            f.write('\n')
            f.close()

    for n7ks in subint_inner_cutover:
        f = open(cutover_dir + "/" + n7ks, "a")
        f.write("! Bring up sub interfaces directly connected to the outer VDCs" + '\n')
        f.write(('\n'.join(subint_inner_cutover[n7ks])))
        f.write('\n')
        f.write('\n')
        f.close()

    for n7ks in bgp_rb_inner:
        f = open(rollback_dir + "/" + n7ks, "a")
        f.write("!!" + '\n')
        f.write("!! N7K VDC - VRF " + vrfmember + '\n')
        f.write("!!" + '\n')
        f.write("configure terminal" + '\n')
        f.write("! Remove new sub interfaces to Outer VDCs" + '\n')
        f.write(('\n'.join(bgp_rb_inner[n7ks]['subint'])))
        f.write('\n')
        f.write('\n')
        f.write("! Re-enable SVI to FW-Inner" + '\n')
        f.write(('\n'.join(bgp_rb_inner[n7ks]['svi'])))
        f.write('\n')
        f.write('\n')
        f.write("! Remove direct BGP adjacencies to Outer VDCs and re-enable BGP Adjacencies to FW-Inner" + '\n')
        f.write(('\n'.join(bgp_rb_inner[n7ks]['neighbors'])))
        f.write('\n')
        f.close()

        f = open(dir_path + "/" + "N7K_ROLLBACK" + "/" + "execute_rollback_" + vrfmember + ".sh", "a")
        f.write("../push_to_n7k.py -f " + vrfmember + "/" + n7ks + " -c ../" + n7ks + "_creds" + '\n')
        f.write("echo FINISHED UPDATING " + n7ks + '\n\n')
        f.close()
        fshowc = open(dir_path + "/" + "N7K_ROLLBACK" + "/" + "execute_show_commands_" + vrfmember + ".sh", "a")
        fshowc.write(
            "../push_to_n7k.py -f " + vrfmember + "/" + n7ks + "_inner_show_commands" + " -c ../" + n7ks + "_creds" + ' > ' + n7ks + "_inner_output" + '\n\n')
        fshowc.close()

    for n7ks in bgp_rb_outer:
        f = open(rollback_dir + "/" + n7ks, "a")
        f.write("!!" + '\n')
        f.write("!! N7K Outer VDC - VRF " + vrfmember + '\n')
        f.write("!!" + '\n')
        f.write("configure terminal" + '\n')
        f.write("! Remove new sub interfaces to Inner VDCs" + '\n')
        f.write(('\n'.join(bgp_rb_outer[n7ks]['subint'])))
        f.write('\n')
        f.write('\n')
        f.write("! Remove direct BGP adjacencies to Inner VDCs and re-enable BGP Adjacencies to FW-Outer" + '\n')
        f.write(('\n'.join(bgp_rb_outer[n7ks]['neighbors'])))
        f.write('\n')
        f.write('\n')
        f.write("! Re-enable SVI to FW-Outer" + '\n')
        f.write(('\n'.join(bgp_rb_outer[n7ks]['svi'])))
        f.write('\n')
        f.close()

        f = open(dir_path + "/" + "N7K_ROLLBACK" + "/" + "execute_rollback_" + vrfmember + ".sh", "a")
        f.write("../push_to_n7k.py -f " + vrfmember + "/" + n7ks + " -c ../" + n7ks + "_creds" + '\n')
        f.write("echo FINISHED UPDATING " + n7ks + '\n\n')
        f.close()
        fshowc = open(dir_path + "/" + "N7K_ROLLBACK" + "/" + "execute_show_commands_" + vrfmember + ".sh", "a")
        fshowc.write(
            "../push_to_n7k.py -f " + vrfmember + "/" + n7ks + "_outer_show_commands" + " -c ../" + n7ks + "_creds" + ' > ' + n7ks + "_outer_output" + '\n\n')
        fshowc.close()

    for n7ks in svi_cleanup:
        f = open(cleanup_dir + "/" + n7ks, "a")
        f.write("!!" + '\n')
        f.write("!! SVI Removal N7K Inner VDC - VRF " + vrfmember + '\n')
        f.write("!!" + '\n')
        f.write("configure terminal" + '\n')
        f.write(('\n'.join(svi_cleanup[n7ks]['svi'])))
        f.write('\n')
        f.close()

        f = open(dir_path + "/" + "N7K_NEXT_CLEANUP" + "/" + "execute_cleanup_" + vrfmember + ".sh", "a")
        f.write("../push_to_n7k.py -f " + vrfmember + "/" + n7ks + " -c ../" + n7ks + "_creds" + '\n')
        f.write("echo FINISHED UPDATING " + n7ks + '\n\n')
        f.close()

    for n7ks in n7k_data:
        if bool(re.search('inner', n7ks, re.IGNORECASE)):
            fwints = n7k_data[n7ks][vrfmember]['fw_trunk_int']
            svi = n7k_data[n7ks][vrfmember]['svi']
            f = open(cleanup_dir + "/" + n7ks, "a")
            f.write('\n')
            f.write("!! Remove VLAN from firewall trunk" + '\n')
            for fw in fwints:
                f.write("interface Ethernet" + fw + '\n')
                f.write(" switchport trunk allowed vlan remove " + svi + '\n')
                f.write('\n')
            f.close()

    for n7ks in bgp_cleanup_inner:
        f = open(cleanup_dir + "/" + n7ks, "a")
        f.write("!!" + '\n')
        f.write("!! Remove BGP adjacency to the outer N7K VDC connected to the FW " + '\n')
        f.write("!!" + '\n')
        f.write("configure terminal" + '\n')
        f.write(('\n'.join(bgp_cleanup_inner[n7ks]['neighbors'])))
        f.write('\n')
        f.close()

    for n7ks in bgp_cleanup_outer:
        f = open(cleanup_dir + "/" + n7ks, "a")
        f.write("!!" + '\n')
        f.write("!! Remove BGP adjacency to the inner N7K VDC connected to the FW " + '\n')
        f.write("!!" + '\n')
        f.write("configure terminal" + '\n')
        f.write(('\n'.join(bgp_cleanup_outer[n7ks]['neighbors'])))
        f.write('\n')
        f.close()

        f = open(dir_path + "/" + "N7K_NEXT_CLEANUP" + "/" + "execute_cleanup_" + vrfmember + ".sh", "a")
        f.write("../push_to_n7k.py -f " + vrfmember + "/" + n7ks + " -c ../" + n7ks + "_creds" + '\n')
        f.write("echo FINISHED UPDATING " + n7ks + '\n\n')
        f.close()

    # Show commands
    # Inner SVI,BGP and P2P Links - cutover and rollback
    for n7k in svi_inner_cutover:
        f = open(cutover_dir + "/" + n7k + "_inner_show_commands", "a")
        frb = open(rollback_dir + "/" + n7k + "_inner_show_commands", "a")
        f.write("! Show commands" + '\n')
        f.write("show ip interface brief vrf " + vrfmember + '\n')
        f.write("show ip bgp summary vrf " + vrfmember + '\n')
        frb.write("! Show commands" + '\n')
        frb.write("show ip interface brief vrf " + vrfmember + '\n')
        frb.write("show ip bgp summary vrf " + vrfmember + '\n')
        f.close()
        frb.close()

    # Outer SVI, BGP and P2P Links show commands
    # bgp_rb_outer has all the IP info we need -  no need to worry about name of variable
    for n7ks in bgp_rb_outer:
        f = open(cutover_dir + "/" + n7ks + "_outer_show_commands", "a")
        frb = open(rollback_dir + "/" + n7ks + "_outer_show_commands", "a")
        f.write("show ip bgp summary" + '\n')
        frb.write("show ip bgp summary" + '\n')
        f.write("show interface status" + '\n')
        frb.write("show interface status" + '\n')
        f.close()
        frb.close()
    """
		for cmds in bgp_rb_outer[n7ks]['neighbors']:
			if bool(re.search('neighbor \d+',cmds,re.IGNORECASE)):
				ip = re.findall( r'[0-9]+(?:\.[0-9]+){3}', cmds )
				f = open(cutover_dir + "/" +  n7ks + "_outer_show_commands", "a")
                		frb = open(rollback_dir + "/" +  n7ks + "_outer_show_commands", "a")
				f.write("show ip bgp summary | inc " + ip[0] + '\n' )
				frb.write("show ip bgp summary | inc " + ip[0] + '\n' )
				f.close()
				frb.close()

		for svis in  bgp_rb_outer[n7ks]['svi']:
			if bool(re.search('interface Vlan',svis,re.IGNORECASE)):
				svi_num = re.findall( r'[0-9]+', svis )
				f = open(cutover_dir + "/" +  n7ks + "_outer_show_commands", "a")
                                frb = open(rollback_dir + "/" +  n7ks + "_outer_show_commands", "a")			
				f.write("show int status | inc Vlan" + svi_num[0] + '\n' )
				frb.write("show int status | inc Vlan" + svi_num[0] + '\n' )
				f.close()
				frb.close()

	for n7ks in subint_outer_cutover:
		for cmds in subint_outer_cutover[n7ks]:
			if bool(re.search('interface ethernet',cmds,re.IGNORECASE)):
				cmds = cmds.replace('interface ','show interface status | inc ')
				cmds = cmds.replace('Ethernet','Eth')
				f = open(cutover_dir + "/" +  n7ks + "_outer_show_commands", "a")
                                frb = open(rollback_dir + "/" +  n7ks + "_outer_show_commands", "a")			
				f.write(cmds + '\n' )
				frb.write(cmds + '\n' )
				f.close()
				frb.close()
	"""


def get_inner_outer_mapping(dc, district):
    dc = dc.lower()
    district = district.lower()

    mapping = {}

    # Hard code GIS/SOE mapping because it does not exist in the port map file

    if district.upper() == 'GIS':
        numn7k = [1, 2, 3, 4]
        outerint = ['5/29', '5/30', '5/31', '5/32']
        innerint = ['2/31', '2/32', '2/33', '2/34']

    if district.upper() == 'SOE':
        numn7k = [1, 2, 3, 4]
        outerint = ['5/25', '5/26', '5/27', '5/28']
        innerint = ['2/15', '2/16', '2/17', '2/18']

    # Inner to Outer Mapping - SOE GIS ONLY

    if district.upper() == 'GIS' or district.upper() == 'SOE':

        for i in numn7k:
            name = dc + 'dcinxc' + str(i) + district + 'inner'
            mapping[name] = {}
            for idx, j in enumerate(innerint):
                mapping[name][j] = {}
                mapping[name][j][outerint[i - 1]] = {}
                mapping[name][j][outerint[i - 1]][dc + 'dcinxc' + str(idx + 1) + 'dciouter'] = {}

        # Outer to inner mapping now
        for i in numn7k:
            name = dc + 'dcinxc' + str(i) + 'dciouter'
            mapping[name] = {}
            for idx, j in enumerate(outerint):
                mapping[name][j] = {}
                mapping[name][j][innerint[i - 1]] = {}
                mapping[name][j][innerint[i - 1]][dc + 'dcinxc' + str(idx + 1) + district + 'inner'] = {}

    pattern = '*' + dc.upper() + '*Port Map*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
            if not bool(re.search('^~', name, re.IGNORECASE)):
                filename = name
                break
    worksheets = []
    wb = openpyxl.load_workbook(filename, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('7706')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for cells in ws.iter_rows(min_row=row_start, min_col=1, max_col=24):
        for vals in cells:
            if vals.value == 'A':
                n7k_local = 1
            if vals.value == 'B':
                n7k_local = 2
            if vals.value == 'C':
                n7k_local = 3
            if vals.value == 'D':
                n7k_local = 4
            if vals.value == 'E':
                n7k_local = 1
            if vals.value == 'F':
                n7k_local = 2

            if bool(re.search("nxc(.*)inner", str(vals.value), re.IGNORECASE)) or bool(
                    re.search("nxc(.*)outer", str(vals.value), re.IGNORECASE)):
                remote_n7k = vals.value

                local_interface = ws[vals.column + str(vals.row + 1)].value

                if local_interface is None:
                    local_interface = ws[vals.column + str(vals.row - 1)].value

                local_interface = local_interface.replace("-", "/")
                xvals = vals.value.split("_E")
                remote_interface = xvals[1]
                remote_n7k = xvals[0]

                # Determine local N7K now
                if bool(re.search("sde", remote_n7k, re.IGNORECASE)) and bool(
                        re.search("inner", remote_n7k, re.IGNORECASE)):
                    local_n7k = dc + 'sde' + 'nxc' + str(n7k_local) + 'sdeouter'
                if bool(re.search("sde", remote_n7k, re.IGNORECASE)) and bool(
                        re.search("outer", remote_n7k, re.IGNORECASE)):
                    local_n7k = dc + 'sde' + 'nxc' + str(n7k_local) + 'sdeinner'

                if bool(re.search("gis", remote_n7k, re.IGNORECASE)) and bool(
                        re.search("inner", remote_n7k, re.IGNORECASE)):
                    # local_n7k = dc + 'gis' + 'nxc' + str(n7k_local) + 'gisouter'
                    local_n7k = dc + 'dcinxc' + str(n7k_local) + 'dciouter'
                if bool(re.search("gis", remote_n7k, re.IGNORECASE)) and bool(
                        re.search("outer", remote_n7k, re.IGNORECASE)):
                    # local_n7k = dc + 'gis' + 'nxc' + str(n7k_local) + 'gisinner'
                    local_n7k = dc + 'dcinxc' + str(n7k_local) + 'gisinner'

                if bool(re.search("soe", remote_n7k, re.IGNORECASE)) and bool(
                        re.search("inner", remote_n7k, re.IGNORECASE)):
                    local_n7k = dc + 'dcinxc' + str(n7k_local) + 'dciouter'
                if bool(re.search("soe", remote_n7k, re.IGNORECASE)) and bool(
                        re.search("outer", remote_n7k, re.IGNORECASE)):
                    # local_n7k = dc + 'soe' + 'nxc' + str(n7k_local) + 'soeinner'
                    local_n7k = local_n7k = dc + 'dcinxc' + str(n7k_local) + 'soeinner'

                if local_n7k not in mapping:
                    mapping[local_n7k] = {}

                if local_interface not in mapping[local_n7k]:
                    mapping[local_n7k][local_interface] = {}

                if remote_interface not in mapping[local_n7k][local_interface]:
                    mapping[local_n7k][local_interface][remote_interface] = {}

                if remote_n7k not in mapping[local_n7k][local_interface][remote_interface]:
                    mapping[local_n7k][local_interface][remote_interface][remote_n7k] = {}

    return mapping


def get_bgp_int_vlan(dc, district, vrfs):
    dc = dc.lower()
    data = {}
    p2p_n7k_mapping = get_inner_outer_mapping(dc, district)

    if district.lower() in ('sde'):
        numfiles = ['1', '2']
    else:
        numfiles = ['1', '2', '3', '4']

    for i in numfiles:
        for j in ['inner', 'outer']:
            if district.upper() == 'SDE':
                filename = dc + district + 'nxc' + str(i) + district.lower() + j + '.log'
                n7k = dc + district + 'nxc' + str(i) + district + j
            else:
                if j == 'inner':
                    filename = dc + 'dcinxc' + str(i) + district.lower() + j + '.log'
                    n7k = dc + 'dcinxc' + str(i) + district.lower() + j
                else:
                    filename = dc + 'dcinxc' + str(i) + 'dciouter.log'
                    n7k = dc + 'dcinxc' + str(i) + 'dciouter'

            # GIS/SOE shared - this does not apply.  Have to seperate names for each district
            # n7k = dc + district + 'nxc' + str(i) + district + j

            # Need to code this
            # n7k_mapping = get_inner_outer_mapping(dc)

            data[n7k] = {}
            parse = CiscoConfParse(filename)

            # GET SVI
            for obj in parse.find_objects("interface Vlan"):
                svi = obj.text
                svi = svi.replace("interface Vlan", "")

                if obj.hash_children != 0:
                    for c in obj.children:

                        if bool((re.search('vrf member', c.text, re.IGNORECASE))):
                            vrf = c.text
                            vrf = vrf.replace("vrf member ", "")
                            vrf = vrf.lstrip()

                        # For outer, set VRF to SVI since outer does not have vrf
                        if bool((re.search('outer', n7k, re.IGNORECASE))):
                            vrf = svi

                        if bool((re.search('ip address', c.text, re.IGNORECASE))):
                            svi_ip = c.text
                            svi_ip = c.text.replace("  ip address ", "")
                            svi_ip = re.sub('/.*', "", svi_ip)

                    data[n7k][vrf] = {}
                    data[n7k]['P2P'] = {}
                    data[n7k][vrf] = {'svi': svi, 'shutdown': 'N', 'fw_trunk_int': [], 'remote_as': 'N/A',
                                      'neighbors': [], 'svi_ip': svi_ip, 'local_as': 'N/A'}

            # GET FW INT
            vlallowed = []
            for obj in parse.find_objects("interface Ethernet"):
                if obj.hash_children != 0:
                    if obj.re_search_children("switchport trunk allowed"):
                        for c in obj.children:
                            if bool((re.search('switchport trunk allowed', c.text, re.IGNORECASE))):
                                fwint = obj.text
                                fwint = fwint.replace("interface Ethernet", "")
                                c.text = c.text.replace("switchport trunk allowed vlan", "")
                                vlist = c.text.split(",")
                                for v in vlist:
                                    if bool((re.search('-', v, re.IGNORECASE))):
                                        vv = v.split("-")
                                        vv[1] = int(vv[1]) + int(1)
                                        for vvv in xrange(int(vv[0]), int(vv[1])):
                                            vlallowed.append(vvv)
                                    else:
                                        vlallowed.append(int(v))

                                for vls in vlallowed:
                                    for vrf_vals in data[n7k]:
                                        if vrf_vals != 'P2P':
                                            if vls == int(data[n7k][vrf_vals]['svi']):
                                                if len(data[n7k][vrf_vals]['fw_trunk_int']) == 0:
                                                    data[n7k][vrf_vals]['fw_trunk_int'] = [fwint]
                                                else:
                                                    data[n7k][vrf_vals]['fw_trunk_int'].append(fwint)
                                    vlallowed = []

            # Get BGP Neighbors
            for obj in parse.find_objects("router bgp"):
                local_as = obj.text
                if bool((re.search('asn', local_as, re.IGNORECASE))):
                    continue
                local_as = local_as.replace("router bgp ", "")

                if obj.hash_children != 0:
                    # For inner VDC
                    if obj.re_search_children("vrf"):
                        for c in obj.children:
                            if bool((re.search('vrf', c.text, re.IGNORECASE))):
                                vrfmember = c.text
                                for obj2 in parse.find_objects(vrfmember):
                                    if obj2.re_search_children("neighbor"):
                                        for d in obj2.children:
                                            vrfmember = vrfmember.replace("vrf ", "")
                                            vrfmember = vrfmember.lstrip()
                                            if bool((re.search('address-family', d.text, re.IGNORECASE))):
                                                continue
                                            if bool((re.search('NonAff', vrfmember, re.IGNORECASE))):
                                                continue
                                            if bool((re.search('timers bgp', d.text, re.IGNORECASE))):
                                                continue
                                            attribs = d.text
                                            attribs = attribs.lstrip()
                                            attribsx = attribs.split(" ")
                                            neighbor = attribsx[1]
                                            remote_as = attribsx[3]
                                            data[n7k][vrfmember]['remote_as'] = remote_as
                                            if len(data[n7k][vrfmember]['neighbors']) == 0:
                                                data[n7k][vrfmember]['neighbors'] = [neighbor]
                                            else:
                                                data[n7k][vrfmember]['neighbors'].append(neighbor)

                                            data[n7k][vrfmember]['local_as'] = local_as
                    # For outer VDC
                    if bool((re.search('outer', filename, re.IGNORECASE))):

                        for c in obj.children:
                            if bool((re.search('neighbor ', c.text, re.IGNORECASE))):
                                attribs = c.text
                                attribs = attribs.lstrip()
                                attribsx = attribs.split(" ")
                                neighbor = attribsx[1]
                                remote_as = attribsx[3]

                                for vrfmember in data:
                                    for v in data[vrfmember]:
                                        if v == 'P2P':
                                            continue
                                        if data[vrfmember][v]['svi_ip'] == neighbor:
                                            for n in data[vrfmember][v]['neighbors']:
                                                for o in data:
                                                    for p in data[o]:
                                                        if unicode(p).isnumeric():

                                                            if data[o][p]['svi_ip'] == n:
                                                                if 'vrf' not in data[o][p]:
                                                                    data[o][p]['vrf'] = {}
                                                                    del data[o][p]['neighbors']
                                                                if v not in data[o][p]['vrf']:
                                                                    data[o][p]['vrf'][v] = {}
                                                                    data[o][p]['vrf'][v]['neighbor_ip'] = [neighbor]
                                                                else:
                                                                    if neighbor not in data[o][p]['vrf'][v][
                                                                        'neighbor_ip']:
                                                                        data[o][p]['vrf'][v]['neighbor_ip'].append(
                                                                            neighbor)
                                                                data[o][p]['remote_as'] = remote_as
                                                                data[o][p]['local_as'] = local_as

    # Append P2P data
    # P2P dictionary format:
    # Local Int -> remote int -> remote N7K
    for n7k in data:
        data[n7k]['P2P'] = p2p_n7k_mapping[n7k]

    return data


def get_sw_prof_name(dafe_file, leafid):
    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('leaf_switch_profile')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start + 1, row_end + 1):
        cell = 'E' + str(x)
        id = ws[cell].value

        if int(id) == int(leafid):
            cell = 'A' + str(x)
            prof = ws[cell].value
            break

    try:
        prof
    except NameError:
        print "Could not find switch profile for leaf ID " + str(leafid)
        prof = 'N/A'

    return prof


def fix_type_x(write_to_aci_cfg, dc, district):
    vrf_type = {}
    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                for e in write_to_aci_cfg[tenant][vrf][epg]:
                    if e['t_type'] != 'X':
                        tenant_vrf_type_key = tenant + "-" + vrf
                        vrf_type[tenant_vrf_type_key] = e['t_type']
                        continue
    
    #write_to_aci_cfg['Control']['PA0'] = {}
    #	write_to_aci_cfg['Control']['PA0']['CTL-PA0-DC1-SOE-TEST'] = []
    #	write_to_aci_cfg['Control']['PA0']['CTL-PA0-DC1-SOE-TEST'].append({'t_type' : 'A' })
    # One off - theres one EPG in CTL-PA0-DC1 and can't determine if its type A or B - will assume 'A'
    #if dc.upper() == 'DC1' and district.upper() == 'SOE' and 'Control' in write_to_aci_cfg:
    #    #write_to_aci_cfg['Control']['PA0']['CTL-PA0-DC1-SOE-TEST'][0]['t_type'] = 'A'
    #  	write_to_aci_cfg['Control']['PA0'] = {}
    # 	write_to_aci_cfg['Control']['PA0']['CTL-PA0-DC1-SOE-TEST'] = []
    # 	write_to_aci_cfg['Control']['PA0']['CTL-PA0-DC1-SOE-TEST'].append({'t_type' : 'A' })

    if dc.upper() == 'DC1' and district.upper() == 'SOE' and 'Control' in write_to_aci_cfg:
    	vrf_type['Control-PA0'] = 'A' 

    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                for e in write_to_aci_cfg[tenant][vrf][epg]:
                    if e['t_type'] == 'X':
                        tenant_vrf_type_key = tenant + "-" + vrf
                        e.update({'t_type': vrf_type[tenant_vrf_type_key]})


def usage():
    print "Usage: " + sys.argv[
        0] + " -d|--district <soe, gis or sde> -c|--datacenter <dc1 or dc2> -f|file <inputfile> -x <excludefile>"
    print ""
    print "-f|--file:   Pass input file to use for configuration.   Format:"
    print "1 value per line - assume its an EPG"
    print "6 to 18 values per line (depending on district) - assume by VRF: Tenant,VRF,P2P Subnet IP 1, P2P IP Subnet 2, etc"
    print "4 subnets for SDE, 16 subnets for GIS/SOE"
    print "-d|--district: indicates district name, must be <SOE|GIS|SDE>"
    print "-c|--datacenter: indicates datacenter name, must be <DC1 or DC2>"
    print "-x|--exclude: exclude EPGs from the migration.  File format expected is: tenant,vrf,epg (per line)"
    print "-h|--help: print help message"
    sys.exit(1)


def getValueWithMergeLookup(sheet, cell):
    for m_range in sheet.merged_cell_ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(m_range))
        for row in merged_cells:
            if cell in row:
                return sheet.cell(merged_cells[0][0]).value


def check_selector_exists(selector, pgname, dafe_file):
    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('leaf_interface_selector')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start + 1, row_end + 1):
        cell = 'A' + str(x)
        int_selector = ws[cell].value

        cell = 'D' + str(x)
        pg_path = ws[cell].value

        if int_selector == selector and pg_path == 'uni/infra/funcprof/accbundle-' + pgname:
            return True

    return False


def check_pcpg_name(pgname, dafe_file):
    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('pc_vpc_interface_policy_group')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start + 1, row_end + 1):
        cell = 'A' + str(x)
        name = ws[cell].value

        if name == pgname:
            return True

    return False


def get_pc_params(dafe_file):
    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('pc_vpc_interface_policy_group')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start + 1, row_end + 1):
        cell = 'A' + str(x)
        name = ws[cell].value

        if bool((re.search('^pc_', name, re.IGNORECASE))) and bool((re.search('downlink', name, re.IGNORECASE))):
            cell = 'E' + str(x)
            aepname = ws[cell].value

            aepname = aepname.replace("uni/infra/attentp-", "")

            cell = 'F' + str(x)
            linkpolname = ws[cell].value

            cell = 'G' + str(x)
            cdppolname = ws[cell].value

            cell = 'H' + str(x)
            lldpolname = ws[cell].value

            cell = 'I' + str(x)
            stpolname = ws[cell].value

            cell = 'J' + str(x)
            lacpolname = ws[cell].value

            cell = 'M' + str(x)
            mcpolname = ws[cell].value

            try:
                aepname
            except NameError:
                aepname = 'N/A'
                print "ERROR: Could not get AEP name for PC policy group creation"

            try:
                linkpolname
            except NameError:
                linkpolname = 'N/A'
                print "ERROR: Could not get link policy name for PC policy group creation"

            try:
                cdppolname
            except NameError:
                cdppolname = 'N/A'
                print "ERROR: Could not get CDP policy name for PC policy group creation"

            try:
                lldpolname
            except NameError:
                lldpolname = 'N/A'
                print "ERROR: Could not get LLDP policy name for PC policy group creation"

            try:
                stpolname
            except NameError:
                stpolname = 'N/A'
                print "ERROR: Could not get STP policy name for PC policy group creation"

            try:
                lacpolname
            except NameError:
                lacpolname = 'N/A'
            except NameError:
                print "ERROR: Could not get LACP policy name for PC policy group creation"

            try:
                mcpolname
            except NameError:
                mcpolname = 'N/A'
                print "ERROR: Could not get MCP policy name for PC policy group creation"

            break

    return (aepname, linkpolname, cdppolname, lldpolname, stpolname, lacpolname, mcpolname)


def get_epg_from_vrf(dafe_file, vrfs):
    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)
    epgs = []
    bds = []
    found = 0
    bd_vrf_map = {}

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('bridge_domain')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for v in vrfs:
        vs = v.split(",")
        if len(vs) != 6 and len(vs) != 18:
            print "ERROR: Check input file, must be in format <tenant,vrf,p2p ips subnet1, p2p ip subnet 2, etc. Found %s" % v
            sys.exit(9)
        tenant = vs[0]
        vrf = vs[1]
        vrf = vrf.rstrip()

        for x in range(row_start, row_end + 1):
            cell = 'C' + str(x)
            tenantvalue = ws[cell].value

            cell = 'D' + str(x)
            vrfvalue = ws[cell].value

            if tenantvalue == tenant and vrfvalue == vrf:
                cell = 'A' + str(x)
                bdvalue = ws[cell].value
                if bdvalue.endswith("-FW") is False and bdvalue.endswith("-VMFW") is False:
                    bds.append(bdvalue)
		    bd_vrf_map[bdvalue] = vrf
                    found = 1

        if found == 0:
            print "ERROR: Could not find any BD for Tenant: %s, VRF: %s" % (tenant, vrf)
        # sys.exit(9)
        else:
            found = 0

    wb.active = worksheets.index('epg')

    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row
    vrf_no_epg = []

    for b in bds:
        for x in range(row_start, row_end + 1):
            cell = 'E' + str(x)
            bdvalue = ws[cell].value

            if bdvalue == b:
                epgs.append(b)
                found = 1

        if found == 0:
            print "WARNING: Could not find any EPGs for BD: %s. All L3Outs for VRF %s will be targeted" % (b,bd_vrf_map[b])
	    vrf_no_epg.append(bd_vrf_map[b])
            # sys.exit(9)
        else:
            found = 0

    return (epgs,vrf_no_epg)


def get_vrf_to_fw(zones_vl_ip_file, dc, district):
    worksheets = []
    wb = openpyxl.load_workbook(zones_vl_ip_file, data_only=True)
    vrf_to_fw = {}
    fw_to_vrf = {}

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('SOE_SDE_GIS_VRF_RT_Definition')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start, row_end + 1):
        cell = 'A' + str(x)
        districtvalue = getValueWithMergeLookup(ws, cell)
        if districtvalue == district.upper():
            cell = 'B' + str(x)
            tenant = getValueWithMergeLookup(ws, cell)
            if tenant is None:
                tenant = ws[cell].value
            if tenant == 'Tn_NonAff':
                continue
            cell = 'C' + str(x)
            tenant = tenant.rstrip()
            vrf = ws[cell].value
            if vrf == 'N/A':
                continue
            if vrf is not None:

                cell = 'E' + str(x)
                firewall = ws[cell].value

                cell = 'I' + str(x)
                vrfmember = ws[cell].value

                # For DMZ Cells, get the last number of the vrfmember.  If DC2, add 2
                if bool(re.search('#', vrf, re.IGNORECASE)) and district.upper() == 'SOE' and dc.upper() == 'DC1':
                    vrf = vrf.replace('<Cell #>', vrfmember[-1:])
                if bool(re.search('#', vrf, re.IGNORECASE)) and district.upper() == 'SOE' and dc.upper() == 'DC2':
                    plus2 = int(vrfmember[-1:]) + 2
                    vrf = vrf.replace('<Cell #>', str(plus2))

                cell = 'P' + str(x)
                outerencap = getValueWithMergeLookup(ws, cell)

                if outerencap is None:
                    outerencap = ws[cell].value

                if vrfmember == 'N/A':
                    continue

                if dc == 'dc2':
                    firewall = firewall.replace('dc1', 'dc2')

                # One off - ACI config has Audit/DDT -Zones vlans and IPs has Audit/DAT.  Changing to what ACI has
                if tenant == 'Audit' and vrf == 'DAT' and district.upper() == 'SDE':
                    vrf = 'DDT'

                # One off - Zones, VLANs and IPs has Limited/Build. ACI has Limited/BLD.  Changing to what ACI has
                if tenant == 'Limited' and vrf == 'Build' and district.upper() == 'SOE':
                    vrf = 'BLD'

                # One off - Zones, VLANs and IPs has Services/TFR. ACI has Services/TRF.  Changing to what ACI has
                if tenant == 'Services' and vrf == 'TFR':
                    vrf = 'TRF'

                # Another One off - ACI config has Audit/DDT -Zones vlans and IPs has User Access.  Changing to what ACI has
                if tenant == 'User Access':
                    tenant = 'User_Access'

                if tenant in vrf_to_fw:
                    vrf_to_fw[tenant][vrf] = {}
                    vrf_to_fw[tenant][vrf] = {'firewall': firewall, 'encap': outerencap, 'to_delete': 0}
                else:
                    vrf_to_fw[tenant] = {}
                    vrf_to_fw[tenant][vrf] = {}
                    vrf_to_fw[tenant][vrf] = {'firewall': firewall, 'encap': outerencap, 'to_delete': 0}

            # print "District %s, Tenant %s, VRF %s, Firewall: %s" % (district,tenant,vrf, firewall)

    for tenant in vrf_to_fw:
        for vrf in vrf_to_fw[tenant]:
            firewall = vrf_to_fw[tenant][vrf]['firewall']
            outerencap = vrf_to_fw[tenant][vrf]['encap']
            if firewall in fw_to_vrf:
                fw_to_vrf[firewall].append({'tenant': tenant, 'vrf': vrf, 'to_delete': 0, 'encap': outerencap})
            else:
                fw_to_vrf[firewall] = {}
                fw_to_vrf[firewall] = [{'tenant': tenant, 'vrf': vrf, 'to_delete': 0, 'encap': outerencap}]

    return vrf_to_fw, fw_to_vrf


def get_all_epg_from_dafe(tenant, vrf, file):
    worksheets = []
    bd = []
    epgl = []
    wb = openpyxl.load_workbook(file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('bridge_domain')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start, row_end + 1):
        cell = 'C' + str(x)
        tenantvalue = ws[cell].value

        cell = 'D' + str(x)
        vrfvalue = ws[cell].value

        cell = 'K' + str(x)
        l3routingvalue = ws[cell].value

        # if tenantvalue == tenant and vrfvalue == vrf and l3routingvalue == 'yes':
        if tenantvalue == tenant and vrfvalue == vrf:
            cell = 'A' + str(x)
            bdvalue = ws[cell].value
            bd.append(bdvalue)

    wb.active = worksheets.index('epg')

    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start, row_end + 1):
        cell = 'E' + str(x)
        bdvalue = ws[cell].value

        if bdvalue in bd:
            cell = 'A' + str(x)
            epgvalue = ws[cell].value
            if epgvalue.endswith("-VMFW") is False:
                epgl.append(epgvalue)
    return epgl


def get_epg_type(epg, district):
    worksheets = []

    pattern = 'VRF_EPG_Counts*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
            if not bool(re.search('^~', name, re.IGNORECASE)):
                filename = name
                break

    wb = openpyxl.load_workbook(filename, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    # if bool((re.search('sde',epg,re.IGNORECASE))):
    #            tab = 'SDE'
    # if bool((re.search('gis',epg,re.IGNORECASE))):
    #            tab = 'GIS'
    # if bool((re.search('soe',epg,re.IGNORECASE))):
    #            tab = 'SOE'

    # Passed in district  - no need to determine from EPG name
    tab = district.upper()

    wb.active = worksheets.index(tab)
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start, row_end + 1):
        cell = 'C' + str(x)
        value = ws[cell].value
        if value == epg:
            cell = 'D' + str(x)
            t_type = ws[cell].value
            continue

    for x in range(row_start, row_end + 1):
        cell = 'I' + str(x)
        value = ws[cell].value
        if value == epg:
            cell = 'J' + str(x)
            t_type = ws[cell].value
            continue

    wb.close()
    try:
        t_type
    except NameError:
        # print "WARNING: EPG %s not found in %s" % (epg,filename)
        t_type = 'X'

    return t_type


def get_data(filename, epgs, dc, district, p2psubnets):
    p2psubnetvals = {}
    write_back_contract = {}
    all_l3_contracts = {}

    for p in p2psubnets:
        p = p.rstrip()
        p = p.split(",")
        p_tenant = p[0]
        p_vrf = p[1]
        p_subnets = p[2:]

        # Sort IP addresses in case they are not in order (low to high)
        new_list = []

        # Check for Valid IP address
        for element in p_subnets:
            try:
                IPNetwork(element)
            except AddrFormatError:
                print "Invalid Subnet address given for Tenant %s, VRF %s.  %s given" % (p_tenant, p_vrf, element)
                sys.exit(9)

            t_ip = IPNetwork(element)
            tt_ip = str(t_ip)
            tt_ip = re.sub('/.*', "", tt_ip)

            if str(tt_ip) != str(t_ip.network):
                print "ERROR: Invalid Subnet address provided for Tenant %s, VRF %s.  %s given" % (
                p_tenant, p_vrf, element)
                sys.exit(9)
            else:
                new_list.append(IPNetwork(element))

        new_list.sort()
        # cidr = cidr_merge(new_list)
        # spanning_cidr gives the accurate summary prefix
        cidr = spanning_cidr(new_list)
        # Check for discontinuous networks
        # if len(cidr) != 1:

        # For GIS/SOE, the summary subnet of the 16 networks is /25.  For SDE it is /28

        if (cidr.__str__()[-2:] != '25' and cidr.__str__()[-2:] != '26') and district.upper() in ['GIS', 'SOE']:
            print "WARNING: Discontiguous subnets found for Tenant %s, VRF %s.  %s given" % (
            p_tenant, p_vrf, (', '.join(p_subnets)))

        if cidr.__str__()[-2:] != '28' and district.upper() == 'SDE':
            print "WARNING: Discontiguous subnets found for Tenant %s, VRF %s.  %s given" % (
            p_tenant, p_vrf, (', '.join(p_subnets)))
        p = []
        for ee in new_list:
            if ee.prefixlen != 30:
                print "WARNING: P2P IP for Tenant %s, VRF %s is not a /30.  %s given" % (
                p_tenant, p_vrf, ee.cidr.__str__())

            p.append(ee.cidr.__str__())

        p2psubnetvals[p_tenant + "-" + p_vrf] = {}
        p2psubnetvals[p_tenant + "-" + p_vrf] = p

    # Get Pre-Build data

    worksheets = []
    pre_build = {}

    pattern = 'DCT_' + district.upper() + '*Firewalls*Cabling*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
            if not bool(re.search('^~', name, re.IGNORECASE)):
                cabling_file = name
                break
    try:
        cabling_file
    except NameError:
        print "ERROR: File DCT_" + district.upper() + "_PBR_Firewalls_Cabling&P2P_Info.xlsx not Found"
        sys.exit(9)

    wb2 = openpyxl.load_workbook(cabling_file, data_only=True)

    for sheet in wb2:
        worksheets.append(sheet.title)
    wb2.close()

    wb2.active = worksheets.index('P2P')

    ws2 = wb2.active
    row_start = ws2.min_row
    row_end = ws2.max_row

    for x in range(row_start + 1, row_end + 1):
        cell = 'A' + str(x)
        fwvalue = ws2[cell].value

        if fwvalue is not None:
            # fw = fwvalue[:-1]
            if fwvalue[-1:] == 'a':
                fwa = fwvalue
            else:
                fwb = fwvalue
        # fwa = fw + 'a'
        # fwb = fw + 'b'

        cell = 'G' + str(x)
        vrfmember = ws2[cell].value

        if vrfmember is not None:
            cell = 'B' + str(x)
            vlan = ws2[cell].value

	    # Work around for UAC-ENT-VM Security zone
            if vlan is None:
               vlan = unicode("ae1.3327")

            z = vlan.split('.')
            vlan = z[1]

            cell = 'C' + str(x)
            fwbdsubnet = ws2[cell].value

            cell = 'D' + str(x)
            fwvip = ws2[cell].value

            cell = 'K' + str(x)
            bdip = ws2[cell].value

            if vlan is None:
                print "WARNING: No VLAN Found for Security Zone %s" % vrfmember
            if not vlan.isnumeric():
                print "WARNING: VLAN ID not valid.  %s given." % vlan
            else:
                if int(vlan) > 4096:
                    print "WARNING: Invalid VLAN range given.  %s given." % vlan

            try:
                IP(fwvip)
            except:
                print "WARNING: Firewall IP %s for Security Zone %s not valid" % (fwvip, vrfmember)

            try:
                IP(bdip)
            except:
                print "WARNING: New FW Bridge Domain IP %s for Security Zone %s not valid" % (bdip, vrfmember)

            try:
                IP(fwbdsubnet)
            except:
                print "WARNING: New FW Bridge Domain Subnet address %s for Security Zone %s not valid" % (
                fwbdsubnet, vrfmember)

            pre_build[vrfmember] = ({
                'fwa': fwa,
                'fwb': fwb,
                'vlan': vlan,
                'fwvip': fwvip,
                'fwbdsubnet': fwbdsubnet,
                'bdip': bdip,
                'leafa': 'N/A',
                'leafb': 'N/A',
                'leafa_int': [],
                'leafb_int': [],
            })

    worksheets = []

    pattern = 'DCT*FIrewalls_Cisco*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
            if not bool(re.search('^~', name, re.IGNORECASE)):
                fwfile = name
                break

    wb2 = openpyxl.load_workbook(fwfile, data_only=True)

    for sheet in wb2:
        worksheets.append(sheet.title)
    wb2.close()

    wb2.active = worksheets.index(dc.upper())

    ws2 = wb2.active
    row_start = ws2.min_row
    row_end = 21

    for x in range(row_start + 1, row_end + 1):
        intf_range = []
        cell = 'A' + str(x)
        fwvalue = ws2[cell].value
        cell = 'H' + str(x)
        leaf = ws2[cell].value

        cell = 'J' + str(x)
        leaf_int = ws2[cell].value
        i = leaf_int.split(' - ')
        low = i[0]
        high = i[1]
        lowa = low.split("/")
        higha = high.split("/")
        for ii in range(int(lowa[1]), int(higha[1]) + 1):
            intf_range.append(str(lowa[0]) + "/" + str(ii))

        for vrfmember in pre_build:
            if pre_build[vrfmember]['fwa'] == fwvalue:
                pre_build[vrfmember]['leafa'] = leaf
                pre_build[vrfmember]['leafa_int'] = intf_range
            if pre_build[vrfmember]['fwb'] == fwvalue:
                pre_build[vrfmember]['leafb'] = leaf
                pre_build[vrfmember]['leafb_int'] = intf_range

    ########

    worksheets = []
    write_to_aci_cfg = {}

    wb = openpyxl.load_workbook(filename, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    for epg in epgs:
        epg = epg.rstrip()
        # From EPG tab, get bridge domain name
        wb.active = worksheets.index('epg')

        ws = wb.active
        row_start = ws.min_row
        row_end = ws.max_row

        # Get Bridge Domain
        for x in range(row_start, row_end + 1):
            cell = 'A' + str(x)
            value = ws[cell].value
            if value == epg:
                bdcell = 'E' + str(x)
                bd = ws[bdcell].value

                apcell = 'D' + str(x)
                ap = ws[apcell].value
                continue

        try:
            bd
        except NameError:
            print "ERROR: BD Not found for EPG %s.  Please check dafe output" % epg
            sys.exit(9)
            continue

        try:
            ap
        except NameError:
            print "ERROR: AppProfile Not found for EPG %s.  Please check dafe output" % epg
            sys.exit(9)
            continue

        # From bridge domain tab, get tenant, vrf and L2/L3 Y or N
        wb.active = worksheets.index('bridge_domain')
        ws = wb.active

        row_start = ws.min_row
        row_end = ws.max_row

        for x in range(row_start, row_end + 1):
            cell = 'A' + str(x)
            value = ws[cell].value
            if value == bd:
                tenant = ws['C' + str(x)].value
                vrf = ws['D' + str(x)].value
                l3routing = ws['K' + str(x)].value
                l2_unknown_unicast = ws['F' + str(x)].value
                arp_flood = ws['J' + str(x)].value
                continue

        try:
            tenant
        except NameError:
            print "ERROR: tenant Not found for EPG %s.  Check dafe file" % epg
            sys.exit(9)

        try:
            vrf
        except NameError:
            print "ERROR: vrf Not found for EPG %s.  Check dafe file" % epg
            sys.exit(9)

        try:
            l3routing
        except NameError:
            print "ERROR: l3routing Not found for EPG %s.  Check dafe file" % epg
            sys.exit(9)
        
	try:
            l2_unknown_unicast
        except NameError:
            print "ERROR: l2_unknown_unicast Not found for EPG %s.  Check dafe file" % epg
            sys.exit(9)
	
	try:
            arp_flood
        except NameError:
            print "ERROR: arp_flood Not found for EPG %s.  Check dafe file" % epg
            sys.exit(9)

        # Get BD Subnet
        if l3routing == 'yes':
            wb.active = worksheets.index('bd_subnet')

            ws = wb.active
            row_start = ws.min_row
            row_end = ws.max_row

            for x in range(row_start, row_end + 1):
                cell = 'C' + str(x)
                bdvalue = ws[cell].value

                cell = 'D' + str(x)
                tenantvalue = ws[cell].value

                if bdvalue == bd and tenantvalue == tenant:
                    bdsubnetcell = 'A' + str(x)
                    bd_subnet = ws[bdsubnetcell].value

            try:
                bd_subnet
            except NameError:
                if not bool(re.search('DCX', epg, re.IGNORECASE)):
                    print "WARNING: %s, BD Subnet Not found - unicast routing enabled" % epg
                bd_subnet = 'N/A'
        else:
            bd_subnet = 'N/A'

        # From vrf tab, get vrf member name, using tenant and vrf from bridge_domain tab
        # Can use this for naming other constructs and getting L3OUT name
        wb.active = worksheets.index('vrf')
        ws = wb.active

        row_start = ws.min_row
        row_end = ws.max_row

        for x in range(row_start, row_end + 1):
            cell = 'A' + str(x)
            vrfvalue = ws[cell].value

            cell = 'B' + str(x)
            tenantvalue = ws[cell].value
            if vrfvalue == vrf and tenantvalue == tenant:
                vrfmember = ws['H' + str(x)].value
                # Services/Common has 2 bgpAddressFamilyContext values.  Picking first one

                if vrfvalue == 'Common' and tenantvalue == 'Services':
                    two_vals = vrfmember.split(",")
                    vrfmember = two_vals[0]
                continue

        try:
            vrfmember
        except NameError:
            print "ERROR: vrf member name Not found for EPG %s.  Check dafe file" % epg
            sys.exit(9)

        # Get L3Out name, using vrf member pre-pended with 'L3Out' - this is a good assumption
        # updated to use tab bd_l3out to get L3out
        wb.active = worksheets.index('bd_l3out')
        ws = wb.active

        row_start = ws.min_row
        row_end = ws.max_row

        for x in range(row_start, row_end + 1):
            cell = 'A' + str(x)
            bdvalue = ws[cell].value

            cell = 'B' + str(x)
            tenantvalue = ws[cell].value

            if bdvalue == bd and tenantvalue == tenant:
                cell = 'C' + str(x)
                l3out = ws[cell].value

        try:
            l3out
        except NameError:
            # print "ERROR: l3out name Not found for EPG %s.  Check dafe file" % epg
            l3out = 'N/A'

        # Get EPG consumed contracts.  For the EPG, put contracts into an array
        wb.active = worksheets.index('epg_consumed_contract')
        ws = wb.active
        epg_c_contracts = []

        row_start = ws.min_row
        row_end = ws.max_row

        for x in range(row_start, row_end + 1):
            cell = 'C' + str(x)
            epgvalue = ws[cell].value

            if epgvalue == epg:
                cell = 'D' + str(x)
                contract = ws[cell].value
                epg_c_contracts.append(contract)

        if len(epg_c_contracts) == 0 and l3routing == 'yes':
            print "WARNING: %s, No consumed contracts assigned" % epg

        # Get EPG provided contracts.  For the EPG, put contracts into an array

        wb.active = worksheets.index('epg_provide_contract')
        ws = wb.active
        epg_p_contracts = []

        row_start = ws.min_row
        row_end = ws.max_row

        for x in range(row_start, row_end + 1):
            cell = 'C' + str(x)
            epgvalue = ws[cell].value

            if epgvalue == epg:
                cell = 'D' + str(x)
                contract = ws[cell].value
                epg_p_contracts.append(contract)

        if len(epg_p_contracts) == 0 and l3routing == 'yes':
            print "WARNING: %s, No provide contracts assigned" % epg

        # External EPG name from L3Out name along with provided and consumed contract, which have to be split by comma
        wb.active = worksheets.index('external_epg')
        ws = wb.active
        l3out_c_contracts = []
        l3out_p_contracts = []

        row_start = ws.min_row
        row_end = ws.max_row

        # Get External EPG Name
        for x in range(row_start, row_end + 1):
            cell = 'B' + str(x)
            l3outvalue = ws[cell].value

            cell = 'A' + str(x)
            tenantvalue = ws[cell].value

            if l3outvalue == l3out and tenantvalue == tenant:
                cell = 'D' + str(x)
                externalepg = ws[cell].value

        try:
            externalepg
        except NameError:
            externalepg = 'N/A'
        # print "ERROR: external EPG name Not found for EPG %s.  Check dafe file" % epg
        # sys.exit(9)

        # Now get contracts on the external EPG
        for x in range(row_start + 1, row_end + 1):
            cell = 'B' + str(x)
            l3outvalue = str(ws[cell].value)
            if l3outvalue == l3out:
                cell = 'G' + str(x)
                ccontracts = str(ws[cell].value)
                l3out_c_contracts = ccontracts.split(",")
                cell = 'I' + str(x)
                pcontracts = str(ws[cell].value)
                l3out_p_contracts = pcontracts.split(",")
		if tenant not in all_l3_contracts:
			all_l3_contracts[tenant] = {}
		if vrf not in all_l3_contracts[tenant]:
			all_l3_contracts[tenant][vrf] = {}
		if l3out not in all_l3_contracts[tenant][vrf]:
			all_l3_contracts[tenant][vrf][l3out] = {}
		if externalepg not in all_l3_contracts[tenant][vrf][l3out]:
			all_l3_contracts[tenant][vrf][l3out][externalepg] = {}
			all_l3_contracts[tenant][vrf][l3out][externalepg]['provider'] = l3out_p_contracts
			all_l3_contracts[tenant][vrf][l3out][externalepg]['consumer'] = l3out_c_contracts
                continue

        # Got all the contract info for the EPG, now identify the ones to remove or are an issue:
        # On EPG, if contract is on EPG prov/cons and in L3out as provider and consumer, list it for removal
        contracts_to_remove = []
	# Corner case - remove DCX contracts that exist on EPG but can't find L3out associated to the DCX EPG
	if len(l3out_c_contracts) == 0 and len(l3out_p_contracts) == 0 and bool(re.search('DCX',epg,re.IGNORECASE)) and ( len(epg_c_contracts) != 0 or len(epg_p_contracts) != 0):
		for xcc in epg_c_contracts:
			contracts_to_remove.append(xcc)
		for xcc in epg_p_contracts:
			if xcc not in contracts_to_remove:
				contracts_to_remove.append(xcc)
		
        for c in epg_c_contracts:
            if c in l3out_c_contracts and c in l3out_p_contracts and c in epg_p_contracts:
                contracts_to_remove.append(c)

            if l3out == 'N/A' and l3routing == 'yes':
                print "WARNING: %s, is part of BD %s that is not assigned to an L3 Out.  Please confirm if this EPG should be included" % (
                epg, bd)

            # On EPG, if on consumer but not provider, print warning
            if c not in epg_p_contracts:
                print "WARNING: %s, Contract %s not assigned to EPG as provider" % (epg, c)

            # If contract is not on L3out as provider and consumer, print warning
            if c not in l3out_p_contracts and c not in l3out_c_contracts and externalepg != 'N/A':
                print "WARNING: %s, Consumed Contract %s not assigned to external EPG %s as consumer or provider" % (
                epg, c, externalepg)
                if tenant not in write_back_contract:
                    write_back_contract[tenant] = {}

                if ap not in write_back_contract[tenant]:
                    write_back_contract[tenant][ap] = {}

                if epg not in write_back_contract[tenant][ap]:
                    write_back_contract[tenant][ap][epg] = []

                if c not in write_back_contract[tenant][ap][epg]:
                    write_back_contract[tenant][ap][epg].append(c)

            # On EPG, if on consumer but only on either provider or consumer of L3out, print warning
            if c not in l3out_c_contracts and c in l3out_p_contracts and externalepg != 'N/A':
                print "WARNING: %s, Contract %s not assigned to external EPG %s as consumer, but assigned as provider" % (
                epg, c, externalepg)
            if c in l3out_c_contracts and c not in l3out_p_contracts:
                print "WARNING: %s, Contract %s assigned to external EPG %s as consumer, but not assigned as provider" % (
                epg, c, externalepg)

        # Repeat checks for EPG provider
        # no need to check if EPG provider contract exists on EPG consumer and l3out as provider/consumer - done already
        for p in epg_p_contracts:

            # On EPG, if on provider but not consumer, print warning
            if p not in epg_c_contracts:
                print "WARNING: %s, Contract %s not assigned as consumer" % (epg, p)

            # If contract is not on L3out as provider and consumer, print warning
            if p not in l3out_p_contracts and p not in l3out_c_contracts and externalepg != 'N/A':
                print "WARNING: %s, Provided Contract %s not assigned to external EPG %s as consumer or provider" % (
                epg, p, externalepg)

            # On EPG, if on provider but only on either provider or consumer of L3out, print warning
            if p not in l3out_c_contracts and p in l3out_p_contracts and externalepg != 'N/A':
                print "WARNING: %s, Contract %s not assigned to external EPG %s as consumer but assigned as provider" % (
                epg, p, externalepg)
            if p in l3out_c_contracts and p not in l3out_p_contracts:
                print "WARNING: %s, Contract %s assigned to external EPG %s as consumer but not assigned as provider" % (
                epg, p, externalepg)

        # Get  VRF to EPG list - will check to make sure we got it all
        # if the argument passed to the script is by VRF, no need to check, we got it all and inner N7K config can be removed

        # Get Type A or Type B.  Look at file VRF_EPG_Counts
        t_type = get_epg_type(epg, district)

        # Write to dictionary for printing config files
        if tenant not in write_to_aci_cfg:
            write_to_aci_cfg[tenant] = {}
        if vrf not in write_to_aci_cfg[tenant]:
            write_to_aci_cfg[tenant][vrf] = {}

        write_to_aci_cfg[tenant][vrf][epg] = {}

        # One off - rename AUD-DDT to AUD-DAT temporarily
        if bool((re.search('AUD-DDT', vrfmember, re.IGNORECASE))) and district.upper() == 'SDE':
            vrfmember = vrfmember.replace('DDT', 'DAT')

        # One off - rename SVC-TRF to SVC-TFR temporarily
        if bool((re.search('SVC-TRF', vrfmember, re.IGNORECASE))) and district.upper() == 'SOE':
            vrfmember = vrfmember.replace('TRF', 'TFR')

        # One off - rename DMZ-DVT-DC[1 or2]-SDE  to DMZ-WEB-DC[1 or 2]-SDE-CELL1 temporarily
        if vrfmember == 'DMZ-DVT-' + dc.upper() + '-SDE':
            vrfmember = 'DMZ-WEB-' + dc.upper() + '-SDE-CELL1'

        # One off - all CELL1 to VRF Name
        if bool((re.search('AUD-ACC-DC', vrfmember, re.IGNORECASE))) and district.upper() == 'SOE':
            vrfmember = vrfmember + "-CELL1"

        # One off - all CELL1 to VRF Name
        if bool((re.search('AUD-DAT-DC', vrfmember, re.IGNORECASE))) and district.upper() == 'SOE':
            vrfmember = vrfmember + "-CELL1"

        if vrfmember in pre_build:
            vlan = pre_build[vrfmember]['vlan']
            fwvip = pre_build[vrfmember]['fwvip']
            fwbdip = pre_build[vrfmember]['bdip']
            fwbdsubnet = pre_build[vrfmember]['fwbdsubnet']
            leafa = pre_build[vrfmember]['leafa']
            leafb = pre_build[vrfmember]['leafb']
            fwa = pre_build[vrfmember]['fwa']
            fwb = pre_build[vrfmember]['fwb']
            leafa_int = pre_build[vrfmember]['leafa_int']
            leafb_int = pre_build[vrfmember]['leafb_int']

        else:
            print "WARNING: Could not find pre build work for VRF Member %s " % vrfmember
            vlan = 'N/A'
            fwvip = 'N/A'
            fwbdip = 'N/A'
            fwbdsubnet = 'N/A'
            leafa = 'N/A'
            leafb = 'N/A'
            fwa = 'N/A'
            fwb = 'N/A'
            leafa_int = 'N/A'
            leafb_int = 'N/A'

        if bool((re.search('AUD-DAT', vrfmember, re.IGNORECASE))) and district.upper() == 'SDE':
            vrfmember = vrfmember.replace('DAT', 'DDT')

        # Change back the One off
        if vrfmember == 'DMZ-WEB-' + dc.upper() + '-SDE-CELL1':
            vrfmember = 'DMZ-DVT-' + dc.upper() + '-SDE'

        # Change back the One off
        if vrfmember == 'AUD-ACC-' + dc.upper() + '-' + district.upper() + '-CELL1' and district.upper() == 'SOE':
            vrfmember = 'AUD-ACC-' + dc.upper() + '-' + district.upper()

        # Change back the One off
        if vrfmember == 'AUD-DAT-' + dc.upper() + '-' + district.upper() + '-CELL1' and district.upper() == 'SOE':
            vrfmember = 'AUD-DAT-' + dc.upper() + '-' + district.upper()

        # Change back one off - rename SVC-TFR to SVC-TRF
        if bool((re.search('SVC-TFR', vrfmember, re.IGNORECASE))) and district.upper() == 'SOE':
            vrfmember = vrfmember.replace('TFR', 'TRF')

        write_to_aci_cfg[tenant][vrf][epg] = [{

            'bd': bd,
            'l3': l3routing,
            'vrfmember': vrfmember,
            'l2_unknown_unicast': l2_unknown_unicast,
            'arp_flood': arp_flood,
            'l3out': l3out,
            'extepg': externalepg,
            't_type': t_type,
            'ap': ap,
            'fwbdsubnet': fwbdsubnet,
            'vrf': vrf,
            'bd_subnet': bd_subnet,
            'remove_l3_contract': 'yes',
            'contract': contracts_to_remove,
            'curr_c_contracts': epg_c_contracts,
            'curr_p_contracts': epg_p_contracts,
            'vlan': vlan,
            'fwbdname': 'N/A',
            'fwvip': fwvip,
            'fwbdip': fwbdip,
            'leafa': leafa,
            'leafb': leafb,
            'leafa_int': leafa_int,
            'leafb_int': leafb_int,
            'fwaname': fwa,
            'fwbname': fwb,
            'l4l7': 'N/A',
            'sgtname': 'N/A',
            'pbrname': 'N/A',
            'sgcontractname': 'N/A',
            'fwcluster': 'N/A',
            'p2psubnets': p2psubnetvals[tenant + "-" + vrf]
        }]

        """	
	print "Tenant: %s" % tenant	
	print "EPG: %s" % epg
	print "BD: %s" % bd
	print "VRF: %s" % vrf
	print "L3 BD: %s" % l3routing
	print "VRF Member: %s" % vrfmember
	print "L3Out: %s" % l3out
	print "External EPG: %s" % externalepg
	print "Type: %s" % t_type
	print "Contracts to remove:"
	print contracts_to_remove
	print "********"
	print 
	print 
	print
        """
        del tenant
        del epg
        del l2_unknown_unicast
        del arp_flood
        del bd
        del vrf
        del l3routing
        del vrfmember
        del l3out
        del epg_c_contracts
        del epg_p_contracts
        del externalepg
        del l3out_c_contracts
        del l3out_p_contracts
        del contracts_to_remove
        del t_type
        del vlan
        del fwvip
        del fwbdip
        del fwbdsubnet
        del leafa
        del leafb
        del fwa
        del fwb
        del leafa_int
        del leafb_int
        del bd_subnet

    return (write_to_aci_cfg, write_back_contract,all_l3_contracts)


def main(argv):
    
    dir_path = './output'
    if os.path.isdir(dir_path):
        shutil.rmtree(dir_path)

    os.mkdir(dir_path)

    toexclude = False

    # Arguments
    if len(argv) == 0:
        usage()

    try:
        opts, args = getopt.getopt(argv, "d:c:f:hx:", ["district=", "datacenter=", "file=", "help", "--exclude"])
    except getopt.GetoptError as err:
        print str(err)
        sys.exit(2)
    else:
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                usage()
            if opt in ("-d", "--district"):
                district = arg
            if opt in ("-c", "--datacenter"):
                dc = arg
            if opt in ("-f", "--file"):
                infile = arg
            if opt in ("-x", "--exclude"):
                excludefile = arg
                toexclude = True
    try:
        district
    except NameError:
        print "District option not passed (-d|--district)"
        sys.exit(9)

    try:
        infile
    except NameError:
        print "Input file not passed (-f|--file)"
        sys.exit(9)

    try:
        dc
    except NameError:
        print "Datacenter option not passed (-c|--datacenter)"
        sys.exit(9)

    if district.upper() not in ('SDE', 'GIS', 'SOE'):
        print "ERROR: District must be SOE, GIS, SDE - one of these values"
        sys.exit(9)

    if dc.upper() not in ('DC1', 'DC2'):
        print "ERROR: DataCenter must be DC1 or DC2 - one of these values"
        sys.exit(9)

    if not os.path.isfile(infile):
        print sys.argv[0] + " Input File %s NOT found" % infile
        sys.exit(9)

    if toexclude is True:
        if os.path.isfile(excludefile):
            with open(excludefile) as ex:
                exinfo = ex.read().splitlines()
        else:
            print "ERROR: Exclude option chosen, but exclude file not found"
            sys.exit(9)

    dafe_file = dc.upper() + "_" + district.upper() + "_DAFE.xlsx"
        
    with open(infile) as f:
        epgs = f.readlines()
        numparams = epgs[0].split(",")

    # If theres no comma in the input, its by EPG. 
    # if theres at least 1 comma, its by VRF
    # Deprecated - it's only bt VRF now

    if len(numparams) == 1:
        with open(infile) as f:
            epgs = f.readlines()

    if len(numparams) == 6 and district.upper() == 'SDE':
        with open(infile) as f:
            vrfs = f.readlines()
            (epgs,vrf_no_epg) = get_epg_from_vrf(dafe_file, vrfs)

    if len(numparams) != 6 and district.upper() == 'SDE':
        print "ERROR: Check input file. There must be at least 6 parameters (tenant, vrf, P2P Subnet 1, P2P Subnet 2, etc"
        sys.exit(9)

    if len(numparams) == 18 and (district.upper() == 'SOE' or district.upper() == 'GIS'):
        with open(infile) as f:
            vrfs = f.readlines()
            (epgs,vrf_no_epg) = get_epg_from_vrf(dafe_file, vrfs)

    if len(numparams) != 18 and (district.upper() == 'SOE' or district.upper() == 'GIS'):
        print "ERROR: Check input file. There must be at least 18 parameters (tenant, vrf, P2P Subnet 1, P2P Subnet 2, etc"
        sys.exit(9)

    # Get N7K Data - to be used after ACI configs built
    n7k_data = get_bgp_int_vlan(dc, district, vrfs)

    # Major one off - CTL/PA0 has no BDs but is in scope.
    for v in vrfs:
        xxz = v.split(',')
        if xxz[0] == 'Control' and xxz[1] == 'PA0' and dc.upper() == 'DC2' and district.upper() == 'SOE':
            p2psubnets = []
            for i in range(2, 18):
                p2psubnets.append(xxz[i].rstrip())
            write_new_n7k_configs('CTL-PA0-DC2-SOE', p2psubnets, dc, district, n7k_data)
    p2psubnets = []

    (aepname, linkpolname, cdppolname, lldpolname, stpolname, lacpolname, mcpolname) = get_pc_params(dafe_file)
    (write_to_aci_cfg, write_back_contract,all_l3_contracts) = get_data(dafe_file, epgs, dc, district, vrfs)

    # Print out pre-migration planning info
    migration_planning_file = './output/PRE_MIGRATION_PLANNING.txt'
    f = open(migration_planning_file, "a")

    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            ecount = 0
            nomigr = []
            for epg in write_to_aci_cfg[tenant][vrf]:
                for e in write_to_aci_cfg[tenant][vrf][epg]:
                    bdip = e['bd_subnet']
                    isl3 = e['l3']
                    fw = e['fwaname']
                    vrftype = e['t_type']
		    arp_flood = e['arp_flood']
		    l2_unknown_unicast = e['l2_unknown_unicast']
                    #if isl3 == 'yes' or ( l2_unknown_unicast == 'proxy' and arp_flood == 'no' ) or bool(re.search('DCX', epg, re.IGNORECASE)):
		    # 10-15-2020: Exclude DCX even if L2
                    if isl3 == 'yes' or ( l2_unknown_unicast == 'proxy' and arp_flood == 'no' ):
                        f.write(epg + ',' + bdip + ',' + fw + ',' + vrftype + ',' + vrf + ',' + tenant + ',' + '\n')
                        ecount = ecount + 1
                    else:
                        nomigr.append(epg)
            f.write('\n' + "Number of EPGs migrated: " + str(ecount) + '\n')
            f.write("Number of L2 EPGs not migrated: " + str(len(nomigr)) + '\n')
            for n in nomigr:
                f.write(n + '\n')
            f.write('\n')
            f.write('*' * 12 + '\n')
    f.close()

    pattern = 'Zones*Vlans*IPs*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
            if not bool(re.search('^~', name, re.IGNORECASE)):
                zvlipfile = name
                break

    vrf_to_fw, fw_to_vrf = get_vrf_to_fw(zvlipfile, dc, district)

    # Checks
    vrfstodelete = []
    # 1. all EPGs migrated in VRF for the input file? Y or N - Will determine if the n7K inner config gets modified
    for tenant in write_to_aci_cfg:
        missing_from_input_file = []
        missing_from_dafe = []
        epglist = []

        for vrf in write_to_aci_cfg[tenant]:

            for e in write_to_aci_cfg[tenant][vrf]:
                epglist.append(e)

            # Get EPG list from DAFE output
            dafe_epg = get_all_epg_from_dafe(tenant, vrf, dafe_file)

            for e in dafe_epg:
                if e not in epglist:
                    missing_from_input_file.append(e)

            for e in epglist:
                if e not in dafe_epg:
                    missing_from_dafe.append(e)

            if len(missing_from_dafe) > 0:
                print "ERROR: The following EPGs provided are not in tenant %s, vrf %s" % (tenant, vrf)
                for e in missing_from_dafe:
                    print e

            if len(missing_from_input_file) > 0:
                f = open(migration_planning_file, "a")
                f.write(
                    "WARNING: The following EPGs in tenant %s, vrf %s are not being migrated.  Inner and outer VDC config should not be modified and contract on L3Out should not be removed" % (
                    tenant, vrf) + '\n')
                for e in missing_from_input_file:
                    f.write(e + '\n')
                f.close()

                # find out which tenant this EPG is in and set all the EPGs in that tenant to remove_l3_contract to no
                for tenant in write_to_aci_cfg:
                    for vrf in write_to_aci_cfg[tenant]:
                        for ep in write_to_aci_cfg[tenant][vrf]:
                            for a in write_to_aci_cfg[tenant][vrf][ep]:
                                a['remove_l3_contract'] = 'no'

            else:
                # All inner is being migrated, remove the tenant/VRF from the list of fw_to_epg for the outer, then check this later on to see if all VRFs are removed from FW
                # Write N7K Config here (Inner and outer and shutdown inner config)
                # Still have to check if we can shutdown outer config - done later
                # Check for cleanup as well - later
                f = open(migration_planning_file, "a")
                f.write("Inner config can be removed for tenant %s, vrf %s" % (tenant, vrf) + '\n')
                f.close()

                for ep in write_to_aci_cfg[tenant][vrf]:
                    vrfmember = write_to_aci_cfg[tenant][vrf][ep][0]['vrfmember']
                    p2psubnets = write_to_aci_cfg[tenant][vrf][ep][0]['p2psubnets']
                    break

                # if district.upper() == 'SOE' and tenant == 'Limited' and vrf == 'BLD':
                #	vrf_to_fw[tenant]['Build']['to_delete'] = 1
                # else:
                vrf_to_fw[tenant][vrf]['to_delete'] = 1
                for n7k in n7k_data:
                    if bool((re.search('inner', n7k, re.IGNORECASE))):
                        n7k_data[n7k][vrfmember]['shutdown'] = 'Y'
                        vrfstodelete.append(vrfmember)
                        continue

                write_new_n7k_configs(vrfmember, p2psubnets, dc, district, n7k_data)

            epglist = []
            dafe_epg = []
            missing_from_dafe = []
            missing_from_input_file = []

    f = open('n7k_data.json', 'w')
    f.write(json.dumps(n7k_data))
    f.close()

    # See what's left in the outer
    # If all the VRFs are being removed from the FW, shutdown the SVI on the outer and remove the VLAN from the trunk on inner and outer
    stillexist = []
    targeted = 0
    count = {}

    cutover_dir = dir_path + "/" + "N7K_CUTOVER"
    cleanup_dir = dir_path + "/" + "N7K_NEXT_CLEANUP"

    for n7k in n7k_data:
        if bool((re.search('outer', n7k, re.IGNORECASE))):
            for svi in n7k_data[n7k]:
                if svi == 'P2P':
                    continue
                if 'vrf' not in n7k_data[n7k][svi]:
                    continue
                for vrfs in n7k_data[n7k][svi]['vrf']:
                    if vrfs not in vrfstodelete:
                        stillexist.append(vrfs)
                    else:
                        targeted = 1

                # Write SVI shutdown and VLAN removal to the VRF cutover file which has the most EPG's in the group of VRFs on the trunk - this will most likely be targeted last during the migration
                # In the inner, remove the VLAN from the FW trunk for every VRF
                # Get EPG count
                for v in n7k_data[n7k][svi]['vrf']:
                    for tenant in write_to_aci_cfg:
                        for vvvrf in write_to_aci_cfg[tenant]:
                            for epg in write_to_aci_cfg[tenant][vvvrf]:
                                if write_to_aci_cfg[tenant][vvvrf][epg][0]['vrfmember'] == v:
                                    num = len(write_to_aci_cfg[tenant][vvvrf])
                                    count[v] = num
                                    break;
                if len(stillexist) == 0 and targeted == 1:
                    sorted_d = sorted(count.items(), key=operator.itemgetter(1))
                    fmigr = open(migration_planning_file, "a")
                    fmigr.write(
                        "Outer encap " + svi + " can be removed on " + n7k + "." + "  SVI shutdown config will be written to: " +
                        sorted_d[-1][0] + " because it has the most EPGs in this VRF/SVI" + '\n')
                    fmigr.close()
                    f = open(cutover_dir + "/" + sorted_d[-1][0] + "/" + n7k, "a")
                    f.close()

                if len(stillexist) != 0 and targeted == 1:
                    sorted_d = sorted(count.items(), key=operator.itemgetter(1))
                    fmigr = open(migration_planning_file, "a")
                    fmigr.write(
                        "WARNING: Outer encap " + svi + " on " + n7k + " has VRFs, but SVI shutdown config will be written to " +
                        sorted_d[-1][
                            0] + ".  Please remove this config if not needed. The following VRFs still exist:" + ','.join(
                            stillexist) + '\n')
                    fmigr.close()
                    f = open(cutover_dir + "/" + sorted_d[-1][0] + "/" + n7k, "a")
                    f.write(
                        "!! Verify if the SVI shutdown and firewall config VLAN removal should be executed. VRFs " + ','.join(
                            stillexist) + " still exist on this VLAN per config" + '\n')
                    f.close()

                if targeted == 1:
                    fwints = n7k_data[n7k][svi]['fw_trunk_int']

                    # write the shutdown for the outer VLAN in the middle of the file
                    foundintvl = 0
                    f = open(cutover_dir + "/" + sorted_d[-1][0] + "/" + n7k, "r")
                    contents = f.readlines()
                    f.close()
                    f = open(cutover_dir + "/" + sorted_d[-1][0] + "/" + n7k, "w")
                    for c in contents:
                        if c == '\n' and foundintvl == 0:
                            f.write(c)
                            f.write("!Shutdown Outer SVI for VRF " + sorted_d[-1][0] + '\n')
                            f.write("interface Vlan" + svi + '\n')
                            f.write(" shutdown" + '\n')
                            f.write(c)
                            foundintvl = 1
                        else:
                            f.write(c)
                    foundintvl = 0

                    # f = open(cutover_dir + "/" + sorted_d[-1][0] + "/" + n7k, "a")
                    # f.write("interface Vlan" + svi + '\n')
                    # f.write(" shutdown" + '\n')
                    # f.close()
                    f = open(cleanup_dir + "/" + sorted_d[-1][0] + "/" + n7k, "a")
                    f.write("!!Remove VLAN Interface for VRF " + sorted_d[-1][0] + '\n')
                    f.write("no interface Vlan" + svi + '\n')
                    f.write('\n')
                    for fw in fwints:
                        f.write("!!Remove VLAN from FW trunk for VRF " + sorted_d[-1][0] + '\n')
                        f.write("interface Ethernet" + fw + '\n')
                        f.write(" switchport trunk allowed vlan remove " + svi + '\n')
                        f.write('\n')
                    f.close()

                count = {}
                stillexist = []
                targeted = 0

    # Write ACI configs
    # Constraints
    # Can't remove contract from L3Out if all EPG's aren't being migrated

    # Fix EPGs/VRFs that have type = 'X'
    fix_type_x(write_to_aci_cfg, dc, district)

    # Associate leaf interfaces to PC and create leaf selectors.  Already done, not needed for 3/13.  Again check previous files

    # Check if output dir exists - if not, create it
    # if exists, delete it and re-create it

    os.mkdir(dir_path + "/" + "ACI_PRE_WORK")
    os.mkdir(dir_path + "/" + "ACI_CONTRACT_VERIFICATION")

    # Create empty creds file - used for later
    open(dir_path + "/" + "ACI_CONTRACT_VERIFICATION" + "/" + "aci_creds", 'a').close()

    # Pre work
    # Create port channel  and interface selectors
    pc_creation = {}
    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                for d in write_to_aci_cfg[tenant][vrf][epg]:
                    pca_key = d['fwaname']
                    if d['fwaname'] == 'N/A':
                        print "WARNING: No firewall information found for EPG: " + epg
                        continue
                    pca_key = 'pc_' + pca_key + ":" + aepname + ":" + linkpolname + ":" + cdppolname + ":" + lldpolname + ":" + stpolname + ":" + lacpolname + ":" + mcpolname

                    pcb_key = d['fwbname']
                    pcb_key = 'pc_' + pcb_key + ":" + aepname + ":" + linkpolname + ":" + cdppolname + ":" + lldpolname + ":" + stpolname + ":" + lacpolname + ":" + mcpolname
                    if pca_key in pc_creation:
                        continue
                    else:
                        pc_creation[pca_key] = {'leafid': d['leafa'], 'intf': d['leafa_int']}

                    if pcb_key in pc_creation:
                        continue
                    else:
                        pc_creation[pcb_key] = {'leafid': d['leafb'], 'intf': d['leafb_int']}

    f = open(
        dir_path + "/" + "ACI_PRE_WORK/3.6  - Create port channel interface policy group for the new firewall connections.csv",
        "a")
    f.write("PC_PG,LINK,CDP,MCP,LLDP,BPDU,LACP,AEP" + '\n')
    f.close()

    f = open(
        dir_path + "/ACI_PRE_WORK/3.7 - Create int selectors and associate leaf interfaces to port channel interface policy group.csv",
        "a")
    f.write("INTPROFILE,PC_POLICY,PORT,PN" + '\n')
    f.close()

    # pc_creation['pc_dc1sdenwa1sbx01:AEP_STATIC:10GB_Auto:CDP_ENABLE:LLDP_ENABLE:BPDU_GUARD_ENABLED:LACP_ACTIVE:MCP_ENABLE'] =  {"leafid": 212, "intf": ["1/27", "1/28", "1/29", "1/17"]}

    for pc_key in pc_creation:
        tmp = pc_key.split(':')
        pgname = tmp[0]
        link = tmp[2]
        cdppol = tmp[3]
        mcp_pol = tmp[7]
        lldppol = tmp[4]
        stppol = tmp[5]
        lacppol = tmp[6]
        aeppol = tmp[1]
        leafid = pc_creation[pc_key]['leafid']
        intf = pc_creation[pc_key]['intf']
        # print json.dumps(pc_creation)
        swprofname = get_sw_prof_name(dafe_file, leafid)

        # if port channel policy group is created, skip creating it
        pgname_exists = check_pcpg_name(pgname, dafe_file)
        if pgname_exists is False:
            f = open(
                dir_path + "/ACI_PRE_WORK/3.6  - Create port channel interface policy group for the new firewall connections.csv",
                "a")
            f.write(
                pgname + "," + link + "," + cdppol + "," + mcp_pol + "," + lldppol + "," + stppol + "," + lacppol + "," + aeppol + '\n')
            f.close()
        else:
	    ff = open("pc_and_int_sel.log","a")
            ff.write("OK: Port channel policy group name " + pgname + " exists, not creating it")
	    ff.close()

        f = open(
            dir_path + "/ACI_PRE_WORK/3.7 - Create int selectors and associate leaf interfaces to port channel interface policy group.csv",
            "a")
        for n in intf:
            i = n.split("/")
            portnum = i[1]
            selector = "E_" + i[0] + "_" + portnum
            selector_exists = check_selector_exists(selector, pgname, dafe_file)

            if selector_exists is False:
                f.write(swprofname + "," + pgname + "," + selector + "," + portnum + '\n')
            else:
		ff = open("pc_and_int_sel.log","a")
                ff.write("OK: Interface selector " + selector + " already exists in policy group " + pgname + ".  Selector will not be created")
		ff.close()
        f.close()

    # 3.8 to 3.14 - Create new construct names, save it and access it later 

    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                for d in write_to_aci_cfg[tenant][vrf][epg]:
                    tshortname = epg.split("-")
                    tenantshort = tshortname[0]
                    shortfirewall = d['fwaname']
                    shortfirewall = shortfirewall[:-1]
                    fwbdname = tenantshort + "-" + vrf + "-BD-" + dc.upper() + "-" + district.upper() + "-FW"
                    if bool((re.search('C-', fwbdname, re.IGNORECASE))):
                        fwbdname = fwbdname.replace("C-", "C_")

                    d.update({'fwbdname': fwbdname})

                    # Create L4L7 name and update dictionary
                    l4l7name = tenantshort + "-" + vrf + "-L4L7-" + shortfirewall.upper()
                    d.update({'l4l7': l4l7name})

                    # Create service graph template name and update dictionary
                    sgtname = tenantshort + "-" + vrf + "-SG-" + dc.upper() + "-" + district.upper()
                    d.update({'sgtname': sgtname})

                    # Create PBR name and update dictionary
                    pbrname = tenantshort + "-" + vrf + "-PBR-" + shortfirewall.upper()
                    d.update({'pbrname': pbrname})

                    # Create contract name and update dictionary
                    sgcontractname = tenantshort + "-" + vrf + "-SG-PBR-Permit_Any"
                    d.update({'sgcontractname': sgcontractname})

                    # update FW cluster name
                    d.update({'fwcluster': shortfirewall.upper()})

    # Prepare all the headers for the pre-work files
    f = open(dir_path + "/ACI_PRE_WORK/3.8 - Create a bridge domain for the policy-based routing policy.csv", "a")
    f.write("TENANT,BD,SUBNET,VRF" + '\n')
    f.close()

    f = open(dir_path + "/ACI_PRE_WORK/3.9 - Create the L4-L7 device.csv", "a")
    f.write("TENANT,FW,DEVICE1,LEAF_ID1,PC_NAME1,DEVICE2,LEAF_ID2,PC_NAME2,PHYS_DOMAIN,VLANID,CLUSTER" + '\n')
    f.close()

    f = open(dir_path + "/ACI_PRE_WORK/3.10 - Create the service graph template.csv", "a")
    f.write("TENANT,SGNAME,FW" + '\n')
    f.close()

    f = open(dir_path + "/ACI_PRE_WORK/3.11 - Create the policy-based redirect policies.csv", "a")
    f.write("TENANT,PBRNAME,REDIRECTIP,REDIRECTMAC" + '\n')
    f.close()

    f = open(
        dir_path + "/ACI_PRE_WORK/3.12 - Create a new contract with the Service Graph name defined as the subject.csv",
        "a")
    f.write("TENANT,CONTRACT_NAME,SUBJECT" + '\n')
    f.close()

    f = open(dir_path + "/ACI_PRE_WORK/3.13 - Create device selection policy.csv", "a")
    f.write("TENANT,CONTRACT_NAME,SGNAME,NODENAME,FW,BD,PBR_POLICY,CLUSTER" + '\n')
    f.close()

    f = open(
        dir_path + "/ACI_PRE_WORK/3.14  - Assign the L4-L7 service graph to the new contract in the contract subject.csv",
        "a")
    f.write("TENANT,CONTRACT_NAME,SGNAME,SUBJECT" + '\n')
    f.close()

    # Write the configs
    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                fwbdname = write_to_aci_cfg[tenant][vrf][epg][0]['fwbdname']
                sgtname = write_to_aci_cfg[tenant][vrf][epg][0]['sgtname']
                fwbdip = write_to_aci_cfg[tenant][vrf][epg][0]['fwbdip']
                fwvip = write_to_aci_cfg[tenant][vrf][epg][0]['fwvip']
                l4l7name = write_to_aci_cfg[tenant][vrf][epg][0]['l4l7']
                leafid_a = write_to_aci_cfg[tenant][vrf][epg][0]['leafa']
                leafid_b = write_to_aci_cfg[tenant][vrf][epg][0]['leafb']
                pbrname = write_to_aci_cfg[tenant][vrf][epg][0]['pbrname']
                sgcontractname = write_to_aci_cfg[tenant][vrf][epg][0]['sgcontractname']
                fwaname = write_to_aci_cfg[tenant][vrf][epg][0]['fwaname']
                fwbname = write_to_aci_cfg[tenant][vrf][epg][0]['fwbname']
                fwcluster = write_to_aci_cfg[tenant][vrf][epg][0]['fwcluster']
                vlan = write_to_aci_cfg[tenant][vrf][epg][0]['vlan']
                fwbdsubnetmask_t = write_to_aci_cfg[tenant][vrf][epg][0]['fwbdsubnet'].split("/")
                fwbdsubnetmask = fwbdsubnetmask_t[1]
                f = open(
                    dir_path + "/ACI_PRE_WORK/3.8 - Create a bridge domain for the policy-based routing policy.csv",
                    "a")
                f.write(tenant + "," + fwbdname + "," + fwbdip + "/" + fwbdsubnetmask + "," + vrf + '\n')
                f.close()

                f = open(dir_path + "/ACI_PRE_WORK/3.9 - Create the L4-L7 device.csv", "a")
                f.write(tenant + "," + l4l7name + "," + fwaname.upper() + "," + str(
                    leafid_a) + "," + "pc_" + fwaname + "," + fwbname.upper() + "," + str(
                    leafid_b) + "," + "pc_" + fwbname + "," + dc.upper() + "_" + district.upper() + "_" + "PHYS_DOM" "," + vlan + "," + fwcluster + '\n')
                f.close()

                f = open(dir_path + "/ACI_PRE_WORK/3.10 - Create the service graph template.csv", "a")
                f.write(tenant + "," + sgtname + "," + l4l7name + '\n')
                f.close()

                f = open(dir_path + "/ACI_PRE_WORK/3.11 - Create the policy-based redirect policies.csv", "a")
                f.write(tenant + "," + pbrname + "," + fwvip + "," + '\n')
                f.close()

                f = open(
                    dir_path + "/ACI_PRE_WORK/3.12 - Create a new contract with the Service Graph name defined as the subject.csv",
                    "a")
                f.write(tenant + "," + sgcontractname + "," + "Permit_Any" + '\n')
                f.close()

                f = open(dir_path + "/ACI_PRE_WORK/3.13 - Create device selection policy.csv", "a")
                f.write(
                    tenant + "," + sgcontractname + "," + sgtname + "," + "N1," + l4l7name + "," + fwbdname + "," + pbrname + "," + fwcluster + '\n')
                f.close()

                f = open(
                    dir_path + "/ACI_PRE_WORK/3.14  - Assign the L4-L7 service graph to the new contract in the contract subject.csv",
                    "a")
                f.write(tenant + "," + sgcontractname + "," + sgtname + "," + "Permit_Any" + '\n')
                f.close()
                break

    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                dirname = epg.split("-")
                tenantdir = dirname[0]
                os.mkdir(dir_path + "/ACI_" + tenantdir + "-" + vrf)
                break

    # MIGRATION STEPS - PRINT CONTRACTS
    # Print L3 out associate new contract
    l3out = {}
    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                for d in write_to_aci_cfg[tenant][vrf][epg]:
                    if d['remove_l3_contract'] == 'yes' and d['l3'] == 'yes':
                        dirname = epg.split("-")
                        tenantdir = dirname[0]
                        fname = vrf + " 1 Type " + d['t_type'] + " - Associate contracts to L3Out as consumer.csv"
                        if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
                            f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                            f.write("TENANT,L3OUT,NETWORK,NEW_L3_CONTRACT" + '\n')
                            f.close()

                        l3out_name = d['l3out']
                        ext_epg = d['extepg']
                        newcontract = tenantdir + "-" + vrf + "-" + "SG-PBR-Permit_Any"
                        newcontract_key = tenantdir + "-" + vrf + "-" + l3out_name + "-" + ext_epg + "-" + newcontract
                        if newcontract_key not in l3out and l3out_name != 'N/A':
                            l3out[newcontract_key] = {}
                            f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                            fverify = open(dir_path + "/ACI_CONTRACT_VERIFICATION/" + tenantdir + "-" + vrf, "a")

                            f.write(tenant + "," + l3out_name + "," + ext_epg + "," + newcontract + '\n')
                            fverify.write("L3Out," + tenant + "," + l3out_name + "," + ext_epg + '\n')

                            fverify.close()
                            f.close()

    # Print EPG associate new contract
    # For type-B, remove OLD EPG contract
    save_ttype = {}

    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                if toexclude is True and epg in exinfo:
                    print "OK: Excluding EPG " + epg + " from new contract association"
                    continue
                for d in write_to_aci_cfg[tenant][vrf][epg]:
                    #if d['l3'] == 'yes' or ( d['l2_unknown_unicast'] == 'proxy' and d['arp_flood'] == 'no' ) or bool(re.search('DCX', epg, re.IGNORECASE)):
		    # 10-15-2020 Exclude DCX if L2
                    if d['l3'] == 'yes' or ( d['l2_unknown_unicast'] == 'proxy' and d['arp_flood'] == 'no' ):
			if d['l3'] == 'no' and d['l2_unknown_unicast'] == 'proxy' and d['arp_flood'] == 'no':
				print "WARNING: EPG %s has unicast routing disabled, but ARP flooding set to %s and L2 Unknown unicast set to %s" % (epg,d['arp_flood'],d['l2_unknown_unicast'])
                        s = epg.split("-")
                        tenantdir = s[0]
                        t_type = d['t_type']
                        ap = d['ap']
                        if t_type == 'A':
                            fname = vrf + " 2 Type " + d['t_type'] + " - Assign new contract as provider.csv"
                            if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
                                f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                f.write("TENANT,AP,EPG,NEW_EPG_CONTRACT" + '\n')
                                f.close()
                            f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                            f.write(tenant + "," + ap + "," + epg + "," + tenantdir + "-" + vrf + "-" + "SG-PBR-Permit_Any" + '\n')
                            f.close()

                            fverify = open(dir_path + "/ACI_CONTRACT_VERIFICATION/" + tenantdir + "-" + vrf, "a")
                            fverify.write("EPG," + tenant + "," + ap + "," + epg + '\n')
                            fverify.close()

                        if t_type == 'B':
                            fname = vrf + " 2 Type " + d['t_type'] + " - Assign new contract as provider and delete old contracts.csv"
                            if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
                                f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                f.write("TENANT,AP,EPG,NEW_EPG_CONTRACT,OLD_EPG_CONTRACT" + '\n')
                                f.close()
                            f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                            for c in d['contract']:
                                f.write(tenant + "," + ap + "," + epg + "," + tenantdir + "-" + vrf + "-" + "SG-PBR-Permit_Any" + "," + c + '\n')
                            f.close()

			    if len(d['contract']) == 0:
				print "WARNING: No contracts found on TYPE-B EPG %s.  Adding new contracts only as provider" % epg
                                fname = vrf + " 2 Type " + d['t_type'] + " - Assign new contract as provider only.csv"
                                if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
                                	f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                	f.write("TENANT,AP,EPG,NEW_EPG_CONTRACT" + '\n')
                                	f.close()
				f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                f.write(tenant + "," + ap + "," + epg + "," + tenantdir + "-" + vrf + "-" + "SG-PBR-Permit_Any" + '\n')
				f.close()

                            fverify = open(dir_path + "/ACI_CONTRACT_VERIFICATION/" + tenantdir + "-" + vrf, "a")
                            fverify.write("EPG," + tenant + "," + ap + "," + epg + '\n')
                            fverify.close()

			if vrf not in save_ttype:
				save_ttype[vrf] = t_type

    l3out = {}
    # Contract removal
    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
            for epg in write_to_aci_cfg[tenant][vrf]:
                if toexclude is True and epg in exinfo:
                    print "OK: Excluding EPG " + epg + " from old contract removal"
                    continue
                for d in write_to_aci_cfg[tenant][vrf][epg]:
                    #if d['l3'] == 'yes' or ( d['l2_unknown_unicast'] == 'proxy' and d['arp_flood'] == 'no' ) or bool(re.search('DCX', epg, re.IGNORECASE)):
		    # 10-15-2020 exclude DCX L2
                    if d['l3'] == 'yes' or ( d['l2_unknown_unicast'] == 'proxy' and d['arp_flood'] == 'no' ):
                        d_l3out = d['l3out']
                        d_extepg = d['extepg']
                        ap = d['ap']
                        s = epg.split("-")
                        t_type = d['t_type']
                        tenantdir = s[0]
                        fname = vrf + " 3 Type " + d['t_type'] + " - Remove contract from L3Out and EPG as provider_consumer.csv"
                        if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
                            f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                            f.write("TENANT,AP,EPG,OLD_EPG_CONTRACT,L3OUT,NETWORK,OLD_L3_CONTRACT" + '\n')
                            f.close()

                        # temp fix
                        if epg == 'CTL-PTD-DCX-SDE-ND-MGMT_SVR':
                            f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                            f.write('Control,' + 'Control,' + epg + ',' + 'PTD_Permit_Any' + '\n')
                            f.close()

                        for c in d['contract']: 
   
                            if t_type == 'A':
                                f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                f.write(tenant + "," + ap + "," + epg + "," + c + '\n')
                                f.close()

			    if d_l3out == 'N/A' and d_extepg == 'N/A':
				continue

                            if d_l3out not in l3out:
				if not bool(re.search("F5",d_l3out)):
                                	l3out[d_l3out] = {}
				#else:
					#print "NOT adding d_l3out %s to l3out" % d_l3out
                                if d_l3out in l3out and d_extepg not in l3out[d_l3out]:
                                    l3out[d_l3out][d_extepg] = {}
                                    l3out[d_l3out][d_extepg][c] = {}
                                    f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                    f.write(tenant + ",,,," + d_l3out + "," + d_extepg + "," + c + '\n')
                                    f.close()
                                if d_l3out in l3out and d_extepg in l3out[d_l3out]:
                                    if c not in l3out[d_l3out][d_extepg]:
                                        l3out[d_l3out][d_extepg][c] = {}
                                        f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                        f.write(tenant + ",,,," + d_l3out + "," + d_extepg + "," + c + '\n')
                                        f.close()
                            if d_l3out in l3out:
                                if d_extepg not in l3out[d_l3out]:
                                    l3out[d_l3out][d_extepg] = {}
                                    l3out[d_l3out][d_extepg][c] = {}
                                    f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                    f.write(tenant + ",,,," + d_l3out + "," + d_extepg + "," + c + '\n')
                                    f.close()
                                if d_extepg in l3out[d_l3out]:
                                    if c not in l3out[d_l3out][d_extepg]:
                                        l3out[d_l3out][d_extepg][c] = {}
                                        f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
                                        f.write(tenant + ",,,," + d_l3out + "," + d_extepg + "," + c + '\n')
                                        f.close()
    

    # Remove contracts that exist on the EPG but not on the L3 Out

    for tenant in write_back_contract:
        for ap in write_back_contract[tenant]:
            for epg in write_back_contract[tenant][ap]:
                xx = epg.split("-")
                shorttenant = xx[0]
                shortvrf = xx[1]
		if shortvrf[0] == 'V' and shortvrf[1] == 'N' and shorttenant == 'LTD':
                        shortvrf = 'VND'


                fname = shortvrf + " 4 " + " - Remove contract from EPG as provider_consumer.csv"
                if not os.path.isfile(dir_path + "/ACI_" + shorttenant + "-" + shortvrf + "/" + fname):
                    f = open(dir_path + "/ACI_" + shorttenant + "-" + shortvrf + "/" + fname, "a")
                    f.write("TENANT,AP,EPG,OLD_EPG_CONTRACT" + '\n')
                    f.close()
                f = open(dir_path + "/ACI_" + shorttenant + "-" + shortvrf + "/" + fname, "a")
                for c in write_back_contract[tenant][ap][epg]:
                    f.write(tenant + ',' + ap + ',' + epg + ',' + c + '\n')
                f.close()

    
    f = open('write_to_aci_cfg.json', 'w')
    f.write(json.dumps(write_to_aci_cfg))
    f.close()
     
    total_l3_contracts = get_total_l3_contracts(dafe_file)
    # Get L3outs for VRFs that have no EPGs
    (all_l3_contracts) = append_l3out_no_bd(vrf_no_epg,all_l3_contracts,dafe_file)
    
    # Remove contracts on L3Out that are not in use. Check if contract is in use on the EPGs as well
    # Take into account L3outs that are have no EPGs
    # l3out dictionary has list of contracts to be removed
    
    # Get VRF list
    vrflist = []
    for tenant in all_l3_contracts:
	for vrf in all_l3_contracts[tenant]:
		vrflist.append(vrf)
    
    # this function is to add any L3outs that exist within the VRF (F5/VGI), other than the standard VRF L3out
    # result should be added to dictionary l3out.  l3out -> extepg -> contract (as dictionary) which should list all the contracts to be removed
    (l3out,delta_l3outs) = add_other_l3outs_to_vrfl3out(total_l3_contracts,l3out,vrflist,save_ttype)
    #print json.dumps(l3out)
    
    contracts_to_remove = []
    c_checked_in_l3out = []
  
    output_messages = {}
 
    for tenant in all_l3_contracts:
	for vrf in all_l3_contracts[tenant]:
		# Only need to do this here
		if vrf not in output_messages:
			output_messages[vrf] = []
		#print "\nVRF " + vrf
		for l3outs in all_l3_contracts[tenant][vrf]:
			for extepg in all_l3_contracts[tenant][vrf][l3outs]:
				for ctype in ('provider','consumer'):
					for c in all_l3_contracts[tenant][vrf][l3outs][extepg][ctype]:
					# Check if contract is slated to be removed
						if c in l3out[l3outs][extepg]:
							#print "FOUND %s CONTRACT %s in L3out %s, EXT EPG: %s, SLATED TO BE REMOVED ALREADY" % (ctype,c,l3outs,extepg)
							c_checked_in_l3out.append(c)
							continue
								
						else:
							# Check other L3 outs that are targeted in this change
							for e in l3out:
								for f in l3out[e]:
									if c in l3out[e][f] and e != l3outs and f != extepg:
										if tenant in delta_l3outs:
											if vrf in delta_l3outs[tenant]:
												for ccctype in delta_l3outs[tenant][vrf]:
													if e in delta_l3outs[tenant][vrf][ccctype]:
														if f not in delta_l3outs[tenant][vrf][ccctype][e]:
															c_checked_in_l3out.append(c)
															continue
							#print "COULD NOT FIND %s CONTRACT %s in L3out %s, EXT EPG: %s, checking other ext epgs" % (ctype,c,l3outs,extepg) 
							# Check all other l3outs as provider and consumer
							if c not in c_checked_in_l3out:
								c_checked_in_l3out.append(c)
								found = check_other_l3_outs(total_l3_contracts,c,l3outs,extepg,delta_l3outs,output_messages)
								if found == 'yes':
									# save for later
									#print "WARNING: CONTRACT %s found on L3OUT %s, will not be removed" % (c,l3outs)
									output_messages[vrf].append("WARNING: CONTRACT %s found on L3OUT %s, will not be removed" % (c,l3outs))
									continue
								else:
									#print "COULD NOT FIND %s CONTRACT %s in all L3outs, checking EPGs" % (ctype,c)
									if c not in contracts_to_remove:
										(c_in_use,f_epg,f_ap) = check_epgs_for_contracts(write_to_aci_cfg,c)
										if c_in_use is False and save_ttype[vrf] == 'B' :
											write_contract_removal(tenant,vrf,c,l3outs,extepg,'B',dir_path,f_epg,f_ap,'l3out')
											contracts_to_remove.append(c)
											# save for later
											#print "OK: COULD NOT FIND CONTRACT %s in any L3out or EPG as provider or consumer, ADDING TO L3OUT CLEANUP" % (c)
											output_messages[vrf].append("OK: COULD NOT FIND CONTRACT %s in any L3out or EPG as provider or consumer, ADDING TO L3OUT CLEANUP" % (c))
										if c_in_use is True and save_ttype[vrf] == 'B':
											c_info = []
											l2orl3 = write_to_aci_cfg[tenant][vrf][f_epg][0]['l3']
											if l2orl3 == 'no':
												l2orl3 = 'L2'
												write_contract_removal(tenant,vrf,c,l3outs,extepg,'B',dir_path,f_epg,f_ap,'epg')
												write_contract_removal(tenant,vrf,c,l3outs,extepg,'B',dir_path,f_epg,f_ap,'l3out')
												update_aci_contract_verification_file(tenant,vrf,dir_path,"EPG",f_ap,f_epg)
												contracts_to_remove.append(c)
												#print "OK: COULD NOT FIND CONTRACT %s in any L3out or EPG as provider or consumer, ADDING TO L3OUT CLEANUP" % (c)
												# save for later	
												#print "OK: ** CONTRACT %s is TIED TO L2 EPG %s and will be removed from both the L3Out and EPG **" % (c,f_epg)
												output_messages[vrf].append("OK: ** CONTRACT %s is TIED TO L2 EPG %s and will be removed from both the L3Out and EPG **" % (c,f_epg))
											else:
												l2orl3 = 'L3'
												if c in write_to_aci_cfg[tenant][vrf][f_epg][0]['curr_c_contracts']:
													c_info.append('consumer')
												if c in write_to_aci_cfg[tenant][vrf][f_epg][0]['curr_p_contracts']:
													c_info.append('provider')
												c_info_msg = ','.join(c_info)
												# save for later
												#print "WARNING: CONTRACT %s found on %s EPG %s as %s, will not be removed from L3Out %s" % (c,l2orl3,f_epg,c_info_msg,l3outs)
												output_messages[vrf].append("WARNING: CONTRACT %s found on %s EPG %s as %s, will not be removed from L3Out %s" % (c,l2orl3,f_epg,c_info_msg,l3outs))
												del c_info
												del c_info_msg
		
									 
    # Write the additional L3outs (F5/VGI) to the CSV files only if Type-B using the delta_l3outs dictionary.  Type-A is taken care of with the python script
    #Type-B: Add as provider only
    #Type-B: Remove as provider/consumer
    f_epg = ''
    f_ap = ''
    
    for tenant in delta_l3outs:
	for vrf in delta_l3outs[tenant]:
		#print "\nVRF " + vrf
		for ctype in delta_l3outs[tenant][vrf]:
			if ctype == 'B':
				for l3out in delta_l3outs[tenant][vrf][ctype]:
					for extepg in delta_l3outs[tenant][vrf][ctype][l3out]:
						if  bool(re.search("F5",l3out,re.IGNORECASE)) or bool(re.search("VGI",l3out,re.IGNORECASE)) or bool(re.search("RIA",l3out,re.IGNORECASE)) or bool(re.search("DMA",l3out,re.IGNORECASE)) or bool(re.search("DMB" ,l3out,re.IGNORECASE)):
							# save for later
							#print "OK: INCLUDING VGI/F5 L3Out %s, SG CONTRACT WILL BE ADDED AS PROVIDER" % (l3out) 
							output_messages[vrf].append("OK: INCLUDING VGI/F5 L3Out %s, SG CONTRACT WILL BE ADDED AS PROVIDER" % (l3out)) 
    							write_contract_addition(tenant,vrf,l3out,extepg,dir_path,ctype) 
							for c in delta_l3outs[tenant][vrf][ctype][l3out][extepg]:
								found = check_other_l3_outs(total_l3_contracts,c,l3out,extepg,delta_l3outs,output_messages)
								if found == 'no':
    									write_contract_removal(tenant,vrf,c,l3out,extepg,ctype,dir_path,f_epg,f_ap,'l3out') 
								else:
									# save for later
									#print "WARNING: CONTRACT %s found on L3OUT %s, will not be removed" % (c,l3out)
									output_messages[vrf].append("WARNING: CONTRACT %s found on L3OUT %s, will not be removed" % (c,l3out))
						else:
							# save for later
							#print "OK: INCLUDING VRF L3Out %s, SG CONTRACT WILL BE ADDED AS CONSUMER" % (l3out) 
							output_messages[vrf].append("OK: INCLUDING VRF L3Out %s, SG CONTRACT WILL BE ADDED AS CONSUMER" % (l3out)) 
    							write_contract_addition_consumer(tenant,vrf,l3out,extepg,dir_path,ctype) 
							for c in delta_l3outs[tenant][vrf][ctype][l3out][extepg]:
								found = check_other_l3_outs(total_l3_contracts,c,l3out,extepg,delta_l3outs,output_messages)
								if found == 'no':
    									write_contract_removal(tenant,vrf,c,l3out,extepg,ctype,dir_path,f_epg,f_ap,'l3out') 
								else:
									# save for later
									#print "WARNING: CONTRACT %s found on L3OUT %s, will not be removed" % (c,l3out)
									output_messages[vrf].append("WARNING: CONTRACT %s found on L3OUT %s, will not be removed" % (c,l3out))
						update_aci_contract_verification_file(tenant,vrf,dir_path,"L3Out",l3out,extepg)

    # Print output messages
    for vrf in output_messages:
	if len(output_messages[vrf]) > 0:
		print "\nVRF " + vrf
		output_messages[vrf] = remove_dups(output_messages[vrf])
		output_messages[vrf].sort()
		for m in output_messages[vrf]:
			print (m)
 
    # For Type-A, print out the script to run and which subnets will be moved
    found_A_to_remove = 0
    aci_json_cfg = load_aci_json(dc,district)
    
    for tenant in delta_l3outs:
	for vrf in delta_l3outs[tenant]:
		for ctype in delta_l3outs[tenant][vrf]:
			if ctype == 'A':
				if found_A_to_remove == 0:
    					print "\nOK: VRF %s: RUN THE FOLLOWING SCRIPTS TO MIGRATE TYPE-A F5 L3OUTs" % vrf
					found_A_to_remove = 1
				
				for l3out in delta_l3outs[tenant][vrf][ctype]:
					for extepg in delta_l3outs[tenant][vrf][ctype][l3out]:
						x = l3out.split("-")
						x[0] = x[0].replace("L3Out_","")
						shorttenant = x[0]
						print "./f5_typeA_l3out.py -l %s -t %s -c ACI_CONTRACT_VERIFICATION/aci_creds -a %s-%s-SG-PBR-Permit_Any" % (l3out,tenant,shorttenant,vrf) 
   						subnet_nhip = get_subnet_nhip(l3out,aci_json_cfg,tenant)
						print "LEAF ID/SUBNET LIST TO BE MOVED:"
						for leaf in subnet_nhip:
							snlist = ','.join(subnet_nhip[leaf])
							leaf = leaf.replace("topology/pod-1/","")
							print leaf + ":" + snlist
						print "\n\n\n" 


    #print json.dumps(delta_l3outs)
    #print json.dumps(all_l3_contracts)

    print "\n\n\n"
    print "!" * 50
    print "\n\nFollow the below steps as well\n\n"

    # Put the 3 python scripts here.  If can't find it, print out warning for user to add it
    if os.path.isfile("./getcontracts.py"):
        shutil.copyfile("./getcontracts.py", "./output/ACI_CONTRACT_VERIFICATION/getcontracts.py")
        os.chmod("./output/ACI_CONTRACT_VERIFICATION/getcontracts.py", 0755)

    else:
        print "Make sure to copy getcontracts.py to ./output/ACI_CONTRACT_VERIFICATION folder"

    if os.path.isfile("./push_to_n7k.py"):
        shutil.copyfile("./push_to_n7k.py", "./output/push_to_n7k.py")
        os.chmod("./output/push_to_n7k.py", 0755)

    else:
        print "Make sure to copy push_to_n7k.py to the output folder"

    if os.path.isfile("./f5_typeA_l3out.py"):
        shutil.copyfile("./f5_typeA_l3out.py", "./output/f5_typeA_l3out.py")
        os.chmod("./output/f5_typeA_l3out.py", 0755)

    else:
        print "Make sure to copy f5_typeA_l3out.py to the output folder"
    
    if os.path.isfile("./n7k_verification.sh"):
        shutil.copyfile("./n7k_verification.sh", "./output/n7k_verification.sh")
        os.chmod("./output/n7k_verification.sh", 0755)

    else:
        print "Make sure to copy n7k_verification.sh to the output folder"
    
    if os.path.isfile("./check_n7k_output.py"):
        shutil.copyfile("./check_n7k_output.py", "./output/check_n7k_output.py")
        os.chmod("./output/check_n7k_output.py", 0755)

    else:
        print "Make sure to copy check_n7k_output.py to the output folder"

    print "\n\nDo not include the N7K_NEXT_CLEANUP as part of this change window.  It is to be used for the NEXT change window.  Running the commands in this folder will undo everything you have done!!"
    #print "\n\nDuring migration, run the script f5_typeA_l3out.py to add the static routes to the external EPG. For Type-B VRF's, use the postman scripts"
    #print "\n\nCheck to make sure the generic 'Permit_Any' contract is removed from the L3Out for Type-B VRFs if not in use"

    print "\n\n\n"
    print "!" * 50
    print "\n\n\n"


if __name__ == '__main__':
    main(sys.argv[1:])
