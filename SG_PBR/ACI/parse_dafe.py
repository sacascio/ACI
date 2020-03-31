#!/usr/bin/env python

import openpyxl
import operator
import ipaddress
from netaddr import *
import shutil
import sys
import getopt
import os.path
import fnmatch 
import re
from IPy import IP
#from xlrd import open_workbook, XLRDError
import json
#from fileinput import filename
import warnings
from ciscoconfparse import CiscoConfParse

warnings.filterwarnings("ignore")


def write_new_n7k_configs(vrfmember,p2psubnets,dc,district,n7k_data):
        dir_path = './output'
	inner_bgp_config = {}
	outer_bgp_config = {}
	bgp_shut_inner = {}
	bgp_shut_outer = {}
	bgp_rb_inner = {}
	bgp_rb_outer = {}
	bgp_rb_outer = {}
	svi_cleanup = {}
	
	cutover_dir = dir_path + "/" + "N7K_CUTOVER" + "/" + vrfmember
	rollback_dir = dir_path + "/" + "N7K_ROLLBACK" + "/" + vrfmember
	cleanup_dir = dir_path + "/" + "N7K_NEXT_CLEANUP" + "/" + vrfmember

	if not os.path.exists(dir_path + "/" + "N7K_PREWORK"):
		os.mkdir(dir_path + "/" + "N7K_PREWORK")
	if not os.path.exists(dir_path + "/" + "N7K_CUTOVER"):
		os.mkdir(dir_path + "/" + "N7K_CUTOVER")
	if not os.path.exists(dir_path + "/" + "N7K_ROLLBACK"):
		os.mkdir(dir_path + "/" + "N7K_ROLLBACK")
	if not os.path.exists(dir_path + "/" + "N7K_NEXT_CLEANUP"):
		os.mkdir(dir_path + "/" + "N7K_NEXT_CLEANUP")

	os.mkdir(cutover_dir)
	os.mkdir(rollback_dir)
	os.mkdir(cleanup_dir)

	k = 0

	# Get Inner BGP Details
	for n7k in n7k_data:
		if bool(re.search('outer',n7k, re.IGNORECASE)):
			continue
		bgp_rb_inner[n7k] = {}
		bgp_rb_inner[n7k]['neighbors'] = []
		bgp_rb_inner[n7k]['subint'] = []
		bgp_rb_inner[n7k]['svi'] = []

		neighbors = n7k_data[n7k][vrfmember]['neighbors']
		local_as = n7k_data[n7k][vrfmember]['local_as']	
		remote_as = n7k_data[n7k][vrfmember]['remote_as']	
		bgp_shut_inner[n7k] = {}
		bgp_shut_inner[n7k][vrfmember] = []
		bgp_shut_inner[n7k][vrfmember].append("! Shutdown the BGP adjacency to the outer N7K VDC connected to the FW")
		bgp_shut_inner[n7k][vrfmember].append("router bgp " + local_as)
		bgp_shut_inner[n7k][vrfmember].append(" vrf " + vrfmember)
	
		# Squeeze in rollback for new adjacencies
		bgp_rb_inner[n7k]['neighbors'].append("router bgp " + local_as)
		bgp_rb_inner[n7k]['neighbors'].append(" vrf " + vrfmember)
	
		for n in neighbors:
			bgp_shut_inner[n7k][vrfmember].append(" neighbor " +  n + " remote-as " + remote_as )
			bgp_shut_inner[n7k][vrfmember].append("   shutdown")

			# Rollback
			bgp_rb_inner[n7k]['neighbors'].append("   neighbor " +  n + " remote-as " + remote_as)
			bgp_rb_inner[n7k]['neighbors'].append("     no shutdown")
			

	# Get outer VDC BGP details
	for n7k in n7k_data:
		if bool(re.search('inner',n7k, re.IGNORECASE)):
			continue
		bgp_rb_outer[n7k] = {}
                bgp_rb_outer[n7k]['neighbors'] = []
                bgp_rb_outer[n7k]['subint'] = []
                bgp_rb_outer[n7k]['svi'] = []

		for svi in n7k_data[n7k]:
			if svi == 'P2P':
				continue
			remote_as = n7k_data[n7k][svi]['remote_as']
			local_as = n7k_data[n7k][svi]['local_as']
			if 'vrf' not in n7k_data[n7k][svi]:	
				continue
			for vrf in n7k_data[n7k][svi]['vrf']:
				if vrf == vrfmember:
					outer_neighbors = n7k_data[n7k][svi]['vrf'][vrf]['neighbor_ip']
					bgp_shut_outer[n7k] = []
					bgp_shut_outer[n7k].append("! Shutdown the BGP adjacency to the N7K Inner in VRF " + vrfmember)
					bgp_shut_outer[n7k].append("router bgp " + local_as)

					# Squeeze in rollback for BGP Neighbors and the SVI
					bgp_rb_outer[n7k]['neighbors'].append("router bgp " + local_as)
					bgp_rb_outer[n7k]['svi'].append("interface Vlan" + svi)
					bgp_rb_outer[n7k]['svi'].append(" no shutdown")

					for n in outer_neighbors:
						bgp_shut_outer[n7k].append(" neighbor " +  n + " remote-as " + remote_as )
						bgp_shut_outer[n7k].append("   shutdown")

						# Rollback
						bgp_rb_outer[n7k]['neighbors'].append(" neighbor " +  n + " remote-as " + remote_as )
						bgp_rb_outer[n7k]['neighbors'].append("    no shutdown" )

	if district.upper() == 'SDE':
		numn7k = ['1','2']
	else:
		numn7k = ['1','2','3','4']
		
	for i in numn7k:
		n7kname = dc + district + 'nxc' + i + district +  'inner'
		encap = n7k_data[n7kname][vrfmember]['svi']
		inner_bgp_as = n7k_data[n7kname][vrfmember]['local_as']
		outer_bgp_as = n7k_data[n7kname][vrfmember]['remote_as']
		inner_bgp_config[n7kname] = []

		# Start preparing inner/outer BGP config
		inner_bgp_config[n7kname].append("! Add new BGP neighbors to VRF " + vrfmember + " using the N7K Outer VDC IP addresses")	
		inner_bgp_config[n7kname].append("! These adjacencies will not come up until the N7K Outer VDCs are configured and enabled" + '\n')	

		inner_bgp_config[n7kname].append("router bgp " + inner_bgp_as)	
		inner_bgp_config[n7kname].append(" vrf " + vrfmember)	
		inner_bgp_config[n7kname].append("  address-family ipv4 unicast")	
		inner_bgp_config[n7kname].append("   maximum-paths 8")	
		

		if not os.path.exists(dir_path + "/" + "N7K_PREWORK" + "/" +  n7kname):
			f = open(dir_path + "/" + "N7K_PREWORK" + "/" +  n7kname, "a")
    			f.write("! Create sub interfaces to outer VDCs in VRF " + vrfmember + " in a shutdown state" + '\n')
                        f.write("configure terminal" +  '\n')
			f.close()
		else:
    			f = open(dir_path + "/" + "N7K_PREWORK" + "/" +  n7kname, "a")
    			f.write("! Create sub interfaces to outer VDCs in VRF " + vrfmember + " in a shutdown state" + '\n')
			f.close()

		if not os.path.exists(cutover_dir + "/" +  n7kname):
			f = open(cutover_dir + "/" +  n7kname, "a")
			f.write("configure terminal" + '\n')
			f.close()

		for inner_int in n7k_data[n7kname]['P2P']:
			for outer_int in n7k_data[n7kname]['P2P'][inner_int]:
				for outer_7k in n7k_data[n7kname]['P2P'][inner_int][outer_int]:
					if outer_7k not in outer_bgp_config:
						outer_bgp_config[outer_7k] = []
						outer_bgp_config[outer_7k].append("! Add new BGP neighbors to VRF " + vrfmember + " using the N7K Inner VDC IP addresses")	
						outer_bgp_config[outer_7k].append("! These adjacencies will not come up until the N7K Inner VDCs are configured and enabled" + '\n')	
						outer_bgp_config[outer_7k].append("router bgp " + outer_bgp_as)	
					

					# Write Inner config
					f = open(dir_path + "/" + "N7K_PREWORK" + "/" +  n7kname, "a")
    					f.write("interface Ethernet" + inner_int + "." + encap + '\n')
    					f.write(" description To_" + outer_7k + "_E" + outer_int + "." + encap + '\n')
    					f.write(" shutdown" + '\n')
    					f.write(" vrf member " + vrfmember + '\n')
    					f.write(" mtu 9192" + '\n')
    					f.write(" encapsulation dot1Q " + encap + '\n')
			
					# Write cutover part to enable the new sub interfaces on inner		
					fsub = open(cutover_dir + "/" +  n7kname, "a")
                                        fsub.write("! Enable sub interfaces to outer VDCs for VRF " + vrfmember +  '\n')
                                        fsub.write("interface Ethernet" + inner_int + "." + encap + '\n')
                                        fsub.write(" no shutdown" + '\n')
                                        fsub.write('\n')
					fsub.close()

					# Write Rollback part for inner
					if bool(re.search('inner',n7kname, re.IGNORECASE)):
						bgp_rb_inner[n7kname]['subint'].append("no interface Ethernet" + inner_int + "." + encap)
					
					# Custom IP addressing for SDE because we already started this way
					if k == 2 and district.upper() == 'SDE':
						ipinner = p2psubnets[3]
					elif k == 3 and district.upper() == 'SDE':
						ipinner = p2psubnets[2]
					else:	
						ipinner = p2psubnets[k]

					ipinnerx = ipinner.split('/')
					ipinner = IPAddress(ipinnerx[0])
					mask = ipinnerx[1]
					ipinner = ipinner + 1
					ipouter = ipinner + 1
					ipinner = str(IPAddress(ipinner))
					ipouter = str(IPAddress(ipouter))

					inner_bgp_config[n7kname].append("  neighbor " + ipouter + " remote-as " + outer_bgp_as)
					inner_bgp_config[n7kname].append("   description TO_" + outer_7k)
					inner_bgp_config[n7kname].append("   address-family ipv4 unicast")
					inner_bgp_config[n7kname].append("    send-community both")
		
					# Write rollback part for inner BGP config and write it later
					if bool(re.search('inner',n7kname, re.IGNORECASE)):
						bgp_rb_inner[n7kname]['neighbors'].append("   no neighbor " + ipouter + " remote-as " + outer_bgp_as) 
					
					outer_bgp_config[outer_7k].append("  neighbor " + ipinner + " remote-as " + inner_bgp_as)
					outer_bgp_config[outer_7k].append("   description TO_" + n7kname)
					outer_bgp_config[outer_7k].append("   address-family ipv4 unicast")
					outer_bgp_config[outer_7k].append("    send-community both")
					outer_bgp_config[outer_7k].append("    route-map PERMIT_DEFAULT_ONLY out")
					outer_bgp_config[outer_7k].append("    default-originate")
					
					# Write rollback part for outer BGP config and write it later
					if bool(re.search('outer',outer_7k, re.IGNORECASE)):
						bgp_rb_outer[outer_7k]['neighbors'].append("   no neighbor " + ipinner + " remote-as " + inner_bgp_as) 

    					f.write(" ip address " + ipinner + "/" + mask + '\n')
    					f.write( '\n')
					k = k+1
					f.close()
				
					# Write outer config
					if not os.path.exists(dir_path + "/" + "N7K_PREWORK" + "/" +  outer_7k):
						f = open(dir_path + "/" + "N7K_PREWORK" + "/" +  outer_7k, "a")
						f.write("! Create sub interfaces to inner VDCs for VRF " + vrfmember + " in a shutdown state" + '\n')
						f.write("configure terminal" + '\n')
						f.close()
						
					f = open(dir_path + "/" + "N7K_PREWORK" + "/" +  outer_7k, "a")
					f.write("interface Ethernet" + outer_int + "." + encap + '\n')
                                        f.write(" description To_" + n7kname + "_E" + inner_int + "." + encap + '\n')
                                        f.write(" shutdown" + '\n')
                                        f.write(" mtu 9192" + '\n')
                                        f.write(" encapsulation dot1Q " + encap + '\n')
					f.write(" ip address " + ipouter + "/" + mask + '\n')
                                        f.write( '\n')
    					f.close()
					
					# Write cutover part to enable the new sub interfaces on outer
					if not os.path.exists(cutover_dir + "/" + outer_7k):
						f = open(cutover_dir + "/" +  outer_7k, "a")
						f.write("configure terminal" +  '\n')
						f.close()
					f = open(cutover_dir + "/" +  outer_7k, "a")
					f.write("! Enable sub interfaces to inner VDCs for VRF " + vrfmember +  '\n')
					f.write("interface Ethernet" + outer_int + "." + encap + '\n')
					f.write(" no shutdown" + '\n')
					f.write('\n')
					f.close()

					# Rollback new sub interfaces on outer
					if bool(re.search('outer',outer_7k, re.IGNORECASE)):
						bgp_rb_outer[outer_7k]['subint'].append("no interface Ethernet" + outer_int + "." + encap)
			
		# Shutdown existing SVI on inner
		fsvi = open(cutover_dir + "/" +  n7kname, "a")
    		fsvi.write("! Shutdown inside SVI vrf " + vrfmember + " for firewall cluster" + '\n')
    		fsvi.write("interface Vlan" + encap + '\n')
    		fsvi.write(" shutdown" + '\n')
    		fsvi.write('\n')
		fsvi.close()
		
		# Write rollback part
		if bool(re.search('inner',n7kname, re.IGNORECASE)):
    			bgp_rb_inner[n7kname]['svi'].append("interface Vlan" + encap)
    			bgp_rb_inner[n7kname]['svi'].append(" no shutdown")
		
		# Write cleanup part for next window
		if bool(re.search('inner',n7kname, re.IGNORECASE)):
			if n7kname not in svi_cleanup:
				svi_cleanup[n7kname] = {}
				svi_cleanup[n7kname]['svi'] = []
    			svi_cleanup[n7kname]['svi'].append("no interface Vlan" + str(encap))
				
					
	for n7ks in inner_bgp_config:
		f = open(dir_path + "/" + "N7K_PREWORK" + "/" +  n7ks, "a")
		f.write(('\n'.join(inner_bgp_config[n7ks])))
		f.write('\n')
		f.write('\n')
		f.close()
	
	for n7ks in outer_bgp_config:
		f = open(dir_path + "/" + "N7K_PREWORK" + "/" +  n7ks, "a")
		f.write(('\n'.join(outer_bgp_config[n7ks])))
		f.write('\n')
		f.write('\n')
		f.close()

	for n7ks in bgp_shut_outer:
		f = open(cutover_dir + "/" +  n7ks, "a")
		f.write(('\n'.join(bgp_shut_outer[n7ks])))
		f.write('\n')
		f.write('\n')
		f.close()
	
	for n7ks in bgp_shut_inner:
		for vrfs in bgp_shut_inner[n7ks]:
			f = open(cutover_dir + "/" +  n7ks, "a")
			f.write(('\n'.join(bgp_shut_inner[n7ks][vrfs])))
			f.write('\n')
			f.write('\n')
			f.close()
		
	for n7ks in bgp_rb_inner:
		f = open(rollback_dir + "/" +  n7ks, "a")
		f.write("!!" + '\n')
		f.write("!! N7K VDC - VRF " + vrfmember +  '\n')
		f.write("!!" + '\n')
		f.write("configure terminal" + '\n')
		f.write("! Remove new sub interfaces to Outer VDCs" + '\n')
		f.write(('\n'.join(bgp_rb_inner[n7ks]['subint'])))
		f.write('\n')
		f.write('\n')
		f.write("! Remove direct BGP adjacencies to Outer VDCs and re-enable BGP Adjacencies to FW-Inner" + '\n')
		f.write(('\n'.join(bgp_rb_inner[n7ks]['neighbors'])))
		f.write('\n')
		f.write('\n')
		f.write("! Re-enable SVI to FW-Inner" + '\n')
		f.write(('\n'.join(bgp_rb_inner[n7ks]['svi'])))
		f.write('\n')
		f.close()
	
	for n7ks in bgp_rb_outer:
		f = open(rollback_dir + "/" +  n7ks, "a")
		f.write("!!" + '\n')
		f.write("!! N7K Outer VDC - VRF " + vrfmember +  '\n')
		f.write("!!" + '\n')
		f.write("configure terminal" + '\n')
		f.write("! Remove new sub interfaces to Inner VDCs" + '\n')
		f.write(('\n'.join(bgp_rb_outer[n7ks]['subint'])))
		f.write('\n')
		f.write('\n')
		f.write("! Remove direct BGP adjacencies to Inner VDCs and re-enable BGP Adjacencies to FW-Outer" + '\n')
		f.write(('\n'.join(bgp_rb_outer[n7ks]['neighbors'])))
		f.write('\n')
		f.write('\n')
		f.write("! Re-enable SVI to FW-Outer" + '\n')
		f.write(('\n'.join(bgp_rb_outer[n7ks]['svi'])))
		f.write('\n')
		f.close()
	
	for n7ks in svi_cleanup:
		f = open(cleanup_dir + "/" +  n7ks, "a")
		f.write("!!" + '\n')
		f.write("!! SVI Removal N7K Inner VDC - VRF " + vrfmember +  '\n')
		f.write("!!" + '\n')
                f.write("configure terminal" + '\n')
		f.write(('\n'.join(svi_cleanup[n7ks]['svi'])))
                f.write('\n')

	for n7ks in n7k_data:
		if bool(re.search('inner',n7ks, re.IGNORECASE)):
			fwints = n7k_data[n7ks][vrfmember]['fw_trunk_int']
			svi = n7k_data[n7ks][vrfmember]['svi']
			f = open(cleanup_dir + "/" + n7ks, "a")
			f.write('\n')
			f.write("!! Remove VLAN from firewall trunk" + '\n')
			for fw in fwints:
				f.write("interface Ethernet" + fw + '\n')
				f.write(" switchport trunk allowed vlan remove " + svi + '\n')
				f.write('\n')
			f.close()	
		
def get_inner_outer_mapping(dc):

    pattern = '*' + dc.upper() + '*Port Map*'
  
    files = os.listdir('.') 
    for name in files:
        if fnmatch.fnmatch(name, pattern):
                if not bool(re.search('^~',name, re.IGNORECASE)): 
                        filename = name
                        break
    worksheets = []
    wb = openpyxl.load_workbook(filename, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('7706')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    mapping = {}
        
    for cells in ws.iter_rows(min_row=row_start, min_col=1, max_col=24): 
        for vals in cells:
                if vals.value == 'A':
                        n7k_local = 1
                if vals.value == 'B':
                        n7k_local = 2
                if vals.value == 'C':
                        n7k_local = 3
                if vals.value == 'D':
                        n7k_local = 4
                if vals.value == 'E':
                        n7k_local = 1
                if vals.value == 'F':
                        n7k_local = 2

                if bool(re.search("nxc(.*)inner",str(vals.value), re.IGNORECASE)) or bool(re.search("nxc(.*)outer",str(vals.value), re.IGNORECASE)):
                        remote_n7k = vals.value
                        local_interface = ws[vals.column + str(vals.row+1)].value

                        if local_interface is None:
                                local_interface = ws[vals.column + str(vals.row-1)].value

                        local_interface = local_interface.replace("-","/")
                        xvals = vals.value.split("_E")
                        remote_interface = xvals[1]
                        remote_n7k = xvals[0]

                        # Determine local N7K now
                        if bool(re.search("sde",remote_n7k, re.IGNORECASE)) and bool(re.search("inner",remote_n7k, re.IGNORECASE)):
                                local_n7k = dc + 'sde' + 'nxc' + str(n7k_local) + 'sdeouter'
                        if bool(re.search("sde",remote_n7k, re.IGNORECASE)) and bool(re.search("outer",remote_n7k, re.IGNORECASE)):
                                local_n7k = dc + 'sde' + 'nxc' + str(n7k_local) + 'sdeinner'

                        if bool(re.search("gis",remote_n7k, re.IGNORECASE)) and bool(re.search("inner",remote_n7k, re.IGNORECASE)):
                                local_n7k = dc + 'gis' + 'nxc' + str(n7k_local) + 'gisouter'
                        if bool(re.search("gis",remote_n7k, re.IGNORECASE)) and bool(re.search("outer",remote_n7k, re.IGNORECASE)):
                                local_n7k = dc + 'gis' + 'nxc' + str(n7k_local) + 'gisinner'

                        if bool(re.search("soe",remote_n7k, re.IGNORECASE)) and bool(re.search("inner",remote_n7k, re.IGNORECASE)):
                                local_n7k = dc + 'soe' + 'nxc' + str(n7k_local) + 'soeouter'
                        if bool(re.search("soe",remote_n7k, re.IGNORECASE)) and bool(re.search("outer",remote_n7k, re.IGNORECASE)):
                                local_n7k = dc + 'soe' + 'nxc' + str(n7k_local) + 'soeinner'


                        if local_n7k not in mapping:
                                mapping[local_n7k] = {}

                        if local_interface not in mapping[local_n7k]:
                                mapping[local_n7k][local_interface] = {}

                        if remote_interface not in mapping[local_n7k][local_interface]:
                                mapping[local_n7k][local_interface][remote_interface] = {}

                        if remote_n7k not in mapping[local_n7k][local_interface][remote_interface]:
                                mapping[local_n7k][local_interface][remote_interface][remote_n7k] = {}
    
    return mapping

def get_bgp_int_vlan(dc,district,vrfs):

    data = {}
    p2p_n7k_mapping = get_inner_outer_mapping(dc)
 
    if district.lower() in ('sde'):
        numfiles = ['1','2']
    else:
        numfiles = ['1','2','3','4']

    for i in numfiles:
        for j in ['inner','outer']:
                filename = dc + district + 'nxc' + str(i) + district + j + '.log'
                n7k = dc + district + 'nxc' + str(i) + district + j 

                # Need to code this
                n7k_mapping = get_inner_outer_mapping(dc)

                data[n7k] = {}
                parse = CiscoConfParse(filename)

                # GET SVI
                for obj in parse.find_objects("interface Vlan"):
                        svi = obj.text
                        svi = svi.replace("interface Vlan","")

                        if obj.hash_children != 0:
                                for c in obj.children:

                                        if bool((re.search('vrf member',c.text,re.IGNORECASE))):
                                                vrf = c.text
                                                vrf = vrf.replace("vrf member ","")
                                                vrf = vrf.lstrip()

                                        # For outer, set VRF to SVI since outer does not have vrf
                                        if bool((re.search('outer',n7k,re.IGNORECASE))):
                                                vrf = svi

                                        if bool((re.search('ip address',c.text,re.IGNORECASE))):
                                                        svi_ip = c.text
                                                        svi_ip = c.text.replace("  ip address ","")
                                                        svi_ip = re.sub('/.*',"", svi_ip)

                                data[n7k][vrf] = {}
                                data[n7k]['P2P'] = {}
                                data[n7k][vrf] = { 'svi' : svi, 'shutdown' : 'N', 'fw_trunk_int' : [], 'remote_as' : 'N/A', 'neighbors' : [], 'svi_ip' : svi_ip, 'local_as' : 'N/A' }

                # GET FW INT
                vlallowed = []
                for obj in parse.find_objects("interface Ethernet"):
                        if obj.hash_children != 0:
                                if obj.re_search_children("switchport trunk allowed"):
                                        for c in obj.children:
                                                if bool((re.search('switchport trunk allowed',c.text,re.IGNORECASE))):
                                                        fwint = obj.text
                                                        fwint = fwint.replace("interface Ethernet","")
                                                        c.text = c.text.replace("switchport trunk allowed vlan","")
                                                        vlist = c.text.split(",")
                                                        for v in vlist:
                                                                if bool((re.search('-',v,re.IGNORECASE))):
                                                                        vv = v.split("-")
                                                                        vv[1] = int(vv[1]) + int(1)
                                                                        for vvv in xrange(int(vv[0]),int(vv[1])):
                                                                                vlallowed.append(vvv) 
                                                                else:
                                                                        vlallowed.append(int(v))

                                                        for vls in vlallowed:
                                                                for vrf_vals in data[n7k]:
                                                                        if vrf_vals != 'P2P':
                                                                                if vls == int(data[n7k][vrf_vals]['svi']):
                                                                                        if len(data[n7k][vrf_vals]['fw_trunk_int']) == 0:
                                                                                                data[n7k][vrf_vals]['fw_trunk_int']  = [ fwint ]
                                                                                        else:
                                                                                                data[n7k][vrf_vals]['fw_trunk_int'].append(fwint)
                                                                vlallowed = []



                # Get BGP Neighbors
                for obj in parse.find_objects("router bgp"):
                        local_as = obj.text
                        local_as = local_as.replace("router bgp ","")
                        if obj.hash_children != 0:
                                # For inner VDC
                                if obj.re_search_children("vrf"):
                                        for c in obj.children:
                                                if bool((re.search('vrf',c.text,re.IGNORECASE))):
                                                        vrfmember = c.text
                                                        for obj2 in parse.find_objects(vrfmember):
                                                                if obj2.re_search_children("neighbor"):
                                                                        for d in obj2.children:
                                                                                vrfmember = vrfmember.replace("vrf ","") 
                                                                                vrfmember = vrfmember.lstrip()
                                                                                if bool((re.search('address-family',d.text,re.IGNORECASE))):
                                                                                        continue
                                                                                if bool((re.search('NonAff',vrfmember,re.IGNORECASE))):
                                                                                        continue
                                                                                attribs = d.text
                                                                                attribs = attribs.lstrip()
                                                                                attribsx = attribs.split(" ")
                                                                                neighbor = attribsx[1]
                                                                                remote_as = attribsx[3]
                                                                                data[n7k][vrfmember]['remote_as']  = remote_as
                                                                                if len(data[n7k][vrfmember]['neighbors']) == 0:
                                                                                        data[n7k][vrfmember]['neighbors']  = [ neighbor ]
                                                                                else:
                                                                                        data[n7k][vrfmember]['neighbors'].append(neighbor)

                                                                                data[n7k][vrfmember]['local_as'] = local_as
                                # For outer VDC
                                if bool((re.search('outer',filename,re.IGNORECASE))):
                                        for c in obj.children:
                                                if bool((re.search('neighbor ',c.text,re.IGNORECASE))):
                                                        attribs = c.text
                                                        attribs = attribs.lstrip()
                                                        attribsx = attribs.split(" ")
                                                        neighbor = attribsx[1]
                                                        remote_as = attribsx[3]

                                                        for vrfmember in data:
                                                                for v in data[vrfmember]:
                                                                        if v == 'P2P':
                                                                                continue
                                                                        if data[vrfmember][v]['svi_ip'] == neighbor:
                                                                                for n in data[vrfmember][v]['neighbors']:
                                                                                        for o in data:
                                                                                                for p in data[o]:
                                                                                                        if unicode(p).isnumeric():

                                                                                                                if data[o][p]['svi_ip'] == n:
                                                                                                                        if 'vrf' not in data[o][p]:
                                                                                                                                data[o][p]['vrf'] = {}
                                                                                                                                del data[o][p]['neighbors']
                                                                                                                        if v not in data[o][p]['vrf']:
                                                                                                                                data[o][p]['vrf'][v] = {}
                                                                                                                                data[o][p]['vrf'][v]['neighbor_ip'] = [neighbor]
                                                                                                                        else:
                                                                                                                                if neighbor not in data[o][p]['vrf'][v]['neighbor_ip']:
                                                                                                                                        data[o][p]['vrf'][v]['neighbor_ip'].append(neighbor)
                                                                                                                        data[o][p]['remote_as'] = remote_as
                                                                                                                        data[o][p]['local_as'] = local_as

    # Append P2P data
    # P2P dictionary format:
    # Local Int -> remote int -> remote N7K
    for n7k in data:
        data[n7k]['P2P'] = p2p_n7k_mapping[n7k]

    return data


def get_sw_prof_name(dafe_file,leafid):
    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('leaf_switch_profile')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for x in range(row_start+1,row_end+1):
	cell = 'E' + str(x)
        id = ws[cell].value	

	if int(id) == int(leafid):
		cell =  'A' + str(x)
		prof = ws[cell].value
		break

    try:
	prof
    except NameError:
	print "Could not find switch profile for leaf ID " + str(leafid)
	prof = 'N/A'

    return prof
	

def fix_type_x(write_to_aci_cfg):
	vrf_type = {}
	for tenant in write_to_aci_cfg:
		for vrf in write_to_aci_cfg[tenant]:
			for epg in write_to_aci_cfg[tenant][vrf]:
				for e in write_to_aci_cfg[tenant][vrf][epg]:
					if e['type'] != 'X':
						tenant_vrf_type_key = tenant + "-" + vrf 
						vrf_type[tenant_vrf_type_key] = e['type']
						continue

	for tenant in write_to_aci_cfg:
                for vrf in write_to_aci_cfg[tenant]:
                        for epg in write_to_aci_cfg[tenant][vrf]:
                                for e in write_to_aci_cfg[tenant][vrf][epg]:
                                        if e['type'] == 'X':
						tenant_vrf_type_key = tenant + "-" + vrf
						e.update({'type' : vrf_type[tenant_vrf_type_key]})


def usage():
    print "Usage: " +  sys.argv[0] + " -d|--district <soe, gis or sde> -c|--datacenter <dc1 or dc2> -f|file <inputfile> -x <excludefile>"
    print ""
    print "-f|--file:   Pass input file to use for configuration.   Format:"
    print "1 value per line - assume its an EPG"
    print "6 to 10 values per line (depending on district) - assume by VRF: Tenant,VRF,P2P Subnet IP 1, P2P IP Subnet 2, etc"
    print "4 subnets for SDE, 8 subnets for GIS/SOE"
    print "-d|--district: indicates district name, must be <SOE|GIS|SDE>"
    print "-c|--datacenter: indicates datacenter name, must be <DC1 or DC2>"
    print "-x|--exclude: exclude EPGs from the migration.  File format expected is: tenant,vrf,epg (per line)"
    print "-h|--help: print help message"
    sys.exit(1)
    

def getValueWithMergeLookup(sheet, cell):
    for m_range in sheet.merged_cell_ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(m_range))
        for row in merged_cells:
            if cell in row:
                return sheet.cell(merged_cells[0][0]).value

def check_selector_exists(selector,pgname,dafe_file):

    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('leaf_interface_selector')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for x in range(row_start+1,row_end+1):
            cell = 'A' + str(x)
            int_selector = ws[cell].value
            
	    cell = 'D' + str(x)
            pg_path = ws[cell].value

            if int_selector == selector and pg_path == 'uni/infra/funcprof/accbundle-' + pgname:
                return True

    return False

    
def check_pcpg_name(pgname,dafe_file):

    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('pc_vpc_interface_policy_group')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for x in range(row_start+1,row_end+1):
            cell = 'A' + str(x)
            name = ws[cell].value

    	    if name == pgname:
		return True

    return False	

def get_pc_params(dafe_file):
			
    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('pc_vpc_interface_policy_group')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for x in range(row_start+1,row_end+1):
            cell = 'A' + str(x)
            name = ws[cell].value

	    if bool((re.search('^pc_',name,re.IGNORECASE))) and bool((re.search('downlink',name,re.IGNORECASE))):
		cell = 'E' + str(x)
                aepname = ws[cell].value

		aepname = aepname.replace("uni/infra/attentp-","")

		cell = 'F' + str(x)
                linkpolname = ws[cell].value

		cell = 'G' + str(x)
                cdppolname = ws[cell].value
		
		cell = 'H' + str(x)
                lldpolname = ws[cell].value
		
		cell = 'I' + str(x)
                stpolname = ws[cell].value
		
		cell = 'J' + str(x)
                lacpolname = ws[cell].value
		
		cell = 'M' + str(x)
                mcpolname = ws[cell].value
		
		try:
        		aepname
    		except NameError:
			aepname = 'N/A'
        		print "ERROR: Could not get AEP name for PC policy group creation"
		
		try:
        		linkpolname
    		except NameError:
			linkpolname = 'N/A'
        		print "ERROR: Could not get link policy name for PC policy group creation"
		
		try:
        		cdppolname
    		except NameError:
			cdppolname = 'N/A'
        		print "ERROR: Could not get CDP policy name for PC policy group creation"
		
		try:
        		lldpolname
    		except NameError:
			lldpolname = 'N/A'
        		print "ERROR: Could not get LLDP policy name for PC policy group creation"
		
		try:
        		stpolname
    		except NameError:
        		stpolname = 'N/A'
        		print "ERROR: Could not get STP policy name for PC policy group creation"
		
		try:
        		lacpolname
    		except NameError:
        		lacpolname = 'N/A'
    		except NameError:
        		print "ERROR: Could not get LACP policy name for PC policy group creation"
		
		try:
        		mcpolname
    		except NameError:
        		mcpolname = 'N/A'
        		print "ERROR: Could not get MCP policy name for PC policy group creation"

		break

    return (aepname,linkpolname,cdppolname,lldpolname,stpolname,lacpolname,mcpolname)

def get_epg_from_vrf(dafe_file,vrfs):

    worksheets = []
    wb = openpyxl.load_workbook(dafe_file, data_only=True)
    epgs = []
    bds = []
    found = 0

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('bridge_domain')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for v in vrfs:
	vs = v.split(",")
	if len(vs) != 6 and len(vs) != 10:
		print "ERROR: Check input file, must be in format <tenant,vrf,p2p ips subnet1, p2p ip subnet 2, etc. Found %s" % v
		sys.exit(9)
        tenant = vs[0]
        vrf = vs[1]
        vrf = vrf.rstrip()
    	
	for x in range(row_start,row_end+1):
            cell = 'C' + str(x)
            tenantvalue = ws[cell].value

            cell = 'D' + str(x)
            vrfvalue = ws[cell].value
	    
            if tenantvalue == tenant and vrfvalue == vrf:
		cell = 'A' + str(x)
                bdvalue = ws[cell].value
		bds.append(bdvalue)
		found = 1
     
        if found == 0: 
        	print "ERROR: Could not find any BD for Tenant: %s, VRF: %s" % (tenant,vrf)
		sys.exit(9)
	else:
	    found = 0 
    
    wb.active = worksheets.index('epg')
    
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row
    
    for b in bds:
    	for x in range(row_start,row_end+1):
            cell = 'E' + str(x)
            bdvalue = ws[cell].value
	
	    if bdvalue == b:
		epgs.append(b)
		found = 1

	if found == 0:
		print "ERROR: Could not find any EPGs for BD: %s" % (bd)
                sys.exit(9)
	else:
		found = 0       

    return epgs

def get_vrf_to_fw(zones_vl_ip_file,dc,district):

    worksheets = []
    wb = openpyxl.load_workbook(zones_vl_ip_file, data_only=True)
    vrf_to_fw = {}
    fw_to_vrf = {}

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('SOE_SDE_GIS_VRF_RT_Definition')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for x in range(row_start,row_end+1):
        cell = 'A' + str(x)
        districtvalue = getValueWithMergeLookup(ws,cell)
        if districtvalue == district.upper():
    		cell = 'B' + str(x)
        	tenant = getValueWithMergeLookup(ws,cell)
		if tenant is None:
			tenant = ws[cell].value
		if tenant == 'Tn_NonAff':
			continue
    		cell = 'C' + str(x)
		vrf = ws[cell].value
		if vrf == 'N/A':
			continue
		if vrf is not None:
			cell = 'E' + str(x)
			firewall = ws[cell].value 

			cell = 'I' + str(x)
			vrfmember = ws[cell].value
		
			cell = 'P' + str(x)
			outerencap = getValueWithMergeLookup(ws,cell)
		
			if outerencap is None:
				outerencap = ws[cell].value

			if vrfmember == 'N/A':
				continue		
		
			if dc == 'dc2':
				firewall = firewall.replace('dc1','dc2')
			
			# One off - ACI config has Audit/DDT -Zones vlans and IPs has Audit/DAT.  Changing to what ACI has
			if tenant == 'Audit' and vrf == 'DAT':
				vrf = 'DDT'

			# Another One off - ACI config has Audit/DDT -Zones vlans and IPs has User Access.  Changing to what ACI has
			if tenant == 'User Access':
				tenant = 'User_Access'
	
                	if tenant in vrf_to_fw:
				vrf_to_fw[tenant][vrf] = {}
                        	vrf_to_fw[tenant][vrf] = { 'firewall' : firewall, 'encap' : outerencap, 'to_delete' : 0 }
			else:
				vrf_to_fw[tenant] = {}
				vrf_to_fw[tenant][vrf] = {}
               			vrf_to_fw[tenant][vrf] = { 'firewall' : firewall, 'encap' : outerencap, 'to_delete' : 0 }

			#print "District %s, Tenant %s, VRF %s, Firewall: %s" % (district,tenant,vrf, firewall)

    for tenant in vrf_to_fw:
	for vrf in vrf_to_fw[tenant]:
		firewall = vrf_to_fw[tenant][vrf]['firewall']
		outerencap = vrf_to_fw[tenant][vrf]['encap']
		if firewall in fw_to_vrf:
			fw_to_vrf[firewall].append({'tenant' : tenant, 'vrf' : vrf, 'to_delete' : 0 , 'encap' : outerencap})
		else:
			fw_to_vrf[firewall] = {}
			fw_to_vrf[firewall] =   [{ 'tenant' : tenant, 'vrf' : vrf, 'to_delete' : 0, 'encap' : outerencap  }]
    
    return vrf_to_fw,fw_to_vrf

def get_all_epg_from_dafe(tenant,vrf,file):

    worksheets = []
    bd = []
    epgl = []
    wb = openpyxl.load_workbook(file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()
    
    wb.active = worksheets.index('bridge_domain')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for x in range(row_start,row_end+1):
    	cell = 'C' + str(x)
        tenantvalue = ws[cell].value
    	
	cell = 'D' + str(x)
        vrfvalue = ws[cell].value
	
	cell = 'K' + str(x)
        l3routingvalue = ws[cell].value

	#if tenantvalue == tenant and vrfvalue == vrf and l3routingvalue == 'yes':
	if tenantvalue == tenant and vrfvalue == vrf:
    		cell = 'A' + str(x)
        	bdvalue = ws[cell].value
		bd.append(bdvalue)

    wb.active = worksheets.index('epg')

    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for x in range(row_start,row_end+1):
	cell = 'E' + str(x)
        bdvalue = ws[cell].value	

	if bdvalue in bd:
		cell = 'A' + str(x)
        	epgvalue = ws[cell].value	
		epgl.append(epgvalue)
    return epgl

def get_epg_type(epg):

    worksheets = []

    pattern = 'VRF_EPG_Counts*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
                if not bool(re.search('^~',name, re.IGNORECASE)):
                        filename = name
                        break

    wb = openpyxl.load_workbook(filename, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    if bool((re.search('sde',epg,re.IGNORECASE))):
                tab = 'SDE'
    if bool((re.search('gis',epg,re.IGNORECASE))):
                tab = 'GIS'
    if bool((re.search('soe',epg,re.IGNORECASE))):
                tab = 'SOE'

    wb.active = worksheets.index(tab)
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    for x in range(row_start,row_end+1):
                cell = 'C' + str(x)
                value = ws[cell].value
                if value == epg:
                        cell = 'D' + str(x)
		        type = ws[cell].value
			continue
     
    for x in range(row_start,row_end+1):
                cell = 'I' + str(x)
                value = ws[cell].value
                if value == epg:
                        cell = 'J' + str(x)
		        type = ws[cell].value
			continue

    wb.close()
    try:
	type
    except NameError:
	#print "WARNING: EPG %s not found in %s" % (epg,filename)
	type='X'
    
    return type

def get_data(filename,epgs,dc,district,p2psubnets):
    p2psubnetvals = {}

    for p in p2psubnets:
	p = p.rstrip()
        p = p.split(",")
        p_tenant = p[0]
 	p_vrf = p[1]
        p_subnets = p[2:]

	# Sort IP addresses in case they are not in order (low to high)
        new_list = []

	# Check for Valid IP address
	for element in p_subnets:
		try:
			IPNetwork(element)
		except AddrFormatError:
			print "Invalid Subnet address given for Tenant %s, VRF %s.  %s given" % (p_tenant,p_vrf,element)
			sys.exit(9)
  	
		t_ip = IPNetwork(element)
		tt_ip = str(t_ip) 
		tt_ip = re.sub('/.*',"", tt_ip)	
		
		if str(tt_ip) != str(t_ip.network):
			print "ERROR: Invalid Subnet address provided for Tenant %s, VRF %s.  %s given" % (p_tenant,p_vrf,element)
			sys.exit(9)
		else:
			new_list.append(IPNetwork(element))
	

	new_list.sort()
	cidr = cidr_merge(new_list)

	# Check for discontinuous networks
	if len(cidr) != 1:
		print "WARNING: Discontiguous subnets found for Tenant %s, VRF %s.  %s given" % (p_tenant,p_vrf,(', '.join(p_subnets)))
	p = []
	for ee in new_list:
		if ee.prefixlen != 30:
			print "WARNING: P2P IP for Tenant %s, VRF %s is not a /30.  %s given" % (p_tenant,p_vrf,ee.cidr.__str__())
			
		p.append(ee.cidr.__str__())

		
	
	p2psubnetvals[p_tenant + "-" + p_vrf] = {}
	p2psubnetvals[p_tenant + "-" + p_vrf] = p

    # Get Pre-Build data

    worksheets = []
    pre_build = {}

    pattern = 'DCT_' + district.upper() + '*Firewalls*Cabling*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
                if not bool(re.search('^~',name, re.IGNORECASE)):
                        cabling_file = name
                        break
    
    wb2 = openpyxl.load_workbook(cabling_file, data_only=True)

    for sheet in wb2:
        worksheets.append(sheet.title)
    wb2.close() 

    wb2.active = worksheets.index('P2P')

    ws2 = wb2.active
    row_start = ws2.min_row
    row_end   = ws2.max_row
    
    for x in range(row_start+1,row_end+1):
    	cell = 'A' + str(x)
        fwvalue = ws2[cell].value
	
	if fwvalue is not None:
		fw = fwvalue[:-1]
		fwa = fw + 'a'
		fwb = fw + 'b'

	cell = 'G' + str(x)
        vrfmember = ws2[cell].value 

	if vrfmember is not None:
		cell = 'B' + str(x)
        	vlan = ws2[cell].value
		z = vlan.split('.')
		vlan = z[1]
		
		cell = 'C' + str(x)
        	fwbdsubnet = ws2[cell].value
		
		cell = 'D' + str(x)
        	fwvip = ws2[cell].value
		
		cell = 'K' + str(x)
        	bdip = ws2[cell].value
	
		if vlan is None:
			print "WARNING: No VLAN Found for Security Zone %s" % vrfmember
		if not vlan.isnumeric():
			print "WARNING: VLAN ID not valid.  %s given." % vlan
		else:
			if int(vlan) > 4096:
				print "WARNING: Invalid VLAN range given.  %s given." % vlan

		try:
                    IP(fwvip)
                except:
                    print "WARNING: Firewall IP %s for Security Zone %s not valid" % (fwvip,vrfmember)

		try:
                    IP(bdip)
                except:
                    print "WARNING: New FW Bridge Domain IP %s for Security Zone %s not valid" % (bdip,vrfmember)
		
		try:
                    IP(fwbdsubnet)
                except:
                    print "WARNING: New FW Bridge Domain Subnet address %s for Security Zone %s not valid" % (fwbdsubnet,vrfmember)
	
		pre_build[vrfmember] = ({
					'fwa': fwa,
					'fwb': fwb,
					'vlan' : vlan,
					'fwvip' : fwvip,
					'fwbdsubnet' : fwbdsubnet,
					'bdip' : bdip,
					'leafa' : 'N/A',
					'leafb' : 'N/A',
					'leafa_int' : [],
					'leafb_int' : [],
					})

		
    worksheets = []

    pattern = 'DCT*FIrewalls_Cisco*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
                if not bool(re.search('^~',name, re.IGNORECASE)):
                        fwfile = name
                        break

    wb2 = openpyxl.load_workbook(fwfile, data_only=True)

    for sheet in wb2:
        worksheets.append(sheet.title)
    wb2.close()

    wb2.active = worksheets.index(dc.upper())

    ws2 = wb2.active
    row_start = ws2.min_row
    row_end   = 21
   
    for x in range(row_start+1,row_end+1):
	intf_range = []
        cell = 'A' + str(x)
        fwvalue = ws2[cell].value
	cell = 'H' + str(x)
        leaf = ws2[cell].value
        
	cell = 'J' + str(x)
        leaf_int = ws2[cell].value
	i = leaf_int.split(' - ')
	low = i[0]
        high = i[1]
	lowa = low.split("/")
	higha = high.split("/")
	for ii in range(int(lowa[1]),int(higha[1])+1):
		intf_range.append(str(lowa[0]) + "/" + str(ii))

	for vrfmember in pre_build:
		if pre_build[vrfmember]['fwa'] == fwvalue:
			pre_build[vrfmember]['leafa'] = leaf
			pre_build[vrfmember]['leafa_int'] = intf_range
		if pre_build[vrfmember]['fwb'] == fwvalue:
			pre_build[vrfmember]['leafb'] = leaf
			pre_build[vrfmember]['leafb_int'] = intf_range
    

    ########

    worksheets = []
    write_to_aci_cfg = {}

    wb = openpyxl.load_workbook(filename, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title) 
    wb.close()

    for epg in epgs:
	epg = epg.rstrip()
    	# From EPG tab, get bridge domain name
    	wb.active = worksheets.index('epg')

    	ws = wb.active
    	row_start = ws.min_row
    	row_end   = ws.max_row
        
    	# Get Bridge Domain
    	for x in range(row_start,row_end+1):
        	cell = 'A' + str(x)
        	value = ws[cell].value 
		if value == epg:
			bdcell = 'E' + str(x)
			bd = ws[bdcell].value
			
			apcell = 'D' + str(x)
			ap = ws[apcell].value
			continue

	try:
		bd
	except NameError:
		print "ERROR: BD Not found for EPG %s.  Please check dafe output" % epg
	        sys.exit(9)
		continue
	
	try:
		ap
	except NameError:
		print "ERROR: AppProfile Not found for EPG %s.  Please check dafe output" % epg
	        sys.exit(9)
		continue

			
    	#From bridge domain tab, get tenant, vrf and L2/L3 Y or N
    	wb.active = worksheets.index('bridge_domain')
	ws = wb.active

	row_start = ws.min_row
        row_end   = ws.max_row
        
        for x in range(row_start,row_end+1):
                cell = 'A' + str(x)
                value = ws[cell].value
                if value == bd:
                        tenant        = ws['C' + str(x)].value
                        vrf           = ws['D' + str(x)].value
                        l3routing     = ws['K' + str(x)].value
                        continue
	
        try:
                tenant
        except NameError:
                print "ERROR: tenant Not found for EPG %s.  Check dafe file" % epg
		sys.exit(9)
        
	try:
                vrf
        except NameError:
                print "ERROR: vrf Not found for EPG %s.  Check dafe file" % epg
		sys.exit(9)

	try:
               l3routing 
        except NameError:
                print "ERROR: l3routing Not found for EPG %s.  Check dafe file" % epg
		sys.exit(9)


	# Get BD Subnet
	if l3routing == 'yes':
		wb.active = worksheets.index('bd_subnet')

        	ws = wb.active
        	row_start = ws.min_row
        	row_end   = ws.max_row

        	for x in range(row_start,row_end+1):
                	cell = 'C' + str(x)
                	bdvalue = ws[cell].value
                
			cell = 'D' + str(x)
                	tenantvalue = ws[cell].value
                
			if bdvalue == bd and tenantvalue == tenant:
                        	bdsubnetcell = 'A' + str(x)
                        	bd_subnet = ws[bdsubnetcell].value

        	try:
                	bd_subnet
        	except NameError:
                	print "WARNING: %s, BD Subnet Not found - unicast routing enabled" % epg
                	bd_subnet = 'N/A'
	else:
		bd_subnet = 'N/A'

 
    	#From vrf tab, get vrf member name, using tenant and vrf from bridge_domain tab
    	# Can use this for naming other constructs and getting L3OUT name
	wb.active = worksheets.index('vrf')
        ws = wb.active

        row_start = ws.min_row
        row_end   = ws.max_row

        for x in range(row_start,row_end+1):
                cell = 'A' + str(x)
                vrfvalue = ws[cell].value
                
		cell = 'B' + str(x)
                tenantvalue = ws[cell].value
                if vrfvalue == vrf and tenantvalue == tenant:
                        vrfmember = ws['H' + str(x)].value
                        continue

        try:
                vrfmember
        except NameError:
                print "ERROR: vrf member name Not found for EPG %s.  Check dafe file" % epg
		sys.exit(9)
  

	# Get L3Out name, using vrf member pre-pended with 'L3Out' - this is a good assumption
	# updated to use tab bd_l3out to get L3out
	wb.active = worksheets.index('bd_l3out')
        ws = wb.active

        row_start = ws.min_row
        row_end   = ws.max_row

        for x in range(row_start,row_end+1):
                cell = 'A' + str(x)
                bdvalue = ws[cell].value

                cell = 'B' + str(x)
                tenantvalue = ws[cell].value

                if bdvalue == bd and tenantvalue == tenant:
			cell = 'C' + str(x)
			l3out = ws[cell].value

        try:
                l3out
        except NameError:
                #print "ERROR: l3out name Not found for EPG %s.  Check dafe file" % epg
		l3out = 'N/A'

	 
    	# Get EPG consumed contracts.  For the EPG, put contracts into an array
	wb.active = worksheets.index('epg_consumed_contract')
        ws = wb.active
	epg_c_contracts = []

        row_start = ws.min_row
        row_end   = ws.max_row

        for x in range(row_start,row_end+1):
                cell = 'C' + str(x)
                epgvalue = ws[cell].value

                if epgvalue == epg :
                        cell = 'D' + str(x)
                        contract = ws[cell].value
			epg_c_contracts.append(contract)

	if len(epg_c_contracts) == 0 and l3routing == 'yes':
		print "WARNING: %s, No consumed contracts assigned" % epg

    	# Get EPG provided contracts.  For the EPG, put contracts into an array

	wb.active = worksheets.index('epg_provide_contract')
        ws = wb.active
        epg_p_contracts = []

        row_start = ws.min_row
        row_end   = ws.max_row

        for x in range(row_start,row_end+1):
                cell = 'C' + str(x)
                epgvalue = ws[cell].value

                if epgvalue == epg :
                        cell = 'D' + str(x)
                        contract = ws[cell].value
                        epg_p_contracts.append(contract)

        if len(epg_p_contracts) == 0 and l3routing == 'yes':
                print "WARNING: %s, No provide contracts assigned" % epg	 

    	# External EPG name from L3Out name along with provided and consumed contract, which have to be split by comma
	wb.active = worksheets.index('external_epg')
        ws = wb.active
        l3out_c_contracts = []
        l3out_p_contracts = []

        row_start = ws.min_row
        row_end   = ws.max_row

	# Get External EPG Name
        for x in range(row_start,row_end+1):
                cell = 'B' + str(x)
                l3outvalue = ws[cell].value
                
		cell = 'A' + str(x)
                tenantvalue = ws[cell].value

                if l3outvalue == l3out and tenantvalue == tenant:
                        cell = 'D' + str(x)
                        externalepg = ws[cell].value

        try:
                externalepg
        except NameError:
                externalepg = 'N/A'
		#print "ERROR: external EPG name Not found for EPG %s.  Check dafe file" % epg
		#sys.exit(9)

	# Now get contracts on the external EPG
        for x in range(row_start + 1,row_end+1):
                cell = 'B' + str(x)
                l3outvalue = str(ws[cell].value)
		if l3outvalue == l3out:
                	cell = 'G' + str(x)
                	ccontracts = str(ws[cell].value)
			l3out_c_contracts = ccontracts.split(",")
                	cell = 'I' + str(x)
                	pcontracts = str(ws[cell].value)
			l3out_p_contracts = pcontracts.split(",")
			continue

	# Got all the contract info for the EPG, now identify the ones to remove or are an issue:
	# On EPG, if contract is on EPG prov/cons and in L3out as provider and consumer, list it for removal
	contracts_to_remove = []
	for c in epg_c_contracts:
		if c in l3out_c_contracts and c in l3out_p_contracts and c in epg_p_contracts:
			contracts_to_remove.append(c)

		if l3out == 'N/A' and l3routing == 'yes':
			print "WARNING: %s, is part of BD %s that is not assigned to an L3 Out.  Please confirm if this EPG should be included" % (epg,bd)
	
		# On EPG, if on consumer but not provider, print warning
		if c not in epg_p_contracts:
			print "WARNING: %s, Contract %s not assigned to EPG as provider" % (epg,c)
		
                # If contract is not on L3out as provider and consumer, print warning
                if c not in l3out_p_contracts and c not in l3out_c_contracts and externalepg != 'N/A':
                        print "WARNING: %s, Consumed Contract %s not assigned to external EPG %s as consumer or provider" % (epg,c,externalepg)
		
		# On EPG, if on consumer but only on either provider or consumer of L3out, print warning
		if c not in l3out_c_contracts and c in l3out_p_contracts and externalepg != 'N/A':
			print "WARNING: %s, Contract %s not assigned to external EPG %s as consumer, but assigned as provider" % (epg,c,externalepg)
		if c in l3out_c_contracts and c not in l3out_p_contracts:
			print "WARNING: %s, Contract %s assigned to external EPG %s as consumer, but not assigned as provider" % (epg,c,externalepg)

	# Repeat checks for EPG provider
	# no need to check if EPG provider contract exists on EPG consumer and l3out as provider/consumer - done already
	for p in epg_p_contracts:
        
                # On EPG, if on provider but not consumer, print warning
                if p not in epg_c_contracts:
                        print "WARNING: %s, Contract %s not assigned as consumer" % (epg,p)
                
                # If contract is not on L3out as provider and consumer, print warning
                if p not in l3out_p_contracts and p not in l3out_c_contracts and externalepg != 'N/A':
                        print "WARNING: %s, Provided Contract %s not assigned to external EPG %s as consumer or provider" % (epg,p,externalepg)
        
                # On EPG, if on provider but only on either provider or consumer of L3out, print warning
                if p not in l3out_c_contracts and p in l3out_p_contracts and externalepg != 'N/A':
                        print "WARNING: %s, Contract %s not assigned to external EPG %s as consumer but assigned as provider" % (epg,p,externalepg)
                if p in l3out_c_contracts and p not in l3out_p_contracts:
                        print "WARNING: %s, Contract %s assigned to external EPG %s as consumer but not assigned as provider" % (epg,p,externalepg)

        # Get  VRF to EPG list - will check to make sure we got it all
        # if the argument passed to the script is by VRF, no need to check, we got it all and inner N7K config can be removed

	# Get Type A or Type B.  Look at file VRF_EPG_Counts	
	type = get_epg_type(epg)	

	# Write to dictionary for printing config files
	if tenant not in write_to_aci_cfg:
		write_to_aci_cfg[tenant] = {}
	if vrf not in write_to_aci_cfg[tenant]:
		write_to_aci_cfg[tenant][vrf] = {}

	write_to_aci_cfg[tenant][vrf][epg] = {}

	# One off - rename AUD-DDT to AUD-DAT temporarily
	if bool((re.search('AUD-DDT',vrfmember,re.IGNORECASE))):
		vrfmember = vrfmember.replace('DDT','DAT')

	# One off - rename DMZ-DVT-DC[1 or2]-SDE  to DMZ-WEB-DC[1 or 2]-SDE-CELL1 temporarily
	if vrfmember == 'DMZ-DVT-' + dc.upper()+ '-SDE':
		vrfmember = 'DMZ-WEB-' + dc.upper() + '-SDE-CELL1'
	
        if vrfmember in pre_build:
		vlan = pre_build[vrfmember]['vlan']
		fwvip = pre_build[vrfmember]['fwvip']
		fwbdip = pre_build[vrfmember]['bdip']
		fwbdsubnet = pre_build[vrfmember]['fwbdsubnet']
		leafa = pre_build[vrfmember]['leafa']
		leafb = pre_build[vrfmember]['leafb']
		fwa = pre_build[vrfmember]['fwa']
		fwb = pre_build[vrfmember]['fwb']
		leafa_int = pre_build[vrfmember]['leafa_int']
		leafb_int = pre_build[vrfmember]['leafb_int']

	else:
		print "WARNING: Could not find pre build work for VRF Member %s " % vrfmember
		vlan = 'N/A'
		fwvip = 'N/A'
		fwbdip = 'N/A'
		fwbdsubnet = 'N/A'
		leafa = 'N/A'
		leafb = 'N/A'
		fwa = 'N/A'
		fwb = 'N/A'
		leafa_int = 'N/A'
		leafb_int = 'N/A' 

	if bool((re.search('AUD-DAT',vrfmember,re.IGNORECASE))):
		vrfmember = vrfmember.replace('DAT','DDT')
	
	# Change back the One off
	if vrfmember == 'DMZ-WEB-' + dc.upper() + '-SDE-CELL1':
		vrfmember = 'DMZ-DVT-' + dc.upper() + '-SDE'
	
	write_to_aci_cfg[tenant][vrf][epg] = [{

						'bd' : bd,
						'l3' : l3routing,
						'vrfmember' : vrfmember,
						'l3out' : l3out,
						'extepg' : externalepg,
						'type' : type,
						'ap'   : ap,
						'fwbdsubnet' : fwbdsubnet,
						'vrf'  : vrf,
						'bd_subnet' : bd_subnet,
						'remove_l3_contract' : 'yes',
						'contract' : contracts_to_remove,
						'curr_c_contracts' : epg_c_contracts,
						'curr_p_contracts' : epg_p_contracts,
						'vlan' : vlan,
						'fwbdname' : 'N/A',
						'fwvip' : fwvip,
						'fwbdip' : fwbdip,
						'leafa' : leafa,
						'leafb' : leafb,
						'leafa_int' : leafa_int,
						'leafb_int' : leafb_int,
						'fwaname'    : fwa, 
						'fwbname'    : fwb,
						'l4l7'       : 'N/A',
						'sgtname'       : 'N/A',
						'pbrname'       : 'N/A',
						'sgcontractname'       : 'N/A',
						'fwcluster' : 'N/A',
						'p2psubnets' :  p2psubnetvals[tenant + "-" + vrf]
				     	      }]



	"""	
	print "Tenant: %s" % tenant	
	print "EPG: %s" % epg
	print "BD: %s" % bd
	print "VRF: %s" % vrf
	print "L3 BD: %s" % l3routing
	print "VRF Member: %s" % vrfmember
	print "L3Out: %s" % l3out
	print "External EPG: %s" % externalepg
	print "Type: %s" % type
	print "Contracts to remove:"
	print contracts_to_remove
	print "********"
	print 
	print 
	print
        """ 
        del tenant
	del epg	
	del bd
	del vrf
	del l3routing
	del vrfmember
	del l3out
	del epg_c_contracts
	del epg_p_contracts
	del externalepg
	del l3out_c_contracts
	del l3out_p_contracts
	del contracts_to_remove
	del type 
	del vlan
        del fwvip
        del fwbdip
        del fwbdsubnet
        del leafa
        del leafb
        del fwa
        del fwb
        del leafa_int
        del leafb_int
	del bd_subnet

   
    return write_to_aci_cfg

def main(argv):
    dir_path = './output'
    
    if os.path.isdir(dir_path):
	shutil.rmtree(dir_path)
   
    os.mkdir(dir_path)

    toexclude = False
    
    # Arguments
    if len(argv) == 0:
        usage()
    
    try:
        opts,args = getopt.getopt(argv,"d:c:f:hx:",["district=","datacenter=","file=","help","--exclude"])
    except getopt.GetoptError as err:
        print str(err)
        sys.exit(2)
    else:
        for opt,arg in opts:
            if opt in ("-h","--help"):
                usage()
            if opt in ("-d","--district"):
                district = arg
            if opt in ("-c","--datacenter"):
                dc = arg
            if opt in ("-f","--file"):
                infile = arg
            if opt in ("-x","--exclude"):
                excludefile = arg
		toexclude = True
    try:
	district
    except NameError:
	print "District option not passed (-d|--district)"
        sys.exit(9)
    
    try:
	infile	
    except NameError:
	print "Input file not passed (-f|--file)"
        sys.exit(9)
    
    try:
	dc
    except NameError:
	print "Datacenter option not passed (-c|--datacenter)"
        sys.exit(9)
	    
    if district.upper() not in ('SDE','GIS','SOE'):
	print "ERROR: District must be SOE, GIS, SDE - one of these values"
	sys.exit(9)
    
    if dc.upper() not in ('DC1','DC2'):
	print "ERROR: DataCenter must be DC1 or DC2 - one of these values"
	sys.exit(9)

    if not os.path.isfile(infile):
        print sys.argv[0] + " Input File %s NOT found" % infile
        sys.exit(9)
   
    if toexclude is True: 
    	if os.path.isfile(excludefile):
        	with open (excludefile) as ex:
			exinfo = ex.read().splitlines()
	else:
		print "ERROR: Exclude option chosen, but exclude file not found"
		sys.exit(9)
			
    dafe_file = dc.upper() + "_" + district.upper() + "_DAFE.xlsx"

    with open (infile) as f:
	epgs = f.readlines()
	numparams = epgs[0].split(",")

    # If theres no comma in the input, its by EPG. 
    # if theres at least 1 comma, its by VRF
    # Deprecated - it's only bt VRF now

    if len(numparams) == 1:
    	with open (infile) as f:
		epgs = f.readlines()

    if len(numparams) == 6 and district.upper() == 'SDE':
	with open (infile) as f:
		vrfs = f.readlines()
		epgs = get_epg_from_vrf(dafe_file,vrfs)	
    
    if len(numparams) != 6 and district.upper() == 'SDE':
    	print "ERROR: Check input file. There must be at least 6 parameters (tenant, vrf, P2P Subnet 1, P2P Subnet 2, etc"
	sys.exit(9)
 
    if len(numparams) == 10 and (district.upper == 'SOE' or district.upper == 'GIS' ) :
	with open (infile) as f:
		vrfs = f.readlines()
		epgs = get_epg_from_vrf(dafe_file,vrfs)	
    
    if len(numparams) != 10 and ( district.upper == 'SOE' or district.upper == 'GIS' ) :
    	print "ERROR: Check input file. There must be at least 10 parameters (tenant, vrf, P2P Subnet 1, P2P Subnet 2, etc"
	sys.exit(9)


    # Get N7K Data - to be used after ACI configs built
    n7k_data = get_bgp_int_vlan(dc,district,vrfs)


    (aepname,linkpolname,cdppolname,lldpolname,stpolname,lacpolname,mcpolname) = get_pc_params(dafe_file)
    write_to_aci_cfg = get_data(dafe_file,epgs,dc,district,vrfs)

    # Print out pre-migration planning info
    migration_planning_file = './output/PRE_MIGRATION_PLANNING.txt'
    f = open(migration_planning_file,"a")

    for tenant in write_to_aci_cfg:
	for vrf in write_to_aci_cfg[tenant]:
		ecount = 0
		nomigr = []
		for epg in write_to_aci_cfg[tenant][vrf]:
			for e in write_to_aci_cfg[tenant][vrf][epg]:
				bdip = e['bd_subnet']
				isl3 = e['l3']
				fw = e['fwaname']
				vrftype = e['type']

				if isl3 == 'yes':
					f.write(epg + ',' + bdip + ',' + fw + ',' + vrftype + ',' + vrf + ',' + tenant + ',' + '\n')
					ecount = ecount + 1 
				else:
					nomigr.append(epg)
		f.write('\n' + "Number of EPGs migrated: " + str(ecount) + '\n')	
		f.write("Number of L2 EPGs not migrated: " + str(len(nomigr)) + '\n' )
		for n in nomigr:
			f.write(n + '\n')
		f.write('\n')
		f.write('*' * 12 + '\n')
    f.close()	

    pattern = 'Zones*Vlans*IPs*'

    files = os.listdir('.')
    for name in files:
        if fnmatch.fnmatch(name, pattern):
                if not bool(re.search('^~',name, re.IGNORECASE)):
                        zvlipfile = name
                        break

    vrf_to_fw,fw_to_vrf = get_vrf_to_fw(zvlipfile,dc,district)
 
    # Checks
    vrfstodelete = []
    # 1. all EPGs migrated in VRF for the input file? Y or N - Will determine if the n7K inner config gets modified
    for tenant in write_to_aci_cfg:
    	missing_from_input_file = []
    	missing_from_dafe = []
    	epglist = []
	
	for vrf in write_to_aci_cfg[tenant]:
		
		for e in write_to_aci_cfg[tenant][vrf]:
			epglist.append(e)

	
		# Get EPG list from DAFE output
        	dafe_epg = get_all_epg_from_dafe(tenant,vrf,dafe_file)
      
		for e in dafe_epg:
			if e not in epglist:
				missing_from_input_file.append(e)
	
		for e in epglist:
			if e not in dafe_epg:
				missing_from_dafe.append(e)

		if len(missing_from_dafe) > 0:
			print "ERROR: The following EPGs provided are not in tenant %s, vrf %s" % ( tenant, vrf )
			for e in missing_from_dafe:
				print e

		if len(missing_from_input_file) > 0:
    			f = open(migration_planning_file,"a")
			f.write("WARNING: The following EPGs in tenant %s, vrf %s are not being migrated.  Inner and outer VDC config should not be modified and contract on L3Out should not be removed" % ( tenant, vrf ) + '\n' )
			for e in missing_from_input_file:
				f.write(e + '\n')
			f.close()

			# find out which tenant this EPG is in and set all the EPGs in that tenant to remove_l3_contract to no
			for tenant in write_to_aci_cfg:
				for vrf in write_to_aci_cfg[tenant]:
					for ep in write_to_aci_cfg[tenant][vrf]:
				 		for a in write_to_aci_cfg[tenant][vrf][ep]:
							 a['remove_l3_contract'] = 'no'	
					
		else:
			# All inner is being migrated, remove the tenant/VRF from the list of fw_to_epg for the outer, then check this later on to see if all VRFs are removed from FW
			# Write N7K Config here (Inner and outer and shutdown inner config)
			# Still have to check if we can shutdown outer config - done later
			# Check for cleanup as well - later
    			f = open(migration_planning_file,"a")
			f.write("Inner config can be removed for tenant %s, vrf %s" % (tenant,vrf) + '\n' )
			f.close()
			
			for ep in write_to_aci_cfg[tenant][vrf]:
					vrfmember = write_to_aci_cfg[tenant][vrf][ep][0]['vrfmember']
					p2psubnets =  write_to_aci_cfg[tenant][vrf][ep][0]['p2psubnets']
					break
			vrf_to_fw[tenant][vrf]['to_delete'] = 1
			for n7k in n7k_data:
				if bool((re.search('inner',n7k,re.IGNORECASE))) :	
					n7k_data[n7k][vrfmember]['shutdown'] = 'Y'
					vrfstodelete.append(vrfmember)
					continue

			write_new_n7k_configs(vrfmember,p2psubnets,dc,district,n7k_data)

		epglist = []
        	dafe_epg = []
		missing_from_dafe = []
		missing_from_input_file = []
    
   
    f = open('n7k_data.json', 'w')
    f.write(json.dumps(n7k_data) )
    f.close()

    # See what's left in the outer
    # If all the VRFs are being removed from the FW, shutdown the SVI on the outer and remove the VLAN from the trunk on inner and outer
    stillexist = []   
    targeted = 0 
    count  = {}
    
    cutover_dir = dir_path + "/" + "N7K_CUTOVER" 
    cleanup_dir = dir_path + "/" + "N7K_NEXT_CLEANUP" 
   
    for n7k in n7k_data:
    	if bool((re.search('outer',n7k,re.IGNORECASE))) :
		for svi in n7k_data[n7k]:
			if svi == 'P2P' :
				continue
			if 'vrf' not in n7k_data[n7k][svi]:
				continue
			for vrfs in n7k_data[n7k][svi]['vrf']:
				if vrfs not in vrfstodelete:
					stillexist.append(vrfs)
				else:
					targeted = 1

			# Write SVI shutdown and VLAN removal to the VRF cutover file which has the most EPG's in the group of VRFs on the trunk - this will most likely be targeted last during the migration
			# In the inner, remove the VLAN from the FW trunk for every VRF	
			# Get EPG count
			for v in n7k_data[n7k][svi]['vrf']:
				for tenant in write_to_aci_cfg:
					for vvvrf in write_to_aci_cfg[tenant]:
						for epg in write_to_aci_cfg[tenant][vvvrf]:
							if write_to_aci_cfg[tenant][vvvrf][epg][0]['vrfmember'] == v:
								num = len(write_to_aci_cfg[tenant][vvvrf])
								count[v] = num
								break;
			if len(stillexist) == 0 and targeted == 1:
				sorted_d = sorted(count.items(), key=operator.itemgetter(1))
				fmigr = open(migration_planning_file,"a")
				fmigr.write("Outer encap " + svi + " can be removed on " + n7k + "." + "  SVI shutdown config will be written to: " + sorted_d[-1][0] + " because it has the most EPGs in this VRF/SVI" + '\n')
				fmigr.close()
				f = open(cutover_dir + "/" + sorted_d[-1][0] + "/" + n7k, "a")
				f.write("!! Shutdown VLAN and remove VLAN from firewall - All VRFs migrated " + '\n')
				f.close()	

			if len(stillexist) != 0 and targeted == 1:
				sorted_d = sorted(count.items(), key=operator.itemgetter(1))
				fmigr = open(migration_planning_file,"a")
				fmigr.write("WARNING: Outer encap " + svi + " on " + n7k + " has VRFs, but SVI shutdown config will be written to " + sorted_d[-1][0] + ".  Please remove this config if not needed. The following VRFs still exist:" + ','.join(stillexist) + '\n' )
				fmigr.close()
				f = open(cutover_dir + "/" + sorted_d[-1][0] + "/" + n7k, "a")
				f.write("!! Verify if the SVI shutdown and firewall config VLAN removal should be executed. VRFs " + ','.join(stillexist) + " still exist on this VLAN per config"  + '\n')
				f.close()
	
			if targeted == 1:
				fwints = n7k_data[n7k][svi]['fw_trunk_int']
				f = open(cutover_dir + "/" + sorted_d[-1][0] + "/" + n7k, "a")
				f.write("interface Vlan" + svi + '\n')
				f.write(" shutdown" + '\n')
				f.close()
				f = open(cleanup_dir + "/" + sorted_d[-1][0] + "/" + n7k, "a")
				f.write("!!Remove VLAN Interface for VRF " + sorted_d[-1][0]  + '\n')
				f.write("no interface Vlan" + svi + '\n')
				f.write('\n')
				for fw in fwints:
					f.write("!!Remove VLAN from FW trunk for VRF " + sorted_d[-1][0]  + '\n')
					f.write("interface Ethernet" + fw + '\n')
					f.write(" switchport trunk allowed vlan remove " + svi + '\n')
					f.write('\n')
				f.close()
					
			count = {}
			stillexist = []
			targeted = 0
		
 
    # Write ACI configs
    # Constraints
    # Can't remove contract from L3Out if all EPG's aren't being migrated

    # Fix EPGs/VRFs that have type = 'X'
    fix_type_x(write_to_aci_cfg)

    # Associate leaf interfaces to PC and create leaf selectors.  Already done, not needed for 3/13.  Again check previous files

    # Check if output dir exists - if not, create it
    # if exists, delete it and re-create it

    os.mkdir(dir_path + "/" + "ACI_PRE_WORK")
    os.mkdir(dir_path + "/" + "ACI_CONTRACT_VERIFICATION")
   

    # Pre work
    # Create port channel  and interface selectors
    pc_creation = {} 
    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
                for epg in write_to_aci_cfg[tenant][vrf]:
                        for d in write_to_aci_cfg[tenant][vrf][epg]: 
				pca_key = d['fwaname']
				if d['fwaname'] == 'N/A':
					print "WARNING: No firewall information found for EPG: " + epg
					continue
				pca_key = 'pc_' + pca_key + ":" + aepname + ":" + linkpolname + ":" + cdppolname + ":" + lldpolname + ":" + stpolname + ":" + lacpolname + ":" + mcpolname
				
				pcb_key = d['fwbname']
				pcb_key = 'pc_' + pcb_key + ":" + aepname + ":" + linkpolname + ":" + cdppolname + ":" + lldpolname + ":" + stpolname + ":" + lacpolname + ":" + mcpolname
				if pca_key in pc_creation:
					continue
				else:
					pc_creation[pca_key] = { 'leafid' : d['leafa'], 'intf' : d['leafa_int']}
				
				if pcb_key in pc_creation:
					continue
				else:
					pc_creation[pcb_key] = {'leafid' : d['leafb'], 'intf' : d['leafb_int']}


    f = open(dir_path + "/" + "ACI_PRE_WORK/3.6  - Create port channel interface policy group for the new firewall connections.csv", "a") 
    f.write("PC_PG,LINK,CDP,MCP,LLDP,BPDU,LACP,AEP" + '\n')
    f.close()
    
    f = open(dir_path + "/ACI_PRE_WORK/3.7 - Create int selectors and associate leaf interfaces to port channel interface policy group.csv", "a") 
    f.write("INTPROFILE,PC_POLICY,PORT,PN" + '\n')
    f.close()
    
    #pc_creation['pc_dc1sdenwa1sbx01:AEP_STATIC:10GB_Auto:CDP_ENABLE:LLDP_ENABLE:BPDU_GUARD_ENABLED:LACP_ACTIVE:MCP_ENABLE'] =  {"leafid": 212, "intf": ["1/27", "1/28", "1/29", "1/17"]}

    for pc_key in pc_creation:
	tmp = pc_key.split(':')
	pgname = tmp[0]
	link = tmp[2]
	cdppol = tmp[3]
	mcp_pol = tmp[7]
	lldppol = tmp[4]
	stppol = tmp[5]
   	lacppol = tmp[6]
	aeppol = tmp[1]
	leafid = pc_creation[pc_key]['leafid']
	intf = pc_creation[pc_key]['intf']

	swprofname = get_sw_prof_name(dafe_file,leafid)

	# if port channel policy group is created, skip creating it
	pgname_exists = check_pcpg_name(pgname,dafe_file)
	if pgname_exists is False:
        	f = open(dir_path + "/ACI_PRE_WORK/3.6  - Create port channel interface policy group for the new firewall connections.csv", "a") 
    		f.write(pgname + "," + link + "," + cdppol + "," + mcp_pol + "," + lldppol + "," + stppol + "," + lacppol + "," + aeppol + '\n')
    		f.close()
	else:
		print "OK: Port channel policy group name " + pgname + " exists, not creating it"
        
	f = open(dir_path + "/ACI_PRE_WORK/3.7 - Create int selectors and associate leaf interfaces to port channel interface policy group.csv", "a") 
	for n in intf:
		i = n.split("/")
		portnum = i[1]
		selector = "E_" + i[0] + "_" + portnum
		selector_exists = check_selector_exists(selector,pgname,dafe_file)

		if selector_exists is False:
    			f.write(swprofname + "," +  pgname + "," + selector + "," + portnum + '\n')
		else:
			print "OK: Interface selector " + selector + " already exists in policy group " + pgname + ".  Selector will not be created"
    	f.close()

    # 3.8 to 3.14 - Create new construct names, save it and access it later 
   
    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
		for epg in write_to_aci_cfg[tenant][vrf]:
			for d in write_to_aci_cfg[tenant][vrf][epg]:
				tshortname = epg.split("-")
				tenantshort = tshortname[0]
				shortfirewall = d['fwaname']
				shortfirewall = shortfirewall[:-1]
				fwbdname = tenantshort + "-" + vrf + "-BD-" + dc.upper() + "-" + district.upper() + "-FW"
				if bool((re.search('C-',fwbdname,re.IGNORECASE))) :
					fwbdname = fwbdname.replace("C-","C_")
			
				d.update({'fwbdname' : fwbdname})

				# Create L4L7 name and update dictionary
				l4l7name = tenantshort + "-" + vrf + "-L4L7-" + shortfirewall.upper()
				d.update({'l4l7' : l4l7name})

				# Create service graph template name and update dictionary
				sgtname = tenantshort + "-" + vrf + "-SG-" + dc.upper() + "-" + district.upper() 
				d.update({'sgtname' : sgtname})

				# Create PBR name and update dictionary
				pbrname = tenantshort + "-" + vrf + "-PBR-" + shortfirewall.upper()
				d.update({'pbrname' : pbrname})

				# Create contract name and update dictionary
				sgcontractname = tenantshort + "-" + vrf + "-SG-PBR-Permit_Any"
				d.update({'sgcontractname' : sgcontractname})
				
				# update FW cluster name
				d.update({'fwcluster' : shortfirewall.upper()})

    # Prepare all the headers for the pre-work files
    f = open(dir_path + "/ACI_PRE_WORK/3.8 - Create a bridge domain for the policy-based routing policy.csv", "a")
    f.write("TENANT,BD,SUBNET,VRF" + '\n')
    f.close()

    f = open(dir_path + "/ACI_PRE_WORK/3.9 - Create the L4-L7 device.csv", "a")
    f.write("TENANT,FW,DEVICE1,LEAF_ID1,PC_NAME1,DEVICE2,LEAF_ID2,PC_NAME2,PHYS_DOMAIN,VLANID,CLUSTER" + '\n')
    f.close()
    
    f = open(dir_path + "/ACI_PRE_WORK/3.10 - Create the service graph template.csv", "a")
    f.write("TENANT,SGNAME,FW" + '\n')
    f.close()
    
    f = open(dir_path + "/ACI_PRE_WORK/3.11 - Create the policy-based redirect policies.csv", "a")
    f.write("TENANT,PBRNAME,REDIRECTIP,REDIRECTMAC" + '\n')
    f.close()
    
    f = open(dir_path + "/ACI_PRE_WORK/3.12 - Create a new contract with the Service Graph name defined as the subject.csv", "a" )
    f.write("TENANT,CONTRACT_NAME,SUBJECT" + '\n')
    f.close()
    
    f = open(dir_path + "/ACI_PRE_WORK/3.13 - Create device selection policy.csv", "a")
    f.write("TENANT,CONTRACT_NAME,SGNAME,NODENAME,FW,BD,PBR_POLICY,CLUSTER" + '\n')
    f.close()
    
    f = open(dir_path + "/ACI_PRE_WORK/3.14  - Assign the L4-L7 service graph to the new contract in the contract subject.csv", "a" )
    f.write("TENANT,CONTRACT_NAME,SGNAME,SUBJECT" + '\n')
    f.close()

    # Write the configs
    for tenant in write_to_aci_cfg:
        for vrf in write_to_aci_cfg[tenant]:
                for epg in write_to_aci_cfg[tenant][vrf]:
			fwbdname =  write_to_aci_cfg[tenant][vrf][epg][0]['fwbdname']
			sgtname =  write_to_aci_cfg[tenant][vrf][epg][0]['sgtname']
			fwbdip   =  write_to_aci_cfg[tenant][vrf][epg][0]['fwbdip']
			fwvip   =  write_to_aci_cfg[tenant][vrf][epg][0]['fwvip']
			l4l7name = write_to_aci_cfg[tenant][vrf][epg][0]['l4l7']
			leafid_a = write_to_aci_cfg[tenant][vrf][epg][0]['leafa']
			leafid_b = write_to_aci_cfg[tenant][vrf][epg][0]['leafb']
			pbrname = write_to_aci_cfg[tenant][vrf][epg][0]['pbrname']
			sgcontractname = write_to_aci_cfg[tenant][vrf][epg][0]['sgcontractname']
			fwaname = write_to_aci_cfg[tenant][vrf][epg][0]['fwaname']
			fwbname = write_to_aci_cfg[tenant][vrf][epg][0]['fwbname']
			fwcluster = write_to_aci_cfg[tenant][vrf][epg][0]['fwcluster']
			vlan = write_to_aci_cfg[tenant][vrf][epg][0]['vlan']
			fwbdsubnetmask_t = write_to_aci_cfg[tenant][vrf][epg][0]['fwbdsubnet'].split("/")
			fwbdsubnetmask =  fwbdsubnetmask_t[1]	
			f = open(dir_path + "/ACI_PRE_WORK/3.8 - Create a bridge domain for the policy-based routing policy.csv", "a")
			f.write(tenant +  "," + fwbdname + "," + fwbdip + "/" + fwbdsubnetmask + "," + vrf + '\n')
			f.close()
			
			f = open(dir_path + "/ACI_PRE_WORK/3.9 - Create the L4-L7 device.csv", "a")
			f.write(tenant +  "," + l4l7name + "," + fwaname.upper() + "," + str(leafid_a) + "," + "pc_" + fwaname + "," + fwbname.upper() + "," + str(leafid_b) + "," + "pc_" + fwbname + "," + dc.upper() + "_" + district.upper() + "_" + "PHYS_DOM" "," + vlan + "," + fwcluster +   '\n')
			f.close()
			
			f = open(dir_path + "/ACI_PRE_WORK/3.10 - Create the service graph template.csv", "a")
			f.write(tenant +  "," + sgtname + "," + l4l7name + '\n')
			f.close()
			
			f = open(dir_path + "/ACI_PRE_WORK/3.11 - Create the policy-based redirect policies.csv", "a")
			f.write(tenant +  "," + pbrname + "," + fwvip + "," + '\n')
			f.close()
			
			f = open(dir_path + "/ACI_PRE_WORK/3.12 - Create a new contract with the Service Graph name defined as the subject.csv", "a")
			f.write(tenant +  "," + sgcontractname + "," + "Permit_Any" + '\n')
			f.close()
			
			f = open(dir_path + "/ACI_PRE_WORK/3.13 - Create device selection policy.csv", "a")
			f.write(tenant +  "," + sgcontractname + "," + sgtname + "," + "N1," + l4l7name + "," + fwbdname + "," + pbrname + "," + fwcluster  + '\n')
			f.close()
			
			f = open(dir_path + "/ACI_PRE_WORK/3.14  - Assign the L4-L7 service graph to the new contract in the contract subject.csv", "a")
			f.write(tenant +  "," + sgcontractname + "," + sgtname + "," + "Permit_Any"  + '\n')
			f.close()
			break
     
    for tenant in write_to_aci_cfg:
	for vrf in write_to_aci_cfg[tenant]:
		for epg in write_to_aci_cfg[tenant][vrf]:
			dirname = epg.split("-")
			tenantdir = dirname[0]
			os.mkdir(dir_path + "/ACI_" + tenantdir + "-" + vrf)
			break

    # MIGRATION STEPS - PRINT CONTRACTS
    # Print L3 out associate new contract
    l3out = {}
    for tenant in write_to_aci_cfg:
	for vrf in write_to_aci_cfg[tenant]:
		for epg in write_to_aci_cfg[tenant][vrf]:
			for d in write_to_aci_cfg[tenant][vrf][epg]:
				if d['remove_l3_contract'] == 'yes' and d['l3'] == 'yes' :
					dirname = epg.split("-")
					tenantdir = dirname[0]
					fname = vrf + " 1 Type " + d['type'] +  " - Associate contracts to L3Out as consumer.csv"
					if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
    						f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
    						f.write("TENANT,L3OUT,NETWORK,NEW_L3_CONTRACT" + '\n')
    						f.close()

					l3out_name = d['l3out']
					ext_epg = d['extepg']
					newcontract = tenantdir + "-" + vrf + "-" + "SG-PBR-Permit_Any"
					newcontract_key = tenantdir + "-" + vrf + "-" + l3out_name + "-" + ext_epg + "-" + newcontract
					if newcontract_key not in l3out and l3out_name != 'N/A':
						l3out[newcontract_key] = {}
						f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
						fverify = open(dir_path + "/ACI_CONTRACT_VERIFICATION/" + tenantdir + "-" + vrf, "a") 
						
						f.write(tenant + "," + l3out_name + "," + ext_epg + "," + newcontract + '\n')
						fverify.write(tenant + "," + l3out_name + "," + ext_epg + '\n')
						
						fverify.close()
						f.close()
    
    # Print EPG associate new contract
    # For type-B, remove OLD EPG contract
    for tenant in write_to_aci_cfg:
	for vrf in write_to_aci_cfg[tenant]:
		for epg in write_to_aci_cfg[tenant][vrf]:
			if toexclude is True and epg in exinfo:
				print "OK: Excluding EPG " + epg + " from new contract association"
				continue	
			for d in write_to_aci_cfg[tenant][vrf][epg]:
				if d['l3'] == 'yes' :
					s = epg.split("-")
					tenantdir = s[0]
					type = d['type']
					ap = d['ap']
					if type == 'A':
						fname = vrf + " 2 Type " + d['type'] + " - Assign new contract as provider.csv"
						if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
    							f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
    							f.write("TENANT,AP,EPG,NEW_EPG_CONTRACT" + '\n')
    							f.close()
    						f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
						f.write(tenant + "," + ap + "," + epg + "," + tenantdir + "-" + vrf + "-" + "SG-PBR-Permit_Any" + '\n')
						f.close()

					if type == 'B':
						fname = vrf + " 2 Type " + d['type'] + " - Assign new contract as provider and delete old contracts.csv"
						if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
    							f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
    							f.write("TENANT,AP,EPG,NEW_EPG_CONTRACT,OLD_EPG_CONTRACT" + '\n')
    							f.close()
    						f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a")
						for c in d['contract']: 
							f.write(tenant + "," + ap + "," + epg + "," + tenantdir + "-" + vrf + "-" + "SG-PBR-Permit_Any" + "," + c + '\n')
						f.close()

					fverify = open(dir_path + "/ACI_CONTRACT_VERIFICATION/" + tenantdir + "-" + vrf, "a") 
					fverify.write(tenant + "," + ap + "," + epg + '\n')
					f.close()
    
    l3out = {}
    #Contract removal
    for tenant in write_to_aci_cfg:
	for vrf in write_to_aci_cfg[tenant]:
		for epg in write_to_aci_cfg[tenant][vrf]:
			if toexclude is True and epg in exinfo:
				print "OK: Excluding EPG " + epg + " from old contract removal"
				continue	
			for d in write_to_aci_cfg[tenant][vrf][epg]:
				if d['l3'] == 'yes' :
					d_l3out = d['l3out']
					d_extepg = d['extepg']
					ap = d['ap']
					s = epg.split("-")
					type = d['type']
					tenantdir = s[0]
					fname = vrf + " 3 Type " + d['type'] + " - Remove contract from L3Out and EPG as provider_consumer.csv"	
					if not os.path.isfile(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname):
    						f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
    						f.write("TENANT,AP,EPG,OLD_EPG_CONTRACT,L3OUT,NETWORK,OLD_L3_CONTRACT" + '\n')
    						f.close()
					
					for c in d['contract']:
						if d_l3out not in l3out:
							l3out[d_l3out] = {}
							if d_extepg not in l3out[d_l3out]:
								l3out[d_l3out][d_extepg] = {}
								l3out[d_l3out][d_extepg][c] = {}
    								f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
								f.write(tenant + ",,,," + d_l3out +  "," + d_extepg + "," + c + '\n')
								f.close()
							if d_extepg in l3out[d_l3out]:
								if c not in l3out[d_l3out][d_extepg]:
									l3out[d_l3out][d_extepg][c] = {}
    									f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
									f.write(tenant + ",,,," + d_l3out +  "," + d_extepg + "," + c + '\n')
									f.close()
						if d_l3out in l3out:
							if d_extepg not in l3out[d_l3out]:
                                                                l3out[d_l3out][d_extepg] = {}
                                                                l3out[d_l3out][d_extepg][c] = {}
    								f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
								f.write(tenant + ",,,," + d_l3out +  "," + d_extepg + "," + c + '\n')
								f.close()
							if d_extepg in l3out[d_l3out]:
                                                                if c not in l3out[d_l3out][d_extepg]:
                                                                        l3out[d_l3out][d_extepg][c] = {}
    									f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
									f.write(tenant + ",,,," + d_l3out +  "," + d_extepg + "," + c + '\n')
									f.close()
						if type == 'A':
    							f = open(dir_path + "/ACI_" + tenantdir + "-" + vrf + "/" + fname, "a") 
							f.write(tenant + "," + ap + "," + epg + "," + c + '\n')
						f.close()

    f = open('write_to_aci_cfg.json', 'w') 
    f.write(json.dumps(write_to_aci_cfg) )
    f.close()

if __name__ == '__main__':
	main(sys.argv[1:])