#!/usr/bin/env python
from ciscoconfparse import CiscoConfParse
import re
import json
import sys
import os
import fnmatch 
import openpyxl
import json

def get_inner_outer_mapping(dc):

    pattern = '*' + dc.upper() + '*Port Map*'
  
    files = os.listdir('.') 
    for name in files:
	if fnmatch.fnmatch(name, pattern):
		if not bool(re.search('^~',name, re.IGNORECASE)): 
        		filename = name
			break
    worksheets = []
    wb = openpyxl.load_workbook(filename, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('7706')
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    mapping = {}
        
    for cells in ws.iter_rows(min_row=row_start, min_col=1, max_col=24): 
	for vals in cells:
		if vals.value == 'A':
			n7k_local = 1
		if vals.value == 'B':
			n7k_local = 2
		if vals.value == 'C':
			n7k_local = 3
		if vals.value == 'D':
			n7k_local = 4
		if vals.value == 'E':
			n7k_local = 1
		if vals.value == 'F':
			n7k_local = 2
		
		if bool(re.search("nxc(.*)inner",str(vals.value), re.IGNORECASE)) or bool(re.search("nxc(.*)outer",str(vals.value), re.IGNORECASE)):
			remote_n7k = vals.value
			local_interface = ws[vals.column + str(vals.row+1)].value

			if local_interface is None:
				local_interface = ws[vals.column + str(vals.row-1)].value

			local_interface = local_interface.replace("-","/")
			xvals = vals.value.split("_E")
			remote_interface = xvals[1]
			remote_n7k = xvals[0]
			
			# Determine local N7K now
			if bool(re.search("sde",remote_n7k, re.IGNORECASE)) and bool(re.search("inner",remote_n7k, re.IGNORECASE)):
				local_n7k = dc + 'sde' + 'nxc' + str(n7k_local) + 'sdeouter'
			if bool(re.search("sde",remote_n7k, re.IGNORECASE)) and bool(re.search("outer",remote_n7k, re.IGNORECASE)):
				local_n7k = dc + 'sde' + 'nxc' + str(n7k_local) + 'sdeinner'
			
			if bool(re.search("gis",remote_n7k, re.IGNORECASE)) and bool(re.search("inner",remote_n7k, re.IGNORECASE)):
				local_n7k = dc + 'gis' + 'nxc' + str(n7k_local) + 'gisouter'
			if bool(re.search("gis",remote_n7k, re.IGNORECASE)) and bool(re.search("outer",remote_n7k, re.IGNORECASE)):
				local_n7k = dc + 'gis' + 'nxc' + str(n7k_local) + 'gisinner'
			
			if bool(re.search("soe",remote_n7k, re.IGNORECASE)) and bool(re.search("inner",remote_n7k, re.IGNORECASE)):
				local_n7k = dc + 'soe' + 'nxc' + str(n7k_local) + 'soeouter'
			if bool(re.search("soe",remote_n7k, re.IGNORECASE)) and bool(re.search("outer",remote_n7k, re.IGNORECASE)):
				local_n7k = dc + 'soe' + 'nxc' + str(n7k_local) + 'soeinner'

			
			if local_n7k not in mapping:
				mapping[local_n7k] = {}

			if local_interface not in mapping[local_n7k]:
				mapping[local_n7k][local_interface] = {}

			if remote_interface not in mapping[local_n7k][local_interface]:
				mapping[local_n7k][local_interface][remote_interface] = {}
			
			if remote_n7k not in mapping[local_n7k][local_interface][remote_interface]:
				mapping[local_n7k][local_interface][remote_interface][remote_n7k] = {}
    
    return mapping

def get_bgp_int_vlan(dc,district):

    data = {}
    p2p_n7k_mapping = get_inner_outer_mapping(dc)
 
    if district.lower() in ('sde'):
    	numfiles = ['1','2']
    else:
    	numfiles = ['1','2','3','4']

    for i in numfiles:
	for j in ['inner','outer']:
		filename = dc + district + 'nxc' + str(i) + district + j + '.log'
		n7k = dc + district + 'nxc' + str(i) + district + j 
		
		# Need to code this
		n7k_mapping = get_inner_outer_mapping(dc)
	
	  	data[n7k] = {}	
		parse = CiscoConfParse(filename)

		# GET SVI
		for obj in parse.find_objects("interface Vlan"):
			svi = obj.text
			svi = svi.replace("interface Vlan","")
		
			if obj.hash_children != 0:
				for c in obj.children:
					
					if bool((re.search('vrf member',c.text,re.IGNORECASE))):
						vrf = c.text
						vrf = vrf.replace("vrf member ","")
						vrf = vrf.lstrip()

					# For outer, set VRF to SVI since outer does not have vrf
					if bool((re.search('outer',n7k,re.IGNORECASE))):
						vrf = svi
					
					if bool((re.search('ip address',c.text,re.IGNORECASE))):
							svi_ip = c.text
							svi_ip = c.text.replace("  ip address ","")
							svi_ip = re.sub('/.*',"", svi_ip)

				data[n7k][vrf] = {}
				data[n7k]['P2P'] = {}
				data[n7k][vrf] = { 'svi' : svi, 'shutdown' : 'N', 'fw_trunk_int' : [], 'remote_as' : 'N/A', 'neighbors' : [], 'svi_ip' : svi_ip, 'local_as' : 'N/A' }

		# GET FW INT
		vlallowed = []
		for obj in parse.find_objects("interface Ethernet"):
			if obj.hash_children != 0:
				if obj.re_search_children("switchport trunk allowed"):
					for c in obj.children:
						if bool((re.search('switchport trunk allowed',c.text,re.IGNORECASE))):
							fwint = obj.text
							fwint = fwint.replace("interface Ethernet","")
							c.text = c.text.replace("switchport trunk allowed vlan","")
							vlist = c.text.split(",")
							for v in vlist:
								if bool((re.search('-',v,re.IGNORECASE))):
									vv = v.split("-")
									vv[1] = int(vv[1]) + int(1)
									for vvv in xrange(int(vv[0]),int(vv[1])):
										vlallowed.append(vvv) 
								else:
									vlallowed.append(int(v))

							for vls in vlallowed:
								for vrf_vals in data[n7k]:
									if vrf_vals != 'P2P':
										if vls == int(data[n7k][vrf_vals]['svi']):
											if len(data[n7k][vrf_vals]['fw_trunk_int']) == 0:
												data[n7k][vrf_vals]['fw_trunk_int']  = [ fwint ]
											else:
												data[n7k][vrf_vals]['fw_trunk_int'].append(fwint)
								vlallowed = []

							
							
    		# Get BGP Neighbors		
		for obj in parse.find_objects("router bgp"):
			local_as = obj.text
			local_as = local_as.replace("router bgp ","")
                        if obj.hash_children != 0:
				# For inner VDC
                                if obj.re_search_children("vrf"):
                                        for c in obj.children:
                                                if bool((re.search('vrf',c.text,re.IGNORECASE))):
							vrfmember = c.text
							for obj2 in parse.find_objects(vrfmember):
								if obj2.re_search_children("neighbor"):
									for d in obj2.children:
										vrfmember = vrfmember.replace("vrf ","") 
										vrfmember = vrfmember.lstrip()
										if bool((re.search('address-family',d.text,re.IGNORECASE))):
											continue
										if bool((re.search('NonAff',vrfmember,re.IGNORECASE))):
											continue
										attribs = d.text
										attribs = attribs.lstrip()
										attribsx = attribs.split(" ")
										neighbor = attribsx[1]
										remote_as = attribsx[3]
                                                                                data[n7k][vrfmember]['remote_as']  = remote_as
										if len(data[n7k][vrfmember]['neighbors']) == 0:
                                                                                        data[n7k][vrfmember]['neighbors']  = [ neighbor ]
                                                                                else:
                                                                                        data[n7k][vrfmember]['neighbors'].append(neighbor)

										data[n7k][vrfmember]['local_as'] = local_as
				# For outer VDC
				if bool((re.search('outer',filename,re.IGNORECASE))):
					for c in obj.children:
						if bool((re.search('neighbor ',c.text,re.IGNORECASE))):
							attribs = c.text
                                                	attribs = attribs.lstrip()
							attribsx = attribs.split(" ")
                                                	neighbor = attribsx[1]
                                                	remote_as = attribsx[3]
			
							for vrfmember in data:
								for v in data[vrfmember]:
									if v == 'P2P':
										continue
									if data[vrfmember][v]['svi_ip'] == neighbor:
										for n in data[vrfmember][v]['neighbors']:
											for o in data:
												for p in data[o]:
													if unicode(p).isnumeric():
														
														if data[o][p]['svi_ip'] == n:
															if 'vrf' not in data[o][p]:
																data[o][p]['vrf'] = {}
																del data[o][p]['neighbors']
															if v not in data[o][p]['vrf']:
																data[o][p]['vrf'][v] = {}
																data[o][p]['vrf'][v]['neighbor_ip'] = [neighbor]
															else:
																if neighbor not in data[o][p]['vrf'][v]['neighbor_ip']:
																	data[o][p]['vrf'][v]['neighbor_ip'].append(neighbor)
															data[o][p]['remote_as'] = remote_as
															data[o][p]['local_as'] = local_as
						
    # Append P2P data
    # P2P dictionary format:
    # Local Int -> remote int -> remote N7K
    for n7k in data:
	data[n7k]['P2P'] = p2p_n7k_mapping[n7k]
		 
    print json.dumps(data)


get_bgp_int_vlan('dc2','sde')



