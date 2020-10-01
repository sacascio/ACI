#!/usr/bin/env python3

import getopt
import os.path
import sys
# from xlrd import open_workbook, XLRDError
# from fileinput import filename
import warnings
import json

import openpyxl

warnings.filterwarnings("ignore")

def get_epg_contracts(dafe_file,c_type):

    worksheets = []
    contracts = {}

    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index(c_type)
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start + 1, row_end):

        cell = 'A' + str(x)
        tenant = ws[cell].value

        cell = 'C' + str(x)
        epg = ws[cell].value

        cell = 'D' + str(x)
        contract = ws[cell].value

        if contract not in contracts:
            contracts[contract] = []
            contracts[contract].append(epg)

        else:
            contracts[contract].append(epg)


    return contracts

def get_l3_contracts(dafe_file):
    worksheets = []
    l3contracts = {}

    wb = openpyxl.load_workbook(dafe_file, data_only=True)

    for sheet in wb:
        worksheets.append(sheet.title)
    wb.close()

    wb.active = worksheets.index('external_epg')
    ws = wb.active
    row_start = ws.min_row
    row_end = ws.max_row

    for x in range(row_start + 1, row_end):

        cell = 'A' + str(x)
        tenant = ws[cell].value

        if tenant in ('mgmt', 'infra' ):
            continue

        cell = 'B' + str(x)
        l3out = ws[cell].value

        cell = 'D' + str(x)
        extepg = ws[cell].value

        cell = 'G' + str(x)
        cons = ws[cell].value

        cell = 'I' + str(x)
        prov = ws[cell].value

        if cons is None:
            consumed = "NO_CONTRACT"
        else:
            consumed = cons.split(",")

        if prov is None:
            provided = "NO_CONTRACT"
        else:
            provided = prov.split(",")

        if extepg not in l3contracts:
            l3contracts[extepg] = {}

        if l3out not in l3contracts[extepg] :
            l3contracts[extepg][l3out] = {}

        if tenant not in l3contracts[extepg][l3out]:
            l3contracts[extepg][l3out][tenant] = {}

        l3contracts[extepg][l3out][tenant]['provided'] = provided
        l3contracts[extepg][l3out][tenant]['consumed'] = consumed


    return l3contracts

def main(argv):
    # Arguments
    found = 0
    if len(argv) == 0:
        usage()

    try:
        opts, args = getopt.getopt(argv, "f:he:", ["file=", "help", "extepg"])
    except getopt.GetoptError as err:
        print (str(err))
        sys.exit(2)
    else:
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                usage()
            if opt in ("-f", "--file"):
                dafe_file = arg

            if opt in ("-e", "--extepg"):
                ext_epg_file = arg
    try:
        dafe_file
    except NameError:
        print ("DAFE file not passed -f|--file")
        sys.exit(9)

    if not os.path.isfile(dafe_file):
        print (sys.argv[0] + " Input File %s NOT found" % dafe_file)
        sys.exit(9)

    try:
        ext_epg_file
    except NameError:
        print ("Ext epg file not passed -e|--extepg")
        sys.exit(9)

    if not os.path.isfile(ext_epg_file):
        print (sys.argv[0] + " Input File %s NOT found" % ext_epg_file)
        sys.exit(9)

    f = open(ext_epg_file, 'r+')
    ext_epg_list = [line for line in f.readlines()]
    f.close()

    l3_contracts = get_l3_contracts(dafe_file)
    epg_provided_contracts = get_epg_contracts(dafe_file,'epg_provide_contract')
    epg_consumed_contracts = get_epg_contracts(dafe_file,'epg_consumed_contract')

    for ext_epg_list_file in ext_epg_list:
        ext_epg_list_file = ext_epg_list_file.rstrip()

        for l3out in l3_contracts[ext_epg_list_file]:
            for tenant in l3_contracts[ext_epg_list_file][l3out]:
                for consumed in l3_contracts[ext_epg_list_file][l3out][tenant]['consumed']:
                        if consumed in epg_consumed_contracts:
                            print ("** CONSUMED CONTRACT: %s, EXT_EPG: %s FOUND IN THE FOLLOWING EPGs as CONSUMER: %s" % (consumed,ext_epg_list_file,epg_consumed_contracts[consumed]))
                            #for cc in epg_consumed_contracts[consumed]:
                            #    print (cc)
                        #else:
                        #    print ("OK: CONSUMED CONTRACT: %s, EXT_EPG: %s NOT FOUND IN ANY EPG as CONSUMER" % (consumed,ext_epg_list_file))

                        if consumed in epg_provided_contracts:
                            print ("** CONSUMED CONTRACT: %s, EXT_EPG: %s FOUND IN THE FOLLOWING EPGs as PROVIDER: %s" % (consumed,ext_epg_list_file,epg_provided_contracts[consumed]))
                            #for cc in epg_provided_contracts[consumed]:
                            #    print (cc)
                        #else:
                        #    print ("OK: CONSUMED CONTRACT: %s, EXT_EPG: %s NOT FOUND IN ANY EPG as PROVIDER" % (consumed,ext_epg_list_file))

                        # Now check and see if the contracts assigned to the targeted L3outs exist on other L3 Outs
                        for ext_epg_copy in l3_contracts:
                            for l3out_copy in l3_contracts[ext_epg_copy]:
                                for tenant_copy in l3_contracts[ext_epg_copy][l3out_copy]:
                                    for consumed_copy in l3_contracts[ext_epg_copy][l3out_copy][tenant_copy]['consumed']:
                                        if consumed == consumed_copy and ext_epg_copy != ext_epg_list_file:
                                            print ("** CONSUMED CONTRACT: %s, EXT_EPG: %s FOUND IN THE FOLLOWING EXT_EPG as CONSUMER: %s" % (consumed,ext_epg_list_file,ext_epg_copy))
                                            #print (ext_epg_copy)
                                            #found = 1
                        #if found == 0:
                        #    print ("OK: CONSUMED CONTRACT: %s, EXT_EPG: %s NOT FOUND IN ANY EXT_EPG as CONSUMER" % (consumed,ext_epg_list_file))
                        #else:
                        #    found = 0

                        for ext_epg_copy in l3_contracts:
                            for l3out_copy in l3_contracts[ext_epg_copy]:
                                for tenant_copy in l3_contracts[ext_epg_copy][l3out_copy]:
                                    for provided_copy in l3_contracts[ext_epg_copy][l3out_copy][tenant_copy]['provided']:
                                        if consumed == provided_copy and ext_epg_copy != ext_epg_list_file:
                                            print ("** CONSUMED CONTRACT: %s, EXT_EPG: %s FOUND IN THE FOLLOWING EXT_EPG as PROVIDER: %s" % (consumed, ext_epg_list_file,ext_epg_copy))
                                            #print (ext_epg_copy)
                                            #found = 1

                        #if found == 0:
                        #    print ("OK: CONSUMED CONTRACT: %s, EXT_EPG: %s NOT FOUND IN ANY EXT_EPG as PROVIDER" % (consumed,ext_epg_list_file))
                        #else:
                        #    found = 0

                        # PROVIDER
        for l3out in l3_contracts[ext_epg_list_file]:
            for tenant in l3_contracts[ext_epg_list_file][l3out]:
                for provided in l3_contracts[ext_epg_list_file][l3out][tenant]['provided']:
                        if provided in epg_provided_contracts:
                            print ("** PROVIDED CONTRACT: %s, EXT_EPG: %s FOUND IN THE FOLLOWING EPGs as PROVIDER: %s" % (provided,ext_epg_list_file,epg_provided_contracts[provided]))
                            #for cc in epg_provided_contracts[provided]:
                            #    print (cc)
                        #else:
                        #    print ("OK: PROVIDED CONTRACT: %s, EXT_EPG: %s NOT FOUND IN ANY EPG as PROVIDER" % (provided,ext_epg_list_file))

                        if provided in epg_consumed_contracts:
                            print ("** PROVIDED CONTRACT: %s, EXT_EPG: %s FOUND IN THE FOLLOWING EPGs as CONSUMER: %s" % (provided,ext_epg_list_file,epg_consumed_contracts[provided]))
                            #for cc in epg_consumed_contracts[provided]:
                            #    print (cc)
                        #else:
                        #    print ("OK: PROVIDED CONTRACT: %s, EXT_EPG: %s NOT FOUND IN ANY EPG as CONSUMER" % (provided,ext_epg_list_file))

                        # Now check and see if the contracts assigned to the targeted L3outs exist on other L3 Outs
                        for ext_epg_copy in l3_contracts:
                            for l3out_copy in l3_contracts[ext_epg_copy]:
                                for tenant_copy in l3_contracts[ext_epg_copy][l3out_copy]:
                                    for provided_copy in l3_contracts[ext_epg_copy][l3out_copy][tenant_copy][ 'provided']:
                                        if provided == provided_copy and ext_epg_copy != ext_epg_list_file:
                                            print ("** PROVIDED CONTRACT: %s, EXT_EPG: %s FOUND IN THE FOLLOWING EXT_EPG as PROVIDER: %s" % ( provided, ext_epg_list_file,ext_epg_copy))
                                            #print (ext_epg_copy)
                                            #found = 1
                            #if found == 0:
                            #    print ("OK: PROVIDED CONTRACT: %s, EXT_EPG: %s NOT FOUND IN ANY EXT_EPG as PROVIDER" % ( provided, ext_epg_list_file))
                            #else:
                            #    found = 0

                        for ext_epg_copy in l3_contracts:
                            for l3out_copy in l3_contracts[ext_epg_copy]:
                                for tenant_copy in l3_contracts[ext_epg_copy][l3out_copy]:
                                    for consumed_copy in l3_contracts[ext_epg_copy][l3out_copy][tenant_copy]['consumed']:
                                            if provided == consumed_copy and ext_epg_copy != ext_epg_list_file:
                                                print ("** PROVIDED CONTRACT: %s, EXT_EPG: %s FOUND IN THE FOLLOWING EXT_EPG as CONSUMER: %s" % (provided, ext_epg_list_file,ext_epg_copy))
                                                #print (ext_epg_copy)
                                                #found = 1

                            #if found == 0:
                            #    print ("OK: PROVIDED CONTRACT: %s, EXT_EPG: %s NOT FOUND IN ANY EXT_EPG as CONSUMER" % (provided, ext_epg_list_file))
                            #else:
                            #    found = 0

if __name__ == '__main__':
    main(sys.argv[1:])
