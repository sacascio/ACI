#!/usr/bin/env python

import openpyxl
import json
worksheets = []
ws_definition = "sandwich"
data = {}
data2 = {}

wb = openpyxl.load_workbook('fwint.xlsx', data_only=True)

for sheet in wb:
        worksheets.append(sheet.title) 
wb.close()


try:
    wb.active = worksheets.index(ws_definition)
except:
    print "Worksheet %s not found" % ws_definition
else:
    ws = wb.active
    row_start = ws.min_row
    row_end   = ws.max_row

    # Process ws_definition tab 
    for x in range(row_start,row_end+1):
        cell = 'A' + str(x)
        district = ws[cell].value 
 
        if district is None:
            continue

        cell = 'B' + str(x)
        loc = ws[cell].value 
 
        cell = 'C' + str(x)
        n7k = ws[cell].value 

        cell = 'D' + str(x)
        dc = ws[cell].value 

        cell = 'E' + str(x)
        fw = ws[cell].value 

        cell = 'F' + str(x)
        locint = ws[cell].value 

        cell = 'G' + str(x)
        remint = ws[cell].value 

        if district not in data:
            data[district] = {}
        if loc not in data[district]:
            data[district][loc] = {}
        if n7k not in data[district][loc]:
            data[district][loc][n7k] = {}
        if dc not in data[district][loc][n7k]:
            data[district][loc][n7k][dc] = {}
        if fw not in data[district][loc][n7k][dc]:
            data[district][loc][n7k][dc][fw] = {}
        if locint not in data[district][loc][n7k][dc][fw]:
            data[district][loc][n7k][dc][fw][locint] = {}

        data[district][loc][n7k][dc][fw][remint] = locint

    print json.dumps(data)            


#################################################################

ws_definition = 'juniper'         	   

try:
    wb.active = worksheets.index(ws_definition)
except:
        print "Worksheet %s not found" % ws_definition
else:
        ws = wb.active
        row_start = ws.min_row
        row_end   = ws.max_row


        # Process ws_definition tab 
        for x in range(row_start,row_end+1):
            cell = 'A' + str(x)
            district = ws[cell].value 

            if district is None:
                continue

            cell = 'B' + str(x)
            n7k = ws[cell].value 

            cell = 'C' + str(x)
            dc = ws[cell].value 

            cell = 'D' + str(x)
            jnpnum = ws[cell].value 

            cell = 'E' + str(x)
            localint = ws[cell].value 
  
            cell = 'F' + str(x)
            remint = ws[cell].value 

            cell = 'G' + str(x)
            devname = ws[cell].value 

            if district not in data2:
                data2[district] = {}
            if n7k not in data2[district]:
                data2[district][n7k] = {}
            if dc not in data2[district][n7k]:
                data2[district][n7k][dc] = {}
            if jnpnum not in data2[district][n7k][dc]:
                data2[district][n7k][dc][jnpnum] = {}

            data2[district][n7k][dc][jnpnum] = {'localint' : localint,'remint' : remint,'devname' : devname}

        print json.dumps(data2)
