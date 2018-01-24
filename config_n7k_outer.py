#!/usr/bin/env python

import openpyxl
import sys
import getopt
import os.path
import magic
import re
from xlrd import open_workbook, XLRDError
import json
from IPy import IP
import requests
import warnings

warnings.filterwarnings("ignore")

def getValueWithMergeLookup(sheet, cell):
    for m_range in sheet.merged_cell_ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(m_range))
        for row in merged_cells:
            if cell in row:
                return sheet.cell(merged_cells[0][0]).value


def outer_vdc_config(ws_definition_data,final_all_inner_data,bgp_asn,outer_to_pa_data,n7k_fw_int,loopback_data,outer_jnp_data,configure,lines,detailops,out_to_jnp_fw_int):
  
    commands = []
    loopback_position  = {'N7K-A' : 1, 'N7K-B' : 2, 'N7K-C' : 3, 'N7K-D' : 4, 'N7K-E' : 1, 'N7K-F' : 2}
    
    for info in lines:
        i = info.split(",")
        district  = i[0].upper()
        dc        = i[1].lower()
        nexusvdc  = i[2].upper()
        device_ip = i[3]
        device_un = i[4]
        device_pw = i[5]
        fw_to_vlan = {}
        
        print "!!! District %s, DC %s, nexusVDC %s -- Outer Config" % (district,dc,nexusvdc)
        print "!"
        
        if configure is not True:
                print "!! If prefix-list and route-map commands are already configured on the N7K, they will not be executed"
                commands.append("ip prefix-list DEFAULT_ROUTE seq 5 permit 0.0.0.0/0")
                commands.append("!")
                commands.append("route-map PERMIT_DEFAULT_ONLY permit 10")
                commands.append("   match ip address prefix-list DEFAULT_ROUTE")
                commands.append("route-map PERMIT_DEFAULT_ONLY deny 20")
                print '\n'.join(map(str,commands))
                print "!"
                commands = []

        else:
            prefix_list = send_to_n7k_api_show("show ip prefix-list DEFAULT_ROUTE",device_ip,district,dc,nexusvdc,device_un,device_pw)
            if prefix_list is None:
                commands.append("ip prefix-list DEFAULT_ROUTE seq 5 permit 0.0.0.0/0")

            route_map_output = send_to_n7k_api_show("show route-map PERMIT_DEFAULT_ONLY",device_ip,district,dc,nexusvdc,device_un,device_pw)
            if route_map_output is None:
                commands.append("route-map PERMIT_DEFAULT_ONLY permit 10")
                commands.append("   match ip address prefix-list DEFAULT_ROUTE")
                commands.append("route-map PERMIT_DEFAULT_ONLY deny 20")

            if len(commands) > 0:
                print '\n'.join(map(str,commands))
                print "!"
                print "*** sending above config to %s,%s,%s ***"  %(dc,district,nexusvdc)
                commands = " ; ".join(map(str,commands))
                send_to_n7k_api(device_ip,commands,district,dc,nexusvdc,device_un,device_pw)

        commands = []
  
        for vsys in ws_definition_data[district]:
            gen_int_config = 0
            found = 0
            
            # If there's at least 1 VRF to config per vSYS, write interface config
            for attrib in ws_definition_data[district][vsys]:
              
                if attrib['config'] == 'yes':
                    prifw  =  attrib[dc + 'prifwname']
                    sbyfw  =  attrib[dc + 'sbyfwname']
                    outervlan = attrib['outervdcencap']
                    
                    if prifw in fw_to_vlan: 
                        for v in fw_to_vlan[prifw]:
                            if v['vlan'] == outervlan:
                                found = 1
                                break
                            
                        if found == 0:
                            fw_to_vlan[prifw].append({ 'vlan' : int(outervlan) })
                        else:
                            found = 0
                                
                    else:
                        fw_to_vlan[prifw] = [{ 'vlan' : int(outervlan)}]
                    
                    if sbyfw in fw_to_vlan:
                        for v in fw_to_vlan[sbyfw]:
                            if v['vlan'] == outervlan:
                                found = 1
                                break
                            
                        if found == 0:
                            fw_to_vlan[sbyfw].append({ 'vlan' : int(outervlan) })
                        else:
                            found = 0
                    else:
                        fw_to_vlan[sbyfw] = [{ 'vlan' : int(outervlan)}]
                    
                   
                    
                    gen_int_config = gen_int_config + 1
            
            
            
            if gen_int_config > 0:
                outervdcvlan =  ws_definition_data[district][vsys][0]['outervdcencap']
                ospfarea     =  ws_definition_data[district][vsys][0]['ospf' + dc]
                n7kip        =  outer_to_pa_data[district][vsys][nexusvdc][0][dc + 'n7kip']
               
                commands.append("! Create L2 VLAN")
                commands.append("vlan " + str(outervdcvlan) )
                commands.append("!")
                commands.append("interface vlan " + str(outervdcvlan))
                commands.append("  description L3_%s_%s" % (district,vsys))
                commands.append("  ip address %s 255.255.255.252" % (n7kip))
                commands.append("  mtu 9192")
                commands.append("  ip ospf network point-to-point")
                commands.append("  ip router ospf %s area 0.0.0.%s" % (district,ospfarea))
                commands.append("  no shutdown")
                
                
        
        if 'vlan' in detailops or 'NONE' in detailops :            
            print '\n'.join(map(str,commands))
            print "!"
            
                      
        if configure is True and ( 'vlan' in detailops or 'NONE' in detailops ):            
            commands = " ; ".join(map(str,commands))
            print "*** sending above config to %s,%s,%s ***"  %(dc,district,nexusvdc)
            send_to_n7k_api(device_ip,commands,district,dc,nexusvdc,device_un,device_pw)
        
        commands = []
                
        print "!"
        
        
        # Begin OSPF and loopback 0 config
        n7k_num = loopback_position[nexusvdc]
        
        if district == 'SDE':
            lbname = dc + district.lower() + 'nxc' + str(n7k_num) + district.lower() + 'outer'
        else:
            lbname = dc + 'dcinxc' + str(n7k_num) + 'dciouter'
                    
        for vals in loopback_data[district]:
            if vals[dc + 'hn'] == lbname:
                loopback_address = vals[dc + 'ip']
                break
        
        commands.append("interface loopback0")  
        commands.append("   ip address %s 255.255.255.255" % (loopback_address)  )            
        commands.append("router ospf %s" % (district))
        commands.append("   router-id %s" % (loopback_address))
        
        if 'ospf' in detailops or 'NONE' in detailops:          
            print '\n'.join(map(str,commands))
            print "!"
                            
        if configure is True and ( 'ospf' in detailops or 'NONE' in detailops ):
            commands = " ; ".join(map(str,commands))
            print "*** sending above config to %s,%s,%s ***"  %(dc,district,nexusvdc)
            send_to_n7k_api(device_ip,commands,district,dc,nexusvdc,device_un,device_pw)
        commands = []
                            
        print "!"
        print "!"
      

        # BGP Configuration
        inner_as   = bgp_asn[district]['Inner'][dc]
        outer_as   = bgp_asn[district]['Outer'][dc]
        loopback_address = loopback_address
        
        commands.append("router bgp %s" % (outer_as))
        commands.append(" router-id %s" % (loopback_address))
        commands.append(" log-neighbor-changes")
        commands.append(" address-family ipv4 unicast")
        commands.append("    maximum-paths 8")
                
        for vsys in final_all_inner_data[district]:
            for n7k in sorted(final_all_inner_data[district][vsys]):
                
               
                n7kinneraddress = final_all_inner_data[district][vsys][n7k][0][dc + 'n7kip']
                vrfname         = final_all_inner_data[district][vsys][n7k][0]['n7kvrf' + dc]
                vrfnum          = final_all_inner_data[district][vsys][n7k][0]['vrfnumber']
                vrfnum          = vrfnum.replace("VRF-","")
                n7k_num         = loopback_position[n7k]
                
                if district == 'SDE':
                        innervdc = dc + district.lower() + 'nxc' + str(n7k_num) + district.lower() + 'inner' 
                else:
                        innervdc = dc + 'dcinxc' + str(n7k_num) + district.lower() + 'inner' 
                
                for vsys_tmp in ws_definition_data[district]:
                    for attribs in ws_definition_data[district][vsys_tmp]:
                        if int(attribs['vrfnumber']) == int(vrfnum) and attribs['config'] != 'no' :
                            commands.append(" neighbor %s remote-as %s" % (n7kinneraddress,inner_as ))
                            commands.append(" description TO_%s_%s" % (innervdc,vrfname) )
                            commands.append("    ebgp-multihop 4" )
                            commands.append("    address-family ipv4 unicast")
                            commands.append("       send-community both")
                            commands.append("       route-map PERMIT_DEFAULT_ONLY out")
                            commands.append("       default-originate")
        
        if 'bgp' in detailops or 'NONE' in detailops:        
            print '\n'.join(map(str,commands))    
                          
        if configure is True and ( 'bgp' in detailops or 'NONE' in detailops ):
                            
            commands = " ; ".join(map(str,commands))
            print "*** sending above config to %s,%s,%s ***"  %(dc,district,nexusvdc)
            send_to_n7k_api(device_ip,commands,district,dc,nexusvdc,device_un,device_pw)
        commands = []
        
        print "!"
        print "!"
        
        # Connection to Juniper Router
        config_to_jnp = 0
        commands.append("router bgp %s" % (outer_as))
        
        for jnp in outer_jnp_data[district][nexusvdc]:
            jnpip = outer_jnp_data[district][nexusvdc][jnp][0][dc + 'jnpip']
            outfwint = out_to_jnp_fw_int[district][nexusvdc][dc.upper()][jnp]['localint']
            jnpfwint = out_to_jnp_fw_int[district][nexusvdc][dc.upper()][jnp]['remint']
            jnpname  = out_to_jnp_fw_int[district][nexusvdc][dc.upper()][jnp]['devname']
            commands.append(" neighbor %s remote-as 64710" % (jnpip))
            commands.append(" description TO %s_%s" % (jnpname,jnpfwint))
            commands.append("    address-family ipv4 unicast")
            commands.append("      send-community both")
        
        if 'bb' in detailops or 'NONE' in detailops:
            print '\n'.join(map(str,commands))    
                          
        if configure is True and ( 'bb' in detailops or 'NONE' in detailops ) :
                            
            jnp_cfg = send_to_n7k_api_show("sh bgp sessions",device_ip,district,dc,nexusvdc,device_un,device_pw)

            for vals in jnp_cfg['body']['TABLE_vrf']['ROW_vrf']['TABLE_neighbor']['ROW_neighbor']:
                if vals['remoteas'] == 64710:
                    config_to_jnp = 1
                    break

            if config_to_jnp == 0:
                commands = " ; ".join(map(str,commands))
                print "*** sending above config to %s,%s,%s ***"  %(dc,district,nexusvdc)
                send_to_n7k_api(device_ip,commands,district,dc,nexusvdc,device_un,device_pw)
            else:
                print "Configuration to Juniper already done.  Skipping Juniper config commands"
                config_to_jnp = 0

        
        commands = []
        
        # Get FW interfaces                 
        for fw in fw_to_vlan:
            vlans = []
            for vl in fw_to_vlan[fw]:
                vlans.append(str(vl['vlan']))
                
            vlans.sort()
            
            if fw not in n7k_fw_int[district]['OUTER'][nexusvdc][dc] and ( 'fw' in detailops or 'NONE' in detailops):
                print " WARNING: Firewall " + fw + " does NOT exist in " + dc + " port-map file for district " + district + "," + nexusvdc + "." +  "Vlans " + ",".join(map(str,vlans)) + " will NOT be configured on the FW interface allowed list"
                continue
            
            if 'fw' in detailops or 'NONE' in detailops:  
                for interfaces in n7k_fw_int[district]['OUTER'][nexusvdc][dc][fw]:
                    fwint = interfaces['int']
                
            
                    if configure is True and ( 'fw' in detailops or 'NONE' in detailops) :
                        curr_allowed_list = send_to_n7k_api_show("show int %s switchport" % (fwint),device_ip,district,dc,nexusvdc,device_un,device_pw)
                    else:
                        curr_allowed_list = 'none'
       
                    if curr_allowed_list == 'none' or curr_allowed_list == '1-4094':
                        commands.append("interface %s" % (fwint))
                        commands.append("description TO_%s" % (fw) )
                        commands.append("switchport")
                        commands.append("switchport mode trunk")
                        commands.append("switchport trunk allow vlan " + ','.join(map(str,vlans)))
                        commands.append("no shutdown")
                        commands.append("!")
                    else:
                        commands.append("interface %s" % (fwint))
                        commands.append("switchport")
                        commands.append("switchport mode trunk")
                        commands.append("switchport trunk allow vlan add " + ','.join(map(str,vlans)))
                        commands.append("no shutdown")
                        commands.append("!")
                    
        if 'fw' in detailops or 'NONE' in detailops:   
            print "!"
            print "!"
            print "! Allow VLANs on the firewall"
            print "!"    
            print '\n'.join(map(str,commands))
            print "!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"
            print ""
        
            if configure is True and ('fw' in detailops or 'NONE' in detailops ):
                            
                commands = " ; ".join(map(str,commands))
                print "*** sending above config to %s,%s,%s ***"  %(dc,district,nexusvdc)
                send_to_n7k_api(device_ip,commands,district,dc,nexusvdc,device_un,device_pw)
            commands = []
        if configure is True:
            print "DONE!"
 
def send_to_n7k_api (ip,commands,district,dc,nexusvdc,username,password):
   
    content_type = "json"
    HTTPS_SERVER_PORT = "8080"

    requests.packages.urllib3.disable_warnings()
    
    if commands.endswith(" ; "):
        commands = commands[:-3]
        
    payload = {
        "ins_api": {
            "version": "1.2",
            "type": "cli_conf",
            "chunk": "0",               # do not chunk results
            "sid": "1",
            "input": commands,
            "output_format": "json"
        }
    }
    
    headers={'content-type':'application/%s' % content_type}
    response = requests.post("https://%s:%s/ins" % (ip, HTTPS_SERVER_PORT),
                             auth=(username, password),
                             headers=headers,
                             data=json.dumps(payload),
                             verify=False,                      # disable SSH certificate verification
                             timeout=30)
    
    if response.status_code == 200:
        # verify result if a cli_conf operation was performed
        if "ins_api" in payload:
            if "type" in payload['ins_api'].keys():
                if "cli_conf" in payload['ins_api']['type']:
                    if isinstance(response.json()['ins_api']['outputs']['output'],dict):
                        code =    response.json()['ins_api']['outputs']['output']['code']
                        message = response.json()['ins_api']['outputs']['output']['msg']
                        detail  = response.json()['ins_api']['outputs']['output']['clierror']
                        allcmds = commands.split(" ; ")
                        if code != 200:
                            print("ERROR: partial configuration failed on %s,%s,%s,%s, please verify your configuration!") % (district,dc,nexusvdc,ip)
                            print("ERROR: message is: %s") % (message)
                            print("ERROR: Detailed message is: %s") % (detail)
                            print("ERROR: Failed command is: %s") % allcmds
                            sys.exit(10)
                    else:
                        for result in response.json()['ins_api']['outputs']['output']:
                            if result['code'] != "200":
                                message = result['msg']
                                detail  = result['clierror']
                                command_number  = response.json()['ins_api']['outputs']['output'].index(result)
                                allcmds = commands.split(" ; ")
                                print("ERROR: partial configuration failed on %s,%s,%s,%s, please verify your configuration!") % (district,dc,nexusvdc,ip)
                                print("ERROR: message is: %s") % (message)
                                print("ERROR: Detailed message is: %s") % (detail)
                                print("ERROR: Failed command is: %s") % allcmds[command_number]
                                sys.exit(10)
                            
    else:
        msg = "call to %s failed, status code %d (%s).  Command is %s.  %s,%s,%s" % (ip,
                                                          response.status_code,
                                                          response.content.decode("utf-8"),
                                                          commands,
                                                          district,
                                                          dc,
                                                          nexusvdc
                                                          )
        print(msg)
        sys.exit(10)
        #raise Exception(msg) 
       
def send_to_n7k_api_show(commands, ip,district,dc,nexusvdc,username,password):

    payload = [
        {
            "jsonrpc": "2.0",
            "method": "cli",
            "params": {
                "cmd": commands,
                "version": 1.2
            },
            "id": 1
        }
    ]

    content_type = "json-rpc"
    HTTPS_SERVER_PORT = "8080"

    requests.packages.urllib3.disable_warnings()

    if commands.endswith(" ; "):
        commands = commands[:-3]

    headers={'content-type':'application/%s' % content_type}
    response = requests.post("https://%s:%s/ins" % (ip, HTTPS_SERVER_PORT),
                             auth=(username, password),
                             headers=headers,
                             data=json.dumps(payload),
                             verify=False,                      # disable SSH certificate verification
                             timeout=30)
    
    if response.status_code == 200:
        if  bool((re.search('switchport',commands,re.IGNORECASE))) :
            return response.json()['result']['body']['TABLE_interface']['ROW_interface']['trunk_vlans']
        else:
            return response.json()['result']
    else:
        msg = "call to %s failed, status code %d (%s).  Command is %s.  %s,%s,%s" % (ip,
                                                          response.status_code,
                                                          response.content.decode("utf-8"),
                                                          commands,
                                                          district,
                                                          dc,
                                                          nexusvdc
                                                          )
        print(msg)
        sys.exit(10)
        #raise Exception(msg)
        
def get_outer_to_pa(wb):
        
        ws = wb.active 
        row_start = ws.min_row
        row_end   = ws.max_row
        data = {}

        # Process outer VDC to PA

        for x in range(row_start,row_end+1):
            # Get district and tenant
            cell = 'A' + str(x)
            value = ws[cell].value 
            
            if value is not None and  \
            not bool((re.search('Addressing',value,re.IGNORECASE))) and \
            not bool((re.search('Mask',value,re.IGNORECASE))) and \
            value != "vSYS":
                value = value.strip()

                if value == 'SDE':
                    district = value
                elif value == 'SOE':
                    district = value
                elif value == 'GIS':
                    district = value
                else:
                    tenant = value
                    tenant = tenant.upper()
                    tenant = tenant.replace(" ","_")
                    tenant = tenant.replace('_X000D__X000D_','')
            # Get PA IP DC1
            cell = 'B' + str(x)
            value = ws[cell].value 
            
            if value is None or value == 'DC1':
                continue
           
            if not bool((re.search('FW',value,re.IGNORECASE))):
                
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1paip = value
                    
            # Get N7K IP DC1
            cell = 'C' + str(x)
            value = ws[cell].value 
            
            if value is None or value == 'DC1':
                continue
           
            if not bool((re.search('VDC',value,re.IGNORECASE))) and not bool((re.search('Address',value,re.IGNORECASE))):
                
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1n7kip = value
            
            # Get N7K VDC
            cell = 'D' + str(x)
            value = ws[cell].value 
            
            if value is None:
                continue

            value = value.strip()
            n7k = value
            
            # Get PA IP DC2
            cell = 'E' + str(x)
            value = ws[cell].value 
            
            if value is None or value == 'DC2':
                continue
           
            if not bool((re.search('FW',value,re.IGNORECASE))):
                
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2paip = value
            
            # Get N7K IP DC2
            cell = 'F' + str(x)
            value = ws[cell].value 
            
            if value is None or value == 'DC2':
                continue
           
            if not bool((re.search('Address',value,re.IGNORECASE))):
                
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2n7kip = value
            
            # If tenant and district exist, append attributes as a list
            
            if district in data:
                if tenant in data[district]:
                    if n7k in data[district][tenant]:
                            data[district][tenant][n7k].append(
                                    {
                                     'dc1paip'     : dc1paip,
                                     'dc1n7kip'    : dc1n7kip,
                                     'dc2paip'     : dc2paip,
                                     'dc2n7kip'    : dc2n7kip
                                   })
                            # If district exists, but not tenant, add new key (tenant) and initial attributes
                    else:
                            data[district][tenant][n7k] = [ {  
                                     
                                     'dc1paip'     : dc1paip,
                                     'dc1n7kip'    : dc1n7kip,
                                     'dc2paip'     : dc2paip,
                                     'dc2n7kip'    : dc2n7kip
                                     
                                     } ]
                            
                else:
                    data[district][tenant] = {}        
                    data[district][tenant][n7k] = [ {  
                                  'dc1paip'     : dc1paip,
                                  'dc1n7kip'    : dc1n7kip,
                                  'dc2paip'     : dc2paip,
                                   'dc2n7kip'    : dc2n7kip
                                } ]

            # Initial key/value assignment
            else:
                    data.update({  
                              district :  
                              { 
                                tenant :  
                                    { 
                                      n7k : [ 
                                     {
                                     'dc1paip'     : dc1paip,
                                     'dc1n7kip'    : dc1n7kip,
                                     'dc2paip'     : dc2paip,
                                     'dc2n7kip'    : dc2n7kip
                                    }
                                    ] 
                                }
                              } 
                         })
        

        return data

def get_loopback(wb):
        
        ws = wb.active 
        row_start = ws.min_row
        row_end   = ws.max_row
        data = {}
        
        # Process Loopback assignments

        for x in range(row_start,row_end+1):

            # Get IP DC1
            cell = 'B' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('IP',value,re.IGNORECASE))) and value != 'DC1':
            
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1ip = value

            # Get hostname DC1
            cell = 'C' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('Hostname',value,re.IGNORECASE))) and value != 'DC1':
                dc1hn = value
                dc1hn = dc1hn.strip()
                
            # Get Description
            cell = 'D' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('Description',value,re.IGNORECASE))) and value != 'DC1':
                dc1desc = value
                dc1desc = dc1desc.strip()
                
            # Get IP DC2
            cell = 'E' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('IP',value,re.IGNORECASE))) and value != 'DC2':
            
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2ip = value

            # Get hostname DC2
            cell = 'F' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('Hostname',value,re.IGNORECASE))) and value != 'DC2':
                dc2hn = value
                dc2hn = dc2hn.strip()
            
            # IF row is blank, skip
            else:
                continue
                
            # Get Description
            cell = 'G' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('Description',value,re.IGNORECASE))) and value != 'DC2':
                dc2desc = value
                dc2desc = dc2desc.strip()

            # Push all data to dictionary
            # Use Debug option to print data

            if bool((re.search('soe',value,re.IGNORECASE))):
                district = ['SOE']

            if bool((re.search('sde',value,re.IGNORECASE))):
                district = ['SDE']
            
            if bool((re.search('gis',value,re.IGNORECASE))):
                district = ['GIS']
                
            if bool((re.search('soe/gis',value,re.IGNORECASE))):
                district = ['GIS','SOE']
            
            for d in district: 
                # If district exist, append attributes as a list
                if d in data:
                         data[d].append(
                                     {
                                      'dc1ip'         : dc1ip,
                                      'dc1desc'       : dc1desc,
                                      'dc1hn'         : dc1hn,
                                      'dc2ip'         : dc2ip,
                                      'dc2desc'       : dc2desc,
                                      'dc2hn'         : dc2hn
                                     })
            
                # Initial key/value assignment
                else:
                 data.update({  
                      d :  
                                 [ 
                                  {
                                      'dc1ip'         : dc1ip,
                                      'dc1desc'       : dc1desc,
                                      'dc1hn'         : dc1hn,
                                      'dc2ip'         : dc2ip,
                                      'dc2desc'       : dc2desc,
                                      'dc2hn'         : dc2hn
                                 }
                            ] 
                           } )
        return data
    
def get_outer_jnp(wb):
        
        ws = wb.active 
        row_start = ws.min_row
        row_end   = ws.max_row
        data = {}
        
        # Process outer to PA

        for x in range(row_start,row_end+1):
            # Get N7K or district
            cell = 'A' + str(x)
            value = ws[cell].value 
            
            if value is not None:
                if bool((re.search('SOE/GIS',value,re.IGNORECASE))):
                    district = ['SOE','GIS']
            
                elif bool((re.search('N7K-',value,re.IGNORECASE))):
                    n7kname = value
            
                elif bool((re.search('SDE',value,re.IGNORECASE))):
                    district = ['SDE']
                    del n7kname
                    
                else:
                    continue
            
            else:
                try:
                    n7kname
                except:
                    continue
                
                
            # Get n7k IP address
            cell = 'B' + str(x)
            value = ws[cell].value 
            
            if value is not None:
                
                try:
                    IP(value)
                except:
                    pass
                else:
                    dc1n7kip = value
                    
            else:
                continue      
                
            cell = 'C' + str(x)
            value = ws[cell].value 
            
            if value is not None:
                
                try:
                    IP(value)
                except:
                    pass
                else:
                    dc1jnpip = value
                    
            cell = 'D' + str(x)
            value = ws[cell].value 
            
            if value is not None and bool((re.search('JNP',value,re.IGNORECASE))):
                desc = value
                desc = desc.strip()
            
            cell = 'E' + str(x)
            value = ws[cell].value 
            
            if value is not None:
                
                try:
                    IP(value)
                except:
                    pass
                else:
                    dc2n7kip = value
                    
            cell = 'F' + str(x)
            value = ws[cell].value 
            
            if value is not None:
                
                try:
                    IP(value)
                except:
                    pass
                else:
                    dc2jnpip = value
                    

            # Push all data to dictionary
            # Use Debug option to print data

            
            for d in district: 
                # If district exist, append attributes as a list
                if d in data:
                    if n7kname in data[d]:
                        if desc in data[d][n7kname]:
                            data[d][n7kname][desc].append(
                                     {
                                      'dc1n7kip'       : dc1n7kip,
                                      'dc1jnpip'       : dc1jnpip,
                                      'dc2n7kip'       : dc2n7kip,
                                      'dc2jnpip'       : dc2jnpip
                                     })
                        else:
                            data[d][n7kname][desc] = {}
                            data[d][n7kname][desc] = [ {
                                    
                                      'dc1n7kip'       : dc1n7kip,
                                      'dc1jnpip'       : dc1jnpip,
                                      'dc2n7kip'       : dc2n7kip,
                                      'dc2jnpip'       : dc2jnpip
                            
                                        }]
                    else:
                        data[d][n7kname] = {}
                        data[d][n7kname][desc] = {}
                        data[d][n7kname][desc] = [ {
                                    
                                      'dc1n7kip'       : dc1n7kip,
                                      'dc1jnpip'       : dc1jnpip,
                                      'dc2n7kip'       : dc2n7kip,
                                      'dc2jnpip'       : dc2jnpip
                            
                                        }]
            
                # Initial key/value assignment
                else:
                 data.update({  
                      d :  
                            {
                               n7kname : 
                               { 
                                   desc:
                                 [ 
                                  {
                                      'dc1n7kip'       : dc1n7kip,
                                      'dc1jnpip'       : dc1jnpip,
                                      'dc2n7kip'       : dc2n7kip,
                                      'dc2jnpip'       : dc2jnpip
                                 }
                                ] 
                               }
                            }
                        } )
        return data
            
def get_inner_to_pa(wb,district):
        
        ws = wb.active 
        row_start = ws.min_row
        row_end   = ws.max_row
        data = {}
        
        if district in ("SOE","GIS"):
            max_n7k = 4
        else:
            max_n7k = 2

        # Process ws_soe_inner_to_pa
        for x in range(row_start,row_end+1):
            # Get VRF#
            cell = 'A' + str(x)
            value = ws[cell].value 
            
            if value is not None and bool((re.search('VRF-',value,re.IGNORECASE))):
                vrfnumber = value
            
            # Get description
            cell = 'B'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('Connection',value,re.IGNORECASE))):
                connectdesc = value
                temp = connectdesc[3]
                connectdesc = connectdesc[:3] + "-" + temp

            
            # get N7K IP DC1
            cell = 'C'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('IP Add',value,re.IGNORECASE))) and not bool((re.search('DC',value,re.IGNORECASE))):
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1n7kip = value
            
            # get PA FW IP DC1
            cell = 'D'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('IP Add',value,re.IGNORECASE))) and not bool((re.search('DC',value,re.IGNORECASE))):
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1pafwip = value
            
            # get sub zone name
            cell = 'E'  + str(x)
            value = ws[cell].value
               
            if value != 0 and value is not None and not bool((re.search('Zone',value,re.IGNORECASE))):
                subzone = value
                subzone = subzone.strip()
                subzone = subzone.replace(' ',"_")
                subzone = subzone.upper()
                
                
            # get N7K VRF Name DC1
            cell = 'F'  + str(x)
            value = ws[cell].value
            
            if value != 0 and value is not None and value != 'DC1 ' and value != 'DC2' and not bool((re.search('N7K',value,re.IGNORECASE))) :
                n7kvrfdc1 = value
                n7kvrfdc1 = n7kvrfdc1.strip()
            
            # get N7K VRF Name DC2
            cell = 'G'  + str(x)
            value = ws[cell].value
            
            if value != 0 and value is not None and value != 'DC1 ' and value != 'DC2' and not bool((re.search('N7K',value,re.IGNORECASE))) :
                n7kvrfdc2 = value
                n7kvrfdc2 = n7kvrfdc2.strip()

            # get N7K IP DC2
            cell = 'H'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('IP Add',value,re.IGNORECASE))) and not bool((re.search('DC',value,re.IGNORECASE))):
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2n7kip = value
            
            # get PA FW IP DC2
            cell = 'I'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('IP Add',value,re.IGNORECASE))) and not bool((re.search('DC',value,re.IGNORECASE))):
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2pafwip = value
                   
            # Push all data to dictionary
            # Use Debug option to print data
                
            # If subzone and connection exist, append attributes as a list
                if district in data:
                    if subzone in data[district]:
                        if connectdesc in data[district][subzone]:
                            if len(data[district][subzone]) < max_n7k:
                                data[district][subzone][connectdesc].append(
                                        {
                                         'vrfnumber'     : vrfnumber,
                                         'dc1n7kip'      : dc1n7kip,
                                         'dc1pafwip'     : dc1pafwip,
                                         'n7kvrfdc1'     : n7kvrfdc1,
                                         'n7kvrfdc2'     : n7kvrfdc2,
                                         'dc2n7kip'      : dc2n7kip,
                                         'dc2pafwip'     : dc2pafwip
                                        })
                # If district exists, but not tenant, add new key (tenant) and initial attributes
                        else:
                                data[district][subzone][connectdesc] = [ {  
                                  'vrfnumber'     : vrfnumber,
                                  'dc1n7kip'      : dc1n7kip,
                                  'dc1pafwip'     : dc1pafwip,
                                  'n7kvrfdc1'     : n7kvrfdc1,
                                  'n7kvrfdc2'     : n7kvrfdc2,
                                  'dc2n7kip'      : dc2n7kip,
                                  'dc2pafwip'     : dc2pafwip
                                } ]
                    else:
                        data[district][subzone] = {}        
                        data[district][subzone][connectdesc] = [ {  
                                  'vrfnumber'     : vrfnumber,
                                  'dc1n7kip'      : dc1n7kip,
                                  'dc1pafwip'     : dc1pafwip,
                                  'n7kvrfdc1'     : n7kvrfdc1,
                                  'n7kvrfdc2'     : n7kvrfdc2,
                                  'dc2n7kip'      : dc2n7kip,
                                  'dc2pafwip'     : dc2pafwip
                                } ]
                        

                # Initial key/value assignment
                else:
                    data.update({  
                         district :  
                                 { 
                                      subzone : 
                                             { 
                                         
                                                connectdesc :  [ 
                                                         {
                                                          'vrfnumber'     : vrfnumber,
                                                          'dc1n7kip'      : dc1n7kip,
                                                          'dc1pafwip'     : dc1pafwip,
                                                          'n7kvrfdc1'     : n7kvrfdc1,
                                                          'n7kvrfdc2'     : n7kvrfdc2,
                                                          'dc2n7kip'      : dc2n7kip,
                                                          'dc2pafwip'     : dc2pafwip
                                                          }
                                                         ] 
                                              }
                                 } 
                    })
   

        return data

def getfwint(wb,dc,data):
    ws = wb.active 
    row_start = ws.min_row
    row_end   = ws.max_row
    
    for cells in ws.iter_rows(min_row=row_start, min_col=1, max_col=24):
        for vals in cells:
            
            if vals.value == 'A':
                n7k = 'N7K-A'
                
            if vals.value == 'B':
                n7k = 'N7K-B'
            
            if vals.value == 'C':
                n7k = 'N7K-C'
                
            if vals.value == 'D':
                n7k = 'N7K-D'
                
            if vals.value == 'E':
                n7k = 'N7K-E'
                
            if vals.value == 'F':
                n7k = 'N7K-F'
            
            
            
            if bool(re.search(dc,str(vals.value), re.IGNORECASE)):
                vals.value = vals.value.replace(" ","")
                fwname   = vals.value[:-6].lower()
                district = vals.value[3:6].upper()
                dloc     = vals.value[-5:].upper()
                
                if bool(re.search(dc, ws[vals.column + str(vals.row-1)].value, re.IGNORECASE)):
                    interface = ws[vals.column + str(vals.row+1)].value
                else:
                    interface = ws[vals.column + str(vals.row-1)].value
                    
                interface = "E" + interface
                interface = interface.replace("-","/")
                            
               
                if district in data:
                    if dloc in data[district]:
                        if n7k in data[district][dloc]:
                            if dc in data[district][dloc][n7k]:
                                if fwname in data[district][dloc][n7k][dc]:
                                    data[district][dloc][n7k][dc][fwname].append({'int'  : interface })
                                else:
                                    data[district][dloc][n7k][dc][fwname] = [{'int' : interface}]
                            else:
                                data[district][dloc][n7k][dc] = {}
                                data[district][dloc][n7k][dc][fwname] = [{ 'int' : interface}]
                        else:
                            data[district][dloc][n7k] = {}
                            data[district][dloc][n7k][dc] = {}
                            data[district][dloc][n7k][dc][fwname] = [{ 'int' : interface}]
                    else:
                        data[district][dloc] = {}
                        data[district][dloc][n7k] = {}
                        data[district][dloc][n7k][dc] = {}
                        data[district][dloc][n7k][dc][fwname] = [{ 'int' : interface}]                 
                                    
                else:
                    data.update({  
                            district :  
                            { 
                            dloc :  
                            {   
                                    n7k :
                                    {
                                     dc : 
                                        {
                                            fwname :
                                            [ 
                                                {
                                                'int'     : interface
                                                }
                                            ] 
                                         }
                                             
                                }}}})
               
    return data


def process_xlsx(filename,dc1portmap,dc2portmap,debug):
    worksheets = []
    ws_definition_data = {}

    # Important worksheets
    ws_definition = "SOE_SDE_GIS_VRF_RT_Definition"

    wb = openpyxl.load_workbook(filename, data_only=True)

    # Get all worksheets
    for sheet in wb:
        worksheets.append(sheet.title) 
    wb.close()

    # Set ws_definition as the active worksheet to read from
    # Process SOE_SDE_GIS_VRF_RT_Definition
    
   
    try:
        wb.active = worksheets.index(ws_definition)
    except:
        print "Worksheet %s not found" % ws_definition
    else:
        ws = wb.active
        row_start = ws.min_row
        row_end   = ws.max_row

        # Process ws_definition tab 
        for x in range(row_start,row_end+1):
           
            # Get District
            cell = 'A' + str(x)
            value = ws[cell].value 
            
            if value is not None and value != 'District':
                district = value

            # Go to next row if heading
            if value == 'District':
                continue

            # Get Zone
            cell = 'B' + str(x)
            value = ws[cell].value 
           
            # Go to next row if heading
            if value is not None and bool((re.search('Zone',value,re.IGNORECASE))):
                continue

            if value is not None and value != 'Sub Zone' and value != 'Network Segmentation Zone' and not bool(re.search('Zone',value, re.IGNORECASE)):
                tenant = value
                tenant = tenant.strip()
                tenant = tenant.replace(" ","_")
                tenant = tenant.upper()
         
            # Get Sub Zone
            cell = 'C' + str(x)
            value = ws[cell].value 
           
            # Skip row if there is no value (empty row)
            if value is None:
               continue
            if value is not None and not bool(re.search('Sub Zone',value, re.IGNORECASE)):
                szone = value
                szone = szone.strip()
                szone = szone.replace(" ","_")
                szone = szone.upper()
            
            # Get Firewall (Inside Cell, Internal, Mainframe, Money Movement
            # 01-24-2018: Only get Internal FW
            cell = 'D' + str(x)
            value = ws[cell].value 
            
            if value is not None and bool(re.search('Internal',value, re.IGNORECASE)):
                fwtype = value
            else:
                continue
            
            # Get firewall names - dc1 names listed.  Assume dc2 fw names are the same, except they start with dc2
            cell = 'E' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('Pri',value, re.IGNORECASE)):
                    dc1prifwname = value
                    dc1prifwname = dc1prifwname.strip()
                    
                    dc2prifwname = dc1prifwname.replace("dc1","dc2")
                    
            # Get firewall names - dc1 names listed.  Assume dc2 fw names are the same, except they start with dc2
            cell = 'F' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('Sby',value, re.IGNORECASE)):
                    dc1sbyfwname = value
                    dc1sbyfwname = dc1sbyfwname.strip()
                    
                    dc2sbyfwname = dc1sbyfwname.replace("dc1","dc2")     
            
            # Get VRF #
            cell = 'G' + str(x)
            value = ws[cell].value 
            
            if value is not None and value != 'VRF' and value != '#':
                    vrf = value
            
            
            ## If config is NO, skip the entire row. We are not configuring that VRF
            cell = 'H' + str(x)
            value = ws[cell].value 
            
            if value is not None and bool((re.search('no',value,re.IGNORECASE))):
                config = 'no'
            else:
                config = 'yes'
            
            
            # Get VRF Name for DC1 - skip entire row if VRF name for DC1 is blank 
            cell = 'I' + str(x)
            value = ws[cell].value 
            
            if value is None:
                continue
            
            else:
                    if  not bool(re.search('N7K',value, re.IGNORECASE)) and value != 'DC1' and value != 'DC2':
                        vrfnamedc1 = value
                        vrfnamedc1 = vrfnamedc1.strip()

            # Get VRF Name for DC2
            cell = 'J' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('N7K',value, re.IGNORECASE)) and value != 'DC1' and value != 'DC2':
                    vrfnamedc2 = value
                    vrfnamedc2 = vrfnamedc2.strip()
            
            # RT DC1
            cell = 'K' + str(x)
            value = ws[cell].value 
            
            if value is not None and value != 'RT' and value != 'DC1' and value != 'DC2':
                    rtdc1 = value
                    rtdc1 = rtdc1.strip()

            # RT DC2
            cell = 'L' + str(x)
            value = ws[cell].value 
            
            if value is not None and value != 'RT' and value != 'DC1' and value != 'DC2':
                    rtdc2 = value
                    rtdc2 = rtdc2.strip()
            
            # inner VDC to FW encap

            cell = 'M' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('encapsulation',str(value))) and not bool(re.search('Inner',str(value))):
                    invdcencap = value
            
            # OSPF DC1 
            cell = 'N' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('OSPF',str(value), re.IGNORECASE)) and not bool(re.search('DC',str(value), re.IGNORECASE)):
                    ospfdc1 = value
            
            # OSPF DC2 
            cell = 'O' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('OSPF',str(value), re.IGNORECASE)) and not bool(re.search('DC',str(value), re.IGNORECASE)):
                    ospfdc2 = value


            # encap - outer VDC to FW
            cell = 'P' + str(x)
            value = ws[cell].value 
            
            if not bool(re.search('encapsulation',str(value), re.IGNORECASE)) and not bool(re.search('Inside',str(value), re.IGNORECASE)):
                     outvdcencap = getValueWithMergeLookup(ws, cell)

                     
           # Push data into dictionary
           # Use Debug option to print data

            # If tenant and district exist, append attributes as a list
            if district in ws_definition_data:
                if tenant in ws_definition_data[district]:
                  ws_definition_data[district][tenant].append(
                                    {
                                     'vrfnumber'     : vrf,
                                     'subzone'       : szone,
                                     'firewall'      : fwtype,
                                     'dc1vrf'        : vrfnamedc1,
                                     'dc2vrf'        : vrfnamedc2,
                                     'rtdc1'         : rtdc1,
                                     'rtdc2'         : rtdc2,
                                     'innervdcencap' : invdcencap,
                                     'ospfdc1'       : ospfdc1,
                                     'ospfdc2'       : ospfdc2,
                                     'config'        : config,
                                     'outervdcencap' : outvdcencap,
                                     'dc1prifwname'  : dc1prifwname,
                                     'dc1sbyfwname'  : dc1sbyfwname,
                                     'dc2prifwname'  : dc2prifwname,
                                     'dc2sbyfwname'  : dc2sbyfwname
                                   })
                # If district exists, but not tenant, add new key (tenant) and initial attributes
                else:
                    ws_definition_data[district][tenant] = [ {  
                                     
                                      'vrfnumber'     : vrf,
                                      'subzone'       : szone,
                                      'firewall'      : fwtype,
                                      'dc1vrf'        : vrfnamedc1,
                                      'dc2vrf'        : vrfnamedc2,
                                      'rtdc1'         : rtdc1,
                                      'rtdc2'         : rtdc2,
                                      'innervdcencap' : invdcencap,
                                      'ospfdc1'       : ospfdc1,
                                      'ospfdc2'       : ospfdc2,
                                      'config'        : config,
                                      'outervdcencap' : outvdcencap,
                                      'dc1prifwname'  : dc1prifwname,
                                      'dc1sbyfwname'  : dc1sbyfwname,
                                      'dc2prifwname'  : dc2prifwname,
                                      'dc2sbyfwname'  : dc2sbyfwname
                                     } ]

            # Initial key/value assignment
            else:
                    ws_definition_data.update({  
                              district :  
                              { 
                                tenant :  [ 
                                     {
                                      'vrfnumber'     : vrf,
                                      'subzone'       : szone,
                                      'firewall'      : fwtype,
                                      'dc1vrf'        : vrfnamedc1,
                                      'dc2vrf'        : vrfnamedc2,
                                      'rtdc1'         : rtdc1,
                                      'rtdc2'         : rtdc2,
                                      'innervdcencap' : invdcencap,
                                      'ospfdc1'       : ospfdc1,
                                      'ospfdc2'       : ospfdc2,
                                      'config'        : config,
                                      'outervdcencap' : outvdcencap,
                                      'dc1prifwname'  : dc1prifwname,
                                      'dc1sbyfwname'  : dc1sbyfwname,
                                      'dc2prifwname'  : dc2prifwname,
                                      'dc2sbyfwname'  : dc2sbyfwname
                                    }
                                ] 
                              } 
                         })
       
        if debug == True :
            print '{ ' +  'SOE_SDE_GIS_VRF_RT_Definition' + ':'  + json.dumps(ws_definition_data) + ' } '
            

    ##############################################################################################

    # process <district>_inner_to_pa worksheets - format is similar for all 3 so use 1 function
    
        ws_soe_inner_to_pa = "SOE-VRF P2P Inner-to-PA"
        ws_sde_inner_to_pa = "SDE VRF P2P Inner-to-PA"
        ws_gis_inner_to_pa = "GIS VRF P2P Inner-to-PA"
  
        inner_ws = [ ws_soe_inner_to_pa,ws_sde_inner_to_pa,ws_gis_inner_to_pa]
        final_all_inner_data = {}

        for inner in inner_ws: 

            try:
                    wb.active = worksheets.index(inner)
            except:
                    print "Worksheet %s not found" % inner
                    sys.exit(9)
            else:
                    district = inner[:3]
                    inner_data = get_inner_to_pa(wb,district)
                    final_all_inner_data.update(inner_data)
        
        if debug == True :
           print '{ ' +  'inner_to_pa' + ':'  + json.dumps(final_all_inner_data) + ' } '
            
    
    ############################################################################################## 

    # Load BGP ASN numbers - this is hardcoded as it will not change
    bgp_asn = {}

    bgp_asn =  { 'SOE' : { 'Outer' : { 'dc1' :  65500, 'dc2' : 65510 } } }
    bgp_asn['SOE'].update({'Inner' : { 'dc1' :  65501, 'dc2' : 65511 } } )
    
    bgp_asn.update({'GIS' : { 'Inner' : { 'dc1' :  65502, 'dc2' : 65512 } } } )
    bgp_asn['GIS'].update(  { 'Outer' : { 'dc1' :  65500, 'dc2' : 65510 } } )

    bgp_asn.update({'SDE' : { 'Inner' : { 'dc1' :  65506, 'dc2' : 65516 } } } )
    bgp_asn['SDE'].update(  { 'Outer' : { 'dc1' :  65505, 'dc2' : 65515 } } )
   
    if debug == True:
        print '{ ' +  'bgp_asn' + ':'  + json.dumps(bgp_asn) + ' } '

    #############################################################################################

    # Load Loopback assignments
    ws_loopback = "Loopback Assignment"
    loopback_data = {}

    try:
            wb.active = worksheets.index(ws_loopback)
    except:
            print "Worksheet %s not found" % ws_loopback
            sys.exit(9)
    else:
            loopback_data = get_loopback(wb)
      
    if debug == True :
       print '{ ' +  'loopback' + ':'  + json.dumps(loopback_data) + ' } '


    ##############################################################################################
    # Load outer VDC to PA FW
    ws_outer_to_pa = "P2P-Outer-VDC"
    outer_to_pa_data = {}

    try:
            wb.active = worksheets.index(ws_outer_to_pa)
    except:
            print "Worksheet %s not found" % ws_outer_to_pa
            sys.exit(9)
    else:
            outer_to_pa_data = get_outer_to_pa(wb)
      
    if debug == True :
       print '{ ' +  'outer_to_pa' + ':'  + json.dumps(outer_to_pa_data) + ' } '

    ##############################################################################################
    
    # Interfaces to PA Firewall - hardcoded - this will most likely not change
    pm = openpyxl.load_workbook(dc1portmap, data_only=True)
    sheetname = '7706'
    wsheets = []
    n7k_fw_int = {}

    # Get all worksheets
    for sheet in pm:
        wsheets.append(sheet.title) 
    pm.close()

    try:
        pm.active = wsheets.index(sheetname)
    except:
        print "Worksheet %s not found in workbook %s" % (sheetname,dc1portmap)
    else:
        n7k_fw_int = getfwint(pm,'dc1',n7k_fw_int)
    
    pm.close()
    
    # DC2  
    pm = openpyxl.load_workbook(dc2portmap, data_only=True)
    

    # Get all worksheets
    for sheet in pm:
        wsheets.append(sheet.title) 
    pm.close()

    try:
        pm.active = wsheets.index(sheetname)
    except:
        print "Worksheet %s not found in workbook %s" % (sheetname,dc2portmap)
    else:
        n7k_fw_int = getfwint(pm,'dc2',n7k_fw_int)
    
    pm.close()  
    
      
    '''
    n7k_fw_int = {}

    n7k_fw_int =  { 'SOE' : { 'Outer' : { 'N7K-A' :  { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-B' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-C' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'}  }, 'N7K-D' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} } }  } }
    n7k_fw_int['SOE'].update({'Inner' : { 'N7K-A' :  { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-B' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-C' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'}  }, 'N7K-D' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} } }   })
    
    n7k_fw_int.update({'GIS' : { 'Outer' : { 'N7K-A' :  { 'dc1'  : { 'int1' : 'E2/22', 'int2' : 'E2/30' }, 'dc2' : { 'int1' : 'E2/22', 'int2' : 'E2/30'} }, 'N7K-B' : { 'dc1'  : { 'int1' : 'E2/22', 'int2' : 'E2/30' }, 'dc2' : { 'int1' : 'E2/22', 'int2' : 'E2/30'} }, 'N7K-C' : { 'dc1'  : { 'int1' : 'E2/22', 'int2' : 'E2/30' }, 'dc2' : { 'int1' : 'E2/22', 'int2' : 'E2/30'}  }, 'N7K-D' : { 'dc1'  : { 'int1' : 'E2/22', 'int2' : 'E2/30' }, 'dc2' : { 'int1' : 'E2/22', 'int2' : 'E2/30'} } }  } })
    n7k_fw_int['GIS'].update({'Inner' : { 'N7K-A' :  { 'dc1'  : { 'int1' : 'E2/13', 'int2' : 'E2/14' }, 'dc2' : { 'int1' : 'E2/13', 'int2' : 'E2/14'} }, 'N7K-B' : { 'dc1'  : { 'int1' : 'E2/13', 'int2' : 'E2/14' }, 'dc2' : { 'int1' : 'E2/13', 'int2' : 'E2/14'} }, 'N7K-C' : { 'dc1'  : { 'int1' : 'E2/13', 'int2' : 'E2/14' }, 'dc2' : { 'int1' : 'E2/13', 'int2' : 'E2/14'}  }, 'N7K-D' : { 'dc1'  : { 'int1' : 'E2/13', 'int2' : 'E2/14' }, 'dc2' : { 'int1' : 'E2/13', 'int2' : 'E2/14'} } }   })
    
    n7k_fw_int.update({'SDE' : { 'Outer' : { 'N7K-E' :  { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-F' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }  } } })
    n7k_fw_int['SDE'].update({'Inner' : { 'N7K-E' :  { 'dc1'  : { 'int1' : 'E2/5', 'int2' : 'E2/13' }, 'dc2' : { 'int1' : 'E2/5', 'int2' : 'E2/13'} }, 'N7K-F' : { 'dc1'  : { 'int1' : 'E2/5', 'int2' : 'E2/13' }, 'dc2' : { 'int1' : 'E2/5', 'int2' : 'E2/13'} } }   })
    
    '''
        
    
    if debug == True:
        print '{ ' +  'portmap' + ':'  + json.dumps(n7k_fw_int) + ' } '
    
    
    ##############################################################################################
    # Outer Interfaces to Juniper Routers
    ws_outer_juniper = "P2P JNP"
    outer_jnp_data = {}

    try:
            wb.active = worksheets.index(ws_outer_juniper)
    except:
            print "Worksheet %s not found" % ws_outer_juniper
            sys.exit(9)
    else:
            outer_jnp_data = get_outer_jnp(wb)
      
    if debug == True :
       print '{ ' +  'outer_to_juniper' + ':'  + json.dumps(outer_jnp_data) + ' } '
       
    ##############################################################################################
    
    ##############################################################################################
    # Get interfaces Outer -> Juniper
    
    out_to_jnp_fw_int = {"SOE": {"N7K-D": {"DC2": {"JNP-2": {"remint": "XE-0/1/7", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/7", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/7", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/7", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/1/7", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/7", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/7", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/7", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}, "N7K-B": {"DC2": {"JNP-2": {"remint": "XE-0/1/6", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/6", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/6", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/6", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/1/6", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/6", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/6", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/6", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}, "N7K-C": {"DC2": {"JNP-2": {"remint": "XE-0/0/7", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/7", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/7", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/7", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/0/7", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/7", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/7", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/7", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}, "N7K-A": {"DC2": {"JNP-2": {"remint": "XE-0/0/6", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/6", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/6", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/6", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/0/6", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/6", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/6", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/6", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}}, "GIS": {"N7K-D": {"DC2": {"JNP-2": {"remint": "XE-0/1/7", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/7", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/7", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/7", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/1/7", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/7", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/7", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/7", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}, "N7K-B": {"DC2": {"JNP-2": {"remint": "XE-0/1/6", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/6", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/6", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/6", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/1/6", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/6", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/6", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/6", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}, "N7K-C": {"DC2": {"JNP-2": {"remint": "XE-0/0/7", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/7", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/7", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/7", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/0/7", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/7", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/7", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/7", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}, "N7K-A": {"DC2": {"JNP-2": {"remint": "XE-0/0/6", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/6", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/6", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/6", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/0/6", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/6", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/6", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/6", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}}, "SDE": {"N7K-F": {"DC2": {"JNP-2": {"remint": "XE-0/1/8", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/8", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/8", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/8", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/1/8", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/1/8", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/1/8", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/1/8", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}, "N7K-E": {"DC2": {"JNP-2": {"remint": "XE-0/0/8", "devname": "dc2mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/8", "devname": "dc2mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/8", "devname": "dc2mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/8", "devname": "dc2mx480-pe4", "localint": "E2/26"}}, "DC1": {"JNP-2": {"remint": "XE-0/0/8", "devname": "dc1mx480-pe2", "localint": "E2/18"}, "JNP-3": {"remint": "XE-0/0/8", "devname": "dc1mx480-pe3", "localint": "E2/25"}, "JNP-1": {"remint": "XE-0/0/8", "devname": "dc1mx480-pe1", "localint": "E2/17"}, "JNP-4": {"remint": "XE-0/0/8", "devname": "dc1mx480-pe4", "localint": "E2/26"}}}}}
    
    if debug == True :
       print '{ ' +  'outer_to_juniper_fw_int' + ':'  + json.dumps(out_to_jnp_fw_int) + ' } '
    
    
    return ws_definition_data,final_all_inner_data,bgp_asn,outer_to_pa_data,n7k_fw_int,loopback_data,outer_jnp_data,out_to_jnp_fw_int

def usage():
    print "Usage: " +  sys.argv[0] + " -f|--file <excel file name> -d|--debug -e|--execute <n7k list> -w."
    print ""
    print "-d|--debug:  Prints excel data in JSON format (no switch changes made)"
    print "-f|--file:   Pass input file to use for configuration.   Must use -e option when using -f"
    print "-e|--execute <n7k list>: Invokes N7K API to configure N7K. Must use -f option when using -e.  <n7k list> file must be in this format:"
    print "        <SOE|GIS|SDE>,<DC1|DC2>,<N7K-A|N7K-B|N7K-C|N7K-D>,<IP>,<Username>,<Password>"
    print "-w|--write: Writes the config to the N7k."
    print "-p|--portmap: Pass portmap file to the script, 1 per DC, separated by comma.  NO spaces between commas.  Ex: -p Vanguard\ DC1\ -\ Port\ Map\ v1.23.xlsx,Vanguard\ DC2\ -\ Port\ Map\ v1.19.xlsx"
    print "      Each file passed to the -p option must have DC1 or DC2 in the name to identify the respective data center"
    print "-t|--type : Print parts of the config.  Valid options are bgp,vlan,fw,ospf,bb"
    sys.exit(1)
    

def main(argv):

   
    debug  = False
    configure = False
    errfound = 0
    procpm = False
    procexec = False
    procfile = False
    procdebug = False
    procwrite = False
    detailops = []
    
    if len(argv) == 0:
        usage()

    try:
        opts,args = getopt.getopt(argv,"f:hde:wp:t:",["file=","help","debug","execute=","write","portmap","type"])
    except getopt.GetoptError as err:
        print str(err)
        sys.exit(2)
    else:
        for opt,arg in opts:
            if opt in ("-f","--file"):
                procfile = True
            if opt in ("-d","--debug"):
                procdebug = True
            if opt in ("-w","--write"):
                procwrite = True
            if opt in ("-p","--portmap"):
                procpm = True
            if opt in ("-e","--execute"):
                procexec = True
            if opt in ("-t", "--type"):
                t = arg.split(",")
                for tt in t:
                    tt = tt.lower()
                    if tt not in ('bgp','ospf','vlan','fw','bb'):
                        print "invalid option passed to -t|--type: " + tt
                        print "Valid options are: bgp,ospf,vlan,fw,bb"
                        sys.exit(9)
                    detailops.append(tt)
        
        if len(detailops) == 0:
            detailops.append('NONE')
        else:
            detailops = list(set(detailops))
                
        if procfile is True and procpm is True and procexec is not True:
            print "Missing -e option"
            sys.exit(9)
        
        if procfile is False:
            print "Missing -f option"
            sys.exit(9)
        
        if procpm is False:
            print "Missing -p option"
            sys.exit(9)
        
        if procdebug is True and procwrite is True:
            print "Cannot use option d and option w together.  Use option w or option d"
            sys.exit(9)  
        
                        
        for opt,arg in opts:
            if opt in ("-d","--debug"):
                debug = True
            if opt in ("-w","--write"):
                configure = True
            if opt in ("-e", "--execute"):
                vdcfile = arg
                if not os.path.isfile(vdcfile):
                    print sys.argv[0] + " VDC File list %s NOT found" % vdcfile
                    sys.exit(1)
                else:
                    # Verify file is in the following format: DISTRICT,DC,N7K-[A|B|C|D|E|F],IP
                    with open (vdcfile) as data:
                        lines = data.read().splitlines()
                    
                    for data in lines:
                        d = data.split(",")
                        f_district = d[0]
                        f_dc       = d[1]
                        f_dev      = d[2]
                        f_ip       = d[3]
                       
                        
                        if f_district.upper() not in ('SOE','GIS','SDE'):
                            print "Incorrect district %s, line %s.  Expecting GIS, SOE or SDE." % (f_district,lines.index(data)+1)
                            errfound = 1
                            
                        if f_dc.upper() not in ('DC1','DC2'):
                            print "Incorrect Data Center %s, line %s.  Expecting DC1 or DC2" % (f_dc,lines.index(data)+1)
                            errfound = 1
                            
                        if f_district.upper() in ('SOE','GIS') and f_dev.upper()[4] not in ('A','B','C','D'):
                            print "Incorrect N7K device %s for district %s, line %s.  Expecting N7K-[A-D]" % (f_dev,f_district,lines.index(data)+1)
                            errfound = 1
                            
                        if f_district.upper() in ('SDE') and f_dev.upper()[4] not in ('E','F'):
                            print "Incorrect N7K device %s for district %s, line %s.  Expecting N7K-[E or F]" % (f_dev,f_district,lines.index(data)+1)
                            errfound = 1
                    
                        try:
                            IP(f_ip)
                        except:
                            print "Invalid IP %s, line %s" % (f_ip,lines.index(data)+1)
                            errfound = 1
                       
                            
                        try:
                            f_un=d[4]
                        except IndexError:
                            print "No n7k Username defined in n7k file.  Expecting username in the 5th column, line %s" % (lines.index(data)+1)
                            errfound = 1
                        
                        try:
                            f_pw=d[5]
                        except IndexError:
                            print "No n7k Password passed in n7k file.  Expecting password in the 6th column, line %s" % (lines.index(data)+1)    
                            errfound = 1
                        
                            
                    if errfound:
                        print "\nPlease correct N7K device file passed to the -e option and try again"    
                        sys.exit(9)
                        
                    # Remove duplicates, if any
                    lines = list(set(lines)) 
                     
            # Get Portmap file for DC1 and DC2.  The characters 'DC1' and 'DC2' must exist in the filename
            # If multiple DC1 or multiple DC2 filenames are passed, only the first one that matches DC1 or DC2 is used.
            # Files must be passed comma seperated, with no spaces before and after the comma
            # Firewall names in the excel sheet must have the word 'firewall' in it and begin with SOE, GIS or SDE
            
            if opt in ("-p", "--portmap"):
                dc1 = False
                dc2 = False
                portmapfile = arg
                xportmapfile = portmapfile.split(",")
                xportmapfile = list(set(xportmapfile))
                
                if len(xportmapfile) != 2:
                    print "The number of portmap files should be only 2. %s Files passed %s" % (len(xportmapfile),xportmapfile)
                    sys.exit(10)
                
                for x in xportmapfile:
                    if not os.path.isfile(x):
                        print sys.argv[0] + " Port map file list %s NOT found" % x
                        sys.exit(1)
                    
                    if dc1 is False:
                        if bool(re.search('dc1',str(x), re.IGNORECASE)):
                            dc1 = True
                            dc1portmap = x
                    
                    if dc2 is False:
                        if bool(re.search('dc2',str(x), re.IGNORECASE)):
                            dc2 = True
                            dc2portmap = x
                            
                if dc1 is False:
                    print "Port Map File for DC1 is missing"
                    sys.exit(10)
                if dc2 is False:
                    print "Port Map File for DC2 is missing"
                    sys.exit(10)
                   
        if debug is True and configure is True:
            print "Cannot use option d and option w together.  Use option w or option d"
            sys.exit(9)
            
        try:
            portmapfile
        except NameError:
            print "Port Map file required with -p|--portmap"
            sys.exit(9)
   
      
         
        for opt,arg in opts:
            if opt == '-h' or opt == '--help':
                usage()
            elif opt in ( "-f", "--file"):
                filename = arg
                if not os.path.isfile(filename):
                    print sys.argv[0] + " excel file %s NOT found" % filename
                    sys.exit(1)
                else:
                    try:
                        open_workbook(filename)
                    except XLRDError:
                        print sys.argv[0] + " file %s is not an Excel file" % filename
                    else:
                        if magic.from_file(filename) == 'Microsoft Excel 2007+':
                            (ws_definition_data,final_all_inner_data,bgp_asn,outer_to_pa_data,n7k_fw_int,loopback_data,outer_jnp_data,out_to_jnp_fw_int) = process_xlsx(filename,dc1portmap,dc2portmap,debug)
                            if debug is False:
                                if configure is True:
                                    confirm = raw_input("\n\nSwitch changes are about to be made.  Type N/n to exit or press any key to continue: \n\n")
                                    if confirm.upper() == 'N':
                                        print "\n\nExiting script.  No changes made\n\n"
                                        sys.exit(9)
                                outer_vdc_config(ws_definition_data,final_all_inner_data,bgp_asn,outer_to_pa_data,n7k_fw_int,loopback_data,outer_jnp_data,configure,lines,detailops,out_to_jnp_fw_int)
                        else:
                            print "File must be in .xlsx format"
                            sys.exit(10)
                            
             
        try:
            filename
        except NameError:
            print "Input file not passed with -f|--file.  Exiting"
            sys.exit(10)
                    

if __name__ == '__main__':
    main(sys.argv[1:])
