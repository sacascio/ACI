#!/usr/bin/env python3
import sys
import json
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def set_col_width(ws):

	dims = {}
	for row in ws.rows:
    		for cell in row:
        		if cell.value:
            			dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
	for col, value in dims.items():
		ws.column_dimensions[get_column_letter(col)].width = value

def get_l2_l3_value():

	l2_l3_vals = {}

	for dafe in ('SG_PBR/DC1_GIS_DAFE.xlsx', 'SG_PBR/DC2_GIS_DAFE.xlsx'):
		dc = dafe[7:10]
		l2_l3_vals[dc] = {}
		worksheets = []
		l3out_list = []
		wb = openpyxl.load_workbook(dafe, data_only=True)

		for sheet in wb:
			worksheets.append(sheet.title)
		wb.close()
	
		wb.active = worksheets.index('bridge_domain')
		ws = wb.active


		row_start = ws.min_row
		row_end   = ws.max_row

		
		for x in range(row_start+1,row_end+1):
			cell = 'A' + str(x)
			bd_name = ws[cell].value
		
			cell = 'C' + str(x)
			tenant_name = ws[cell].value

			cell = 'D' + str(x)
			vrf_name = ws[cell].value

			cell = 'K' + str(x)
			isL3 = ws[cell].value

			if bd_name not in l2_l3_vals[dc]:
				l2_l3_vals[dc][bd_name] = {}
			if tenant_name not in l2_l3_vals[dc][bd_name]:
				l2_l3_vals[dc][bd_name][tenant_name] = {}
			if vrf_name not in l2_l3_vals[dc][bd_name][tenant_name]:
				l2_l3_vals[dc][bd_name][tenant_name][vrf_name] = {}

			l2_l3_vals[dc][bd_name][tenant_name][vrf_name] = isL3

	return l2_l3_vals
			

def reconcileL3out(dc,bd,tenant):
	dafe = 'SG_PBR/' + dc + '_GIS_DAFE.xlsx'
	worksheets = []
	l3out_list = []
	wb = openpyxl.load_workbook(dafe, data_only=True)

	for sheet in wb:
		worksheets.append(sheet.title)
	wb.close()
	wb.active = worksheets.index('bd_l3out')
	ws = wb.active
    
	row_start = ws.min_row
	row_end   = ws.max_row

	for x in range(row_start+1,row_end+1):
		cell = 'A' + str(x)
		bd_name = ws[cell].value

		cell = 'B' + str(x)
		tenant_name = ws[cell].value
		
		if bd_name == bd and tenant_name == tenant:
			cell = 'C' + str(x)
			l3out_name = ws[cell].value
			l3out_list.append(l3out_name)

	if len(l3out_list) > 1:
		print ("WARNING: More than 1 L3out found for Tenant %s, BD %s" % (tenant,bd))
	if len(l3out_list) == 0:
		return 'N/A'
	if len(l3out_list) == 1:
		#print ("INFO: Found L3Out for BD %s, Tenant %s" % (bd,tenant))
		return l3out_list[0]

def loadSites(sitefile) :

	siteData = {}

	with open (sitefile) as site:
		data = json.load(site)

	for i in data:
		siteID =  i['_id']['$oid']
		siteName = i['name']

		siteData[siteID] = siteName

	return siteData

def loadTenants(tenantfile) :

	tenantData = {}

	with open (tenantfile) as tenant:
		data = json.load(tenant)

	for i in data:
		tenantID =  i['_id']['$oid']
		tenantName = i['name']

		tenantData[tenantID] = tenantName

	return tenantData

def loadSchemas(schemafile,tenantData,siteData) :

	schemaData = {}
	found = 0

	with open (schemafile) as schema:
		data = json.load(schema)

	for i in data:
		temp_bds = []
		schemaID =  i['_id']['$oid']
		schemaName = i['displayName']
		#print ("checking schema " + schemaName)
		if bool((re.search('VRF-schema',schemaName,re.IGNORECASE))):
			print ("INFO: SKIPPING " +  schemaName)
			continue

		if schemaID not in schemaData:
			schemaData[schemaID] = {}
		else:
			print ("WARNING: Schema ID %s already found for schema Name %s" % (schemaID,schemaName))

		if schemaName not in schemaData[schemaID]:
			schemaData[schemaID][schemaName] = {}
		else:
			print ("WARNING: Schema name %s already found in schema ID %s" % (schemaName,schemaID))

		# Get the BD names where more than 1 BD exists at the site.  This is a workaround because other templates have 0 BD at the site, which seems like a bug
		for zz in i['sites']:
			if len(zz['bds']) > 1:
				siteID = zz['siteId']
				for z in zz['bds']:
					temp_bds.append({'siteID' : siteID, 'bdName' : z['bdRef']['bdName'], 'schemaId' : z['bdRef']['schemaId'], 'templateName' : z['bdRef']['templateName'], 'l3Out' : z['l3Outs']})
				#print (temp_bds)

		for j in i['templates']:
			templateName = j['name']
			tenantName = tenantData[j['tenantId']]						
			schemaData[schemaID][schemaName][templateName] = {}

			schemaData[schemaID][schemaName][templateName]['bd'] = []
			schemaData[schemaID][schemaName][templateName]['vrf'] = []
			schemaData[schemaID][schemaName][templateName]['tenant'] = []
			schemaData[schemaID][schemaName][templateName]['anp'] = []
			schemaData[schemaID][schemaName][templateName]['sites'] = []
			schemaData[schemaID][schemaName][templateName]['epgs'] = []
			schemaData[schemaID][schemaName][templateName]['l3out'] = []
			schemaData[schemaID][schemaName][templateName]['tenant'].append(tenantName)

			if len(j['bds']) != 1:
				print ("WARNING: %s BD's found in schema %s, template %s" % (len(j['bds']), schemaName, templateName))
			
			if len(j['anps']) != 1:
				print ("WARNING: %s ANP's found in schema %s, template %s" % (len(j['anps']), schemaName, templateName))
			
			if len(j['vrfs']) != 0:
				print ("WARNING: %s VRFs found in schema %s, template %s" % (len(j['vrfs']), schemaName, templateName))

			for k in j['anps']:
				anpName = k['name']
			
				schemaData[schemaID][schemaName][templateName]['anp'].append(anpName)	
				
				if k['anpRef']['schemaId'] != schemaID:
					print ("WARNING: Schema ID for schema %s in ANPREF of template %s not the same as schemaID found previously" % (schemaName,templateName))

				if k['anpRef']['templateName'] != templateName:
					print ("WARNING: Template name %s in ANPREF is not the same as template Name found previously" % (templateName))
				
				if k['anpRef']['anpName'] != anpName:
					print ("WARNING: ANP name %s in ANPREF in template %s is not the same as ANP Name found previously" % (anpName,templateName))

				for l in k['epgs']:
					epgName = l['name']
				
					schemaData[schemaID][schemaName][templateName]['epgs'].append(epgName)	

					if l['epgRef']['schemaId'] != schemaID:
						print ("WARNING: Schema ID for schema %s in EPGREF of template %s not the same as schemaID found previously" % (schemaName,templateName))
				
					if l['epgRef']['templateName'] != templateName:
						print ("WARNING: Template name %s in EPGREF is not the same as template Name found previously" % (templateName))
					
					if l['epgRef']['anpName'] != anpName:
						print ("WARNING: ANP name %s in EPGREF is not the same as ANP Name found previously" % (anpName))
					
					if l['epgRef']['epgName'] != epgName:
						print ("WARNING: EPG name %s in EPGREF is not the same as EPG Name found previously" % (epgName))
				
			for m in j['bds']:
				bdName = m['name']
				schemaData[schemaID][schemaName][templateName]['bd'].append(bdName)

				if m['bdRef']['schemaId'] != schemaID:
					print ("WARNING: Schema ID for schema %s in BDREF of template %s not the same as schemaID found previously" % (schemaName,templateName))

				if m['bdRef']['templateName'] != templateName:
					print ("WARNING: Template name %s in BDREF is not the same as template Name found previously" % (templateName))
				
				if m['bdRef']['bdName'] != bdName:
					print ("WARNING: BD name %s in BDREF in template %s is not the same as BD Name found previously" % (bdName,templateName))

				vrf = m['vrfRef']['vrfName']
				schemaData[schemaID][schemaName][templateName]['vrf'].append(vrf)


			for n in i['sites']:
				siteTemplateName = n['templateName'] 
				
				siteID = n['siteId']
				siteName = siteData[siteID]

				if siteTemplateName == templateName:
					schemaData[schemaID][schemaName][siteTemplateName]['sites'].append(siteName)
				else:
					continue

				if len(n['anps']) != 1:
					print ("WARNING: %s ANP's found in schema %s, template %s IN SITE section" % (len(n['anps']), schemaName, siteTemplateName))
			
				if len(n['vrfs']) != 0:
					print ("WARNING: %s VRFs found in schema %s, template %s IN SITE section" % (len(n['vrfs']), schemaName, siteTemplateName))
		

				for o in n['anps']:

					if o['anpRef']['schemaId'] != schemaID:
						print ("WARNING: Schema ID for schema %s in SITE ANPREF of template %s not the same as schemaID found previously" % (schemaName,siteTemplateName))
				
					if o['anpRef']['templateName'] != siteTemplateName:
						print ("WARNING: Template name %s in SITE ANPREF is not the same as template Name found previously" % (siteTemplateName))
				
					if o['anpRef']['anpName'] != anpName:
						print ("WARNING: ANP name %s in SITE ANPREF in template %s is not the same as ANP Name found previously" % (anpName,siteTemplateName))
				
					for p in k['epgs']:

						if p['epgRef']['schemaId'] != schemaID:
							print ("WARNING: Schema ID for schema %s in SITE EPGREF of template %s not the same as schemaID found previously" % (schemaName,siteTemplateName))
				
						if p['epgRef']['templateName'] != siteTemplateName:
							print ("WARNING: Template name %s in SITE EPGREF is not the same as template Name found previously" % (siteTemplateName))
					
						if p['epgRef']['anpName'] != anpName:
							print ("WARNING: ANP name %s in SITE EPGREF is not the same as ANP Name found previously" % (anpName))
					
						if p['epgRef']['epgName'] != epgName:
							print ("WARNING: EPG name %s in SITE EPGREF is not the same as EPG Name found previously" % (epgName))
		
				if len(n['bds']) == 1:
	
					for q in n['bds']:
						if q['bdRef']['schemaId'] != schemaID:
							print ("WARNING: Schema ID for schema %s in SITE BDREF of template %s not the same as schemaID found previously" % (schemaName,siteTemplateName))

						if q['bdRef']['templateName'] != siteTemplateName:
							# Looks like at the site level, BDs for other templates are showing up in the same template.  ex. schema SVC-COM-DCX-DC2-01 
							print ("WARNING: Template name %s in SITE BDREF is not the same as template Name found previously" % (siteTemplateName))
				
						if q['bdRef']['bdName'] != bdName:
							print ("WARNING: BD name %s in SITE BDREF in template %s is not the same as BD Name found previously" % (bdName,siteTemplateName))

						for r in q['l3Outs']:
							schemaData[schemaID][schemaName][siteTemplateName]['l3out'].append(r)

				if len(n['bds']) != 1:
					#print ("checking site template " + siteTemplateName)
					#print (temp_bds)
					for y in temp_bds:
						if y['templateName'] == siteTemplateName and y['siteID'] == siteID:
							found = 1
							
							if y['schemaId'] not in schemaData:
								schemaData[y['schemaId']] = {}
							if schemaName not in schemaData[y['schemaId']]:
								schemaData[y['schemaId']][schemaName] = {}
							if siteTemplateName not in schemaData[y['schemaId']][schemaName]:
								schemaData[y['schemaId']][schemaName][siteTemplateName] = {}
								schemaData[y['schemaId']][schemaName][siteTemplateName]['l3out'] = []
							
							if len(y['l3Out']) > 1:
								print ("WARNING: Multiple L3outs at SITE found for Template %s, Schema %s" % (siteTemplateName,schemaName))
								#print (y['l3Out'])
 
							if len(y['l3Out']) == 1:
								schemaData[y['schemaId']][schemaName][siteTemplateName]['l3out'].append(y['l3Out'][0])
							
				if found == 0 and len(n['bds']) == 0:
					#print (temp_bds)
					if y['schemaId'] not in schemaData:
						schemaData[y['schemaId']] = {}
					if schemaName not in schemaData[y['schemaId']]:
						schemaData[y['schemaId']][schemaName] = {}
					if siteTemplateName not in schemaData[y['schemaId']][schemaName]:
						schemaData[y['schemaId']][schemaName][siteTemplateName] = {}
						schemaData[y['schemaId']][schemaName][siteTemplateName]['l3out'] = []
							
						schemaData[y['schemaId']][schemaName][siteTemplateName]['l3out'].append("N/A")

					#print ("WARNING: Could not find l3Out for template %s, schema %s" % (siteTemplateName, schemaName))	
				else:
					found = 0
					
			
	return schemaData


def main(argv):
	l2_l3_vals = get_l2_l3_value()
	siteData = loadSites('site.json')
	tenantData = loadTenants('tenant.json')
	schemaData = loadSchemas('schema.json',tenantData, siteData)
	#print (json.dumps(schemaData))	

	wb = Workbook()
	file = 'GIS_MSO_DATA.xlsx'
	dc1_ws = wb.active
	dc1_ws.title = 'DC1'

	dc2_ws = wb.create_sheet(title='DC2')

	fields = ['Schema_ID','Schema_Name','Template_Name','BD_NAME','VRF_Name','Tenant_Name','ANP_Name','site_name','EPG_Name','L3Out_Name','isL3Network']

	row = 1
	for col in range (1,12):	
		cell = dc1_ws.cell(column=col, row=row)
		cell.value = fields[col-1]
	
		cell = dc2_ws.cell(column=col, row=row)
		cell.value = fields[col-1]

	dc1_row = 2	
	dc2_row = 2	
	col = 1

	for schemaID in schemaData:
		for schemaName in schemaData[schemaID]:
			for templateName in schemaData[schemaID][schemaName]:
				if 'bd' in schemaData[schemaID][schemaName][templateName]:
					if len(schemaData[schemaID][schemaName][templateName]['bd']) > 1:
						print ("WARNING: More than 1 BD found in Schema %s, template %s" % (schemaName,templateName)) 
					bdName = schemaData[schemaID][schemaName][templateName]['bd'][0]
				else:
					bdName = 'N/A'
				
				if 'vrf' in schemaData[schemaID][schemaName][templateName]:
					if len(schemaData[schemaID][schemaName][templateName]['vrf']) > 1:
						print ("WARNING: More than 1 VRF found in Schema %s, template %s" % (schemaName,templateName)) 
					vrfName = schemaData[schemaID][schemaName][templateName]['vrf'][0]
				else:
					vrfName = 'N/A'
				
				if 'tenant' in schemaData[schemaID][schemaName][templateName]:
					if len(schemaData[schemaID][schemaName][templateName]['tenant']) > 1:
						print ("WARNING: More than 1 Tenant found in Schema %s, template %s" % (schemaName,templateName)) 
					tenantName = schemaData[schemaID][schemaName][templateName]['tenant'][0]
				else:
					tenantName = 'N/A'
				
				if 'anp' in schemaData[schemaID][schemaName][templateName]:
					if len(schemaData[schemaID][schemaName][templateName]['anp']) > 1:
						print ("WARNING: More than 1 ANP found in Schema %s, template %s" % (schemaName,templateName)) 
					anpName = schemaData[schemaID][schemaName][templateName]['anp'][0]
				else:
					anpName = 'N/A'
				
				if 'sites' in schemaData[schemaID][schemaName][templateName]:
					if len(schemaData[schemaID][schemaName][templateName]['sites']) > 1:
						print ("WARNING: More than 1 Site found in Schema %s, template %s" % (schemaName,templateName)) 
					siteName = schemaData[schemaID][schemaName][templateName]['sites'][0]
				else:
					siteName = 'N/A'
				
				if 'epgs' in schemaData[schemaID][schemaName][templateName]:
					if len(schemaData[schemaID][schemaName][templateName]['epgs']) > 1:
						print ("WARNING: More than 1 EPG found in Schema %s, template %s" % (schemaName,templateName)) 
					epgName = schemaData[schemaID][schemaName][templateName]['epgs'][0]
				else:
					epgName = 'N/A'
				
				if 'l3out' in schemaData[schemaID][schemaName][templateName]:
					if len(schemaData[schemaID][schemaName][templateName]['l3out']) > 1:
						print ("WARNING: More than 1 L3Out found in Schema %s, template %s" % (schemaName,templateName)) 
						print (schemaData[schemaID][schemaName][templateName]['l3out'])
					if len(schemaData[schemaID][schemaName][templateName]['l3out']) == 0:
						l3outName = 'N/A'
					else:
						l3outName = schemaData[schemaID][schemaName][templateName]['l3out'][0]
				else:
					l3outName = 'N/A'

				if bool(re.search('DC1',siteName,re.IGNORECASE)):
					cell = dc1_ws.cell(column=1, row=dc1_row)
					cell.value = schemaID

					cell = dc1_ws.cell(column=2, row=dc1_row)
					cell.value = schemaName
					
					cell = dc1_ws.cell(column=3, row=dc1_row)
					cell.value = templateName
					
					cell = dc1_ws.cell(column=4, row=dc1_row)
					cell.value = bdName
					
					cell = dc1_ws.cell(column=5, row=dc1_row)
					cell.value = vrfName
					
					cell = dc1_ws.cell(column=6, row=dc1_row)
					cell.value = tenantName
					
					cell = dc1_ws.cell(column=7, row=dc1_row)
					cell.value = anpName
					
					cell = dc1_ws.cell(column=8, row=dc1_row)
					cell.value = siteName
					
					cell = dc1_ws.cell(column=9, row=dc1_row)
					cell.value = epgName
					
					if l3outName == 'N/A':
						l3outName = reconcileL3out('DC1',bdName,tenantName)
					
					cell = dc1_ws.cell(column=10, row=dc1_row)
					cell.value = l3outName

					cell = dc1_ws.cell(column=11, row=dc1_row)

					if bdName != 'N/A':
						cell.value = l2_l3_vals['DC1'][bdName][tenantName][vrfName]
					else:
						cell.value = 'N/A'
							

					dc1_row = dc1_row + 1

				else:
					cell = dc2_ws.cell(column=1, row=dc2_row)
					cell.value = schemaID

					cell = dc2_ws.cell(column=2, row=dc2_row)
					cell.value = schemaName
					
					cell = dc2_ws.cell(column=3, row=dc2_row)
					cell.value = templateName
					
					cell = dc2_ws.cell(column=4, row=dc2_row)
					cell.value = bdName
					
					cell = dc2_ws.cell(column=5, row=dc2_row)
					cell.value = vrfName
					
					cell = dc2_ws.cell(column=6, row=dc2_row)
					cell.value = tenantName
					
					cell = dc2_ws.cell(column=7, row=dc2_row)
					cell.value = anpName
					
					cell = dc2_ws.cell(column=8, row=dc2_row)
					cell.value = siteName
					
					cell = dc2_ws.cell(column=9, row=dc2_row)
					cell.value = epgName
					
					if l3outName == 'N/A':
						l3outName = reconcileL3out('DC2',bdName,tenantName)

					cell = dc2_ws.cell(column=10, row=dc2_row)
					cell.value = l3outName
					
					cell = dc2_ws.cell(column=11, row=dc2_row)
	
					if bdName != 'N/A':
						cell.value = l2_l3_vals['DC2'][bdName][tenantName][vrfName]
					else:
						cell.value = 'N/A'
					
					dc2_row = dc2_row + 1
							
	print ("Excel file saved to name " + file)
	#print ("Load ACI config and check L3Out")

	set_col_width(dc1_ws)
	set_col_width(dc2_ws)

	dc1_ws.sheet_view.zoomScale = 125
	dc2_ws.sheet_view.zoomScale = 125

	# add filters
	dc1_ws.auto_filter.ref = 'A1:K900'
	dc2_ws.auto_filter.ref = 'A1:K900'

	wb.save(filename = file)
    		

if __name__ == '__main__':
	main(sys.argv[1:])
