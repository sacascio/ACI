#!/usr/bin/env python

import openpyxl
import sys
import getopt
import os.path
import magic
import re
#import numbers
from xlrd import open_workbook, XLRDError
import json
from IPy import IP

def inner_vdc_config(ws_definition_data,final_all_inner_data,bgp_asn,outer_to_pa_data,n7k_fw_int,loopback_data,configure):
   
    vlans = []
    commands = []
    dcs = ['dc1', 'dc2']
    districts = ['SOE','GIS','SDE']
    n7k_prod  = ['N7K-A','N7K-B','N7K-C','N7K-D']
    n7k_dev   = ['N7K-E','N7K-F']
    loopback_position  = {'N7K-A' : 1, 'N7K-B' : 2, 'N7K-C' : 3, 'N7K-D' : 4, 'N7K-E' : 1, 'N7K-F' : 2}
    
    for dc in dcs:
        for district in districts:
            # DC1 config - select N7K naming convention based on prod (GIS/SOE) or dev (SDE)
            if district in ('GIS','SOE'):
                n7k = n7k_prod
            else:
                n7k = n7k_dev

            for nexusvdc in n7k:
                print "!!! District %s, DC %s, nexusVDC %s" % (district,dc,nexusvdc)
                print "!"
                # Get the firewall interfaces
                fwint1 =  n7k_fw_int[district]['Inner'][nexusvdc][dc]['int1']
                fwint2 =  n7k_fw_int[district]['Inner'][nexusvdc][dc]['int2']
                
                for vsys in ws_definition_data[district]:
                    for attribs in ws_definition_data[district][vsys]:
                        innervdcvlan =  attribs['innervdcencap']
                        subzone      =  attribs['subzone']
                        n7kip        =  final_all_inner_data[district][subzone][nexusvdc][0][dc + 'n7kip']
                        commands.append("interface vlan " + str(innervdcvlan))
                        commands.append("  description Layer3_%s_%s" % (vsys,attribs[dc+'vrf']))
                        commands.append("  vrf member %s " % (attribs[dc+'vrf']))
                        commands.append("  ip address %s/30" % (n7kip))
                        commands.append("  ip ospf network point-to-point")
                        commands.append("  ip router ospf %s area 0.0.0.%s" % (vsys.upper(),attribs['ospf' + dc]))
                        commands.append("  no shutdown")
                        
                        print '\n'.join(map(str,commands))
                        print "!"
                        
                        if configure is True:
                            
                            commands = ";".join(map(str,commands))
                            print "*** sending above config to %s,%s,%s ***"  %(dc,district,nexusvdc)
                        
                        commands = []
                        
                    
                        vlans.append(str(innervdcvlan))
                    
                    print "!"
                    # Begin OSPF config
                    print "router ospf %s" % (vsys)  
                          
                    for attribs in ws_definition_data[district][vsys]:    
                            n7k_num = loopback_position[nexusvdc]
                            for vals in loopback_data[district]:
                                if vals[dc + 'hn'] == dc + 'dcinxc' + str(n7k_num) + district.lower() + 'inner':
                                    loopback_address = vals[dc + 'ip']  
                            print " vrf %s" % (attribs[dc+'vrf'])
                            print "   router-id %s" % (loopback_address)
                            print "   log-adjacency-changes"
                            print "!"
                            
                print "!"
                print "!"
                print "!"
                print "!"            
                # BGP Configuration
                inner_as   = bgp_asn[district]['Inner'][dc]
                outer_as   = bgp_asn[district]['Outer'][dc]
                            
                print "router bgp %s" % (inner_as)
                print " router-id %s" % (loopback_address)
                print " address-family l2vpn evpn"
                print "  maximum-paths 8"
                
                for vsys in ws_definition_data[district]:
                    for attribs in ws_definition_data[district][vsys]:
                        vrf = attribs[dc+'vrf']
                        print "vrf %s" % (vrf)
                        for vdc in n7k:
                            n7k_num = loopback_position[vdc]
                            if district == 'SDE':
                                outervdc = dc + district.lower() + 'nxc' + str(n7k_num) + district.lower() + 'outer'
                            else:
                                outervdc = dc + 'dcinxc' + str(n7k_num) + 'dciouter'
                                
                            tname = vsys + "-" + dc.upper() + "-" + district
                            
                            neighbor_ip = outer_to_pa_data[district][vsys][vdc][0][dc + 'n7kip']
                            print " address-family ipv4 unicast"
                            print "  maximum-paths 8"
                            print " neighbor %s remote-as %s" % (neighbor_ip,outer_as) 
                            print " description TO_%s_%s" % (outervdc,tname)
                            print "    ebgp-multihop 4"
                            print "    address-family ipv4 unicast"
                            print "      send-community both"
                           
                # got all vlans for district/subzone per N7K - now add the vlans to the FW Int config
                vlans.sort()
                print "!"
                print "!"
                print "! Allow VLANs on the firewall"
                # Add vlans to allowed list on firewall interfaces
                print "interface %s" % (fwint1)
                print " switchport trunk allow vlan add " + ','.join(map(str,vlans))
                print "!"
                print "!"
                print "interface %s" % (fwint2)
                print " switchport trunk allow vlan add " + ','.join(map(str,vlans))
                print "!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"
                print ""
                vlans = []
                
def get_outer_to_pa(wb):
        
        ws = wb.active 
        row_start = ws.min_row
        row_end   = ws.max_row
        data = {}

        # Process outer VDC to PA

        for x in range(row_start,row_end+1):
            # Get district and tenant
            cell = 'A' + str(x)
            value = ws[cell].value 
            
            if value is not None and  \
            not bool((re.search('Addressing',value,re.IGNORECASE))) and \
            not bool((re.search('Mask',value,re.IGNORECASE))) and \
            value != "vSYS":
                value = value.strip()

                if value == 'SDE':
                    district = value
                elif value == 'SOE':
                    district = value
                elif value == 'GIS':
                    district = value
                else:
                    tenant = value
                    tenant = tenant.upper()
                    tenant = tenant.replace(" ","_")

            # Get PA IP DC1
            cell = 'B' + str(x)
            value = ws[cell].value 
            
            if value is None or value == 'DC1':
                continue
           
            if not bool((re.search('FW',value,re.IGNORECASE))):
                
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1paip = value
                    
            # Get N7K IP DC1
            cell = 'C' + str(x)
            value = ws[cell].value 
            
            if value is None or value == 'DC1':
                continue
           
            if not bool((re.search('VDC',value,re.IGNORECASE))) and not bool((re.search('Address',value,re.IGNORECASE))):
                
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1n7kip = value
            
            # Get N7K VDC
            cell = 'D' + str(x)
            value = ws[cell].value 
            
            if value is None:
                continue

            value = value.strip()
            n7k = value
            
            # Get PA IP DC2
            cell = 'E' + str(x)
            value = ws[cell].value 
            
            if value is None or value == 'DC2':
                continue
           
            if not bool((re.search('FW',value,re.IGNORECASE))):
                
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2paip = value
            
            # Get N7K IP DC2
            cell = 'F' + str(x)
            value = ws[cell].value 
            
            if value is None or value == 'DC2':
                continue
           
            if not bool((re.search('Address',value,re.IGNORECASE))):
                
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2n7kip = value
            
            # If tenant and district exist, append attributes as a list
            
            if district in data:
                if tenant in data[district]:
                    if n7k in data[district][tenant]:
                            data[district][tenant][n7k].append(
                                    {
                                     'dc1paip'     : dc1paip,
                                     'dc1n7kip'    : dc1n7kip,
                                     'dc2paip'     : dc2paip,
                                     'dc2n7kip'    : dc2n7kip
                                   })
                            # If district exists, but not tenant, add new key (tenant) and initial attributes
                    else:
                            data[district][tenant][n7k] = [ {  
                                     
                                     'dc1paip'     : dc1paip,
                                     'dc1n7kip'    : dc1n7kip,
                                     'dc2paip'     : dc2paip,
                                     'dc2n7kip'    : dc2n7kip
                                     
                                     } ]
                            
                else:
                    data[district][tenant] = {}        
                    data[district][tenant][n7k] = [ {  
                                  'dc1paip'     : dc1paip,
                                  'dc1n7kip'    : dc1n7kip,
                                  'dc2paip'     : dc2paip,
                                   'dc2n7kip'    : dc2n7kip
                                } ]

            # Initial key/value assignment
            else:
                    data.update({  
                              district :  
                              { 
                                tenant :  
                                    { 
                                      n7k : [ 
                                     {
                                     'dc1paip'     : dc1paip,
                                     'dc1n7kip'    : dc1n7kip,
                                     'dc2paip'     : dc2paip,
                                     'dc2n7kip'    : dc2n7kip
                                    }
                                    ] 
                                }
                              } 
                         })
        

        return data

def get_loopback(wb):
        
        ws = wb.active 
        row_start = ws.min_row
        row_end   = ws.max_row
        data = {}
        
        # Process Loopback assignments

        for x in range(row_start,row_end):

            # Get IP DC1
            cell = 'B' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('IP',value,re.IGNORECASE))) and value != 'DC1':
            
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1ip = value

            # Get hostname DC1
            cell = 'C' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('Hostname',value,re.IGNORECASE))) and value != 'DC1':
                dc1hn = value
                dc1hn = dc1hn.strip()
                
            # Get Description
            cell = 'D' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('Description',value,re.IGNORECASE))) and value != 'DC1':
                dc1desc = value
                dc1desc = dc1desc.strip()
                
            # Get IP DC2
            cell = 'E' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('IP',value,re.IGNORECASE))) and value != 'DC2':
            
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2ip = value

            # Get hostname DC2
            cell = 'F' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('Hostname',value,re.IGNORECASE))) and value != 'DC2':
                dc2hn = value
                dc2hn = dc2hn.strip()
            
            # IF row is blank, skip
            else:
                continue
                
            # Get Description
            cell = 'G' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool((re.search('Description',value,re.IGNORECASE))) and value != 'DC2':
                dc2desc = value
                dc2desc = dc2desc.strip()

            # Push all data to dictionary
            # Use Debug option to print data

            if bool((re.search('soe',value,re.IGNORECASE))):
                district = ['SOE']

            if bool((re.search('sde',value,re.IGNORECASE))):
                district = ['SDE']
            
            if bool((re.search('gis',value,re.IGNORECASE))):
                district = ['GIS']
                
            if bool((re.search('soe/gis',value,re.IGNORECASE))):
                district = ['GIS','SOE']
            
            for d in district: 
                # If district exist, append attributes as a list
                if d in data:
                         data[d].append(
                                     {
                                      'dc1ip'         : dc1ip,
                                      'dc1desc'       : dc1desc,
                                      'dc1hn'         : dc1hn,
                                      'dc2ip'         : dc2ip,
                                      'dc2desc'       : dc2desc,
                                      'dc2hn'         : dc2hn
                                     })
            
                # Initial key/value assignment
                else:
                 data.update({  
                      d :  
                                 [ 
                                  {
                                      'dc1ip'         : dc1ip,
                                      'dc1desc'       : dc1desc,
                                      'dc1hn'         : dc1hn,
                                      'dc2ip'         : dc2ip,
                                      'dc2desc'       : dc2desc,
                                      'dc2hn'         : dc2hn
                                 }
                            ] 
                           } )
        return data
            
def get_inner_to_pa(wb,district):
        
        ws = wb.active 
        row_start = ws.min_row
        row_end   = ws.max_row
        data = {}
        
        if district in ("SOE","GIS"):
            max_n7k = 4
        else:
            max_n7k = 2

        # Process ws_soe_inner_to_pa
        for x in range(row_start,row_end):
            # Get VRF#
            cell = 'A' + str(x)
            value = ws[cell].value 
            
            if value is not None and bool((re.search('VRF-',value,re.IGNORECASE))):
                vrfnumber = value
            
            # Get description
            cell = 'B'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('Connection',value,re.IGNORECASE))):
                connectdesc = value
                temp = connectdesc[3]
                connectdesc = connectdesc[:3] + "-" + temp

            
            # get N7K IP DC1
            cell = 'C'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('IP Add',value,re.IGNORECASE))) and not bool((re.search('DC',value,re.IGNORECASE))):
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1n7kip = value
            
            # get PA FW IP DC1
            cell = 'D'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('IP Add',value,re.IGNORECASE))) and not bool((re.search('DC',value,re.IGNORECASE))):
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc1pafwip = value
            
            # get sub zone name
            cell = 'E'  + str(x)
            value = ws[cell].value

            if value is not None and not bool((re.search('Zone',value,re.IGNORECASE))):
                subzone = value
                subzone = subzone.strip()
                subzone = subzone.replace(' ',"_")
                subzone = subzone.upper()
                
                
            # get N7K VRF Name DC1
            cell = 'F'  + str(x)
            value = ws[cell].value
            
            if value != 0 and value is not None and value != 'DC1 ' and value != 'DC2' and not bool((re.search('N7K',value,re.IGNORECASE))) :
                n7kvrfdc1 = value
                n7kvrfdc1 = n7kvrfdc1.strip()
            
            # get N7K VRF Name DC2
            cell = 'G'  + str(x)
            value = ws[cell].value
            
            if value != 0 and value is not None and value != 'DC1 ' and value != 'DC2' and not bool((re.search('N7K',value,re.IGNORECASE))) :
                n7kvrfdc2 = value
                n7kvrfdc2 = n7kvrfdc2.strip()

            # get N7K IP DC2
            cell = 'H'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('IP Add',value,re.IGNORECASE))) and not bool((re.search('DC',value,re.IGNORECASE))):
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2n7kip = value
            
            # get PA FW IP DC2
            cell = 'I'  + str(x)
            value = ws[cell].value
            
            if value is not None and not bool((re.search('IP Add',value,re.IGNORECASE))) and not bool((re.search('DC',value,re.IGNORECASE))):
                try:
                    IP(value)
                except:
                    print "Worksheet %s, Cell %s, value is not an IP address" % ws,cell
                    sys.exit(1)
                else:
                    dc2pafwip = value
                   
            # Push all data to dictionary
            # Use Debug option to print data
                
            # If subzone and connection exist, append attributes as a list
                if district in data:
                    if subzone in data[district]:
                        if connectdesc in data[district][subzone]:
                            if len(data[district][subzone]) < max_n7k:
                                data[district][subzone][connectdesc].append(
                                        {
                                         'vrfnumber'     : vrfnumber,
                                         'dc1n7kip'      : dc1n7kip,
                                         'dc1pafwip'     : dc1pafwip,
                                         'n7kvrfdc1'     : n7kvrfdc1,
                                         'n7kvrfdc2'     : n7kvrfdc2,
                                         'dc2n7kip'      : dc2n7kip,
                                         'dc2pafwip'     : dc2pafwip
                                        })
                # If district exists, but not tenant, add new key (tenant) and initial attributes
                        else:
                                data[district][subzone][connectdesc] = [ {  
                                  'vrfnumber'     : vrfnumber,
                                  'dc1n7kip'      : dc1n7kip,
                                  'dc1pafwip'     : dc1pafwip,
                                  'n7kvrfdc1'     : n7kvrfdc1,
                                  'n7kvrfdc2'     : n7kvrfdc2,
                                  'dc2n7kip'      : dc2n7kip,
                                  'dc2pafwip'     : dc2pafwip
                                } ]
                    else:
                        data[district][subzone] = {}        
                        data[district][subzone][connectdesc] = [ {  
                                  'vrfnumber'     : vrfnumber,
                                  'dc1n7kip'      : dc1n7kip,
                                  'dc1pafwip'     : dc1pafwip,
                                  'n7kvrfdc1'     : n7kvrfdc1,
                                  'n7kvrfdc2'     : n7kvrfdc2,
                                  'dc2n7kip'      : dc2n7kip,
                                  'dc2pafwip'     : dc2pafwip
                                } ]
                        

                # Initial key/value assignment
                else:
                    data.update({  
                         district :  
                                 { 
                                      subzone : 
                                             { 
                                         
                                                connectdesc :  [ 
                                                         {
                                                          'vrfnumber'     : vrfnumber,
                                                          'dc1n7kip'      : dc1n7kip,
                                                          'dc1pafwip'     : dc1pafwip,
                                                          'n7kvrfdc1'     : n7kvrfdc1,
                                                          'n7kvrfdc2'     : n7kvrfdc2,
                                                          'dc2n7kip'      : dc2n7kip,
                                                          'dc2pafwip'     : dc2pafwip
                                                          }
                                                         ] 
                                              }
                                 } 
                    })
   

        return data

def process_xlsx(filename,debug):
    worksheets = []
    ws_definition_data = {}

    # Important worksheets
    ws_definition = "SOE_SDE_GIS_VRF_RT_Definition"

    wb = openpyxl.load_workbook(filename, data_only=True)

    # Get all worksheets
    for sheet in wb:
        worksheets.append(sheet.title) 
    wb.close()

    # Set ws_definition as the active worksheet to read from
    # Process SOE_SDE_GIS_VRF_RT_Definition
    
   
    try:
        wb.active = worksheets.index(ws_definition)
    except:
        print "Worksheet %s not found" % ws_definition
    else:
        ws = wb.active
        row_start = ws.min_row
        row_end   = ws.max_row

        # Process ws_definition tab 
        for x in range(row_start,row_end):
            # Get District
            cell = 'A' + str(x)
            value = ws[cell].value 
            
            if value is not None and value != 'District':
                district = value

            # Go to next row if heading
            if value == 'District':
                continue

            # Get Zone
            cell = 'B' + str(x)
            value = ws[cell].value 
           
            # Go to next row if heading
            if value is not None and bool((re.search('Zone',value,re.IGNORECASE))):
                continue

            if value is not None and value != 'Sub Zone' and value != 'Network Segmentation Zone' and not bool(re.search('Zone',value, re.IGNORECASE)):
                tenant = value
                tenant = tenant.strip()
                tenant = tenant.replace(" ","_")
                tenant = tenant.upper()
         
            # Get Sub Zone
            cell = 'C' + str(x)
            value = ws[cell].value 
           
            # Skip row if there is no value (empty row)
            if value is None:
               continue
            if value is not None and not bool(re.search('Sub Zone',value, re.IGNORECASE)):
                szone = value
                szone = szone.strip()
                szone = szone.replace(" ","_")
                szone = szone.upper()

            # Get VRF #
            cell = 'D' + str(x)
            value = ws[cell].value 
            
            if value is not None and value != 'VRF' and value != '#':
                    vrf = value
            
            # Get VRF Name for DC1 - skip entire row if VRF name for DC1 is blank 
            cell = 'E' + str(x)
            value = ws[cell].value 
            
            if value is None:
                continue
            
            else:
                    if  not bool(re.search('N7K',value, re.IGNORECASE)) and value != 'DC1' and value != 'DC2':
                        vrfnamedc1 = value
                        vrfnamedc1 = vrfnamedc1.strip()

            # Get VRF Name for DC2
            cell = 'F' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('N7K',value, re.IGNORECASE)) and value != 'DC1' and value != 'DC2':
                    vrfnamedc2 = value
                    vrfnamedc2 = vrfnamedc2.strip()
            
            # RT DC1
            cell = 'G' + str(x)
            value = ws[cell].value 
            
            if value is not None and value != 'RT' and value != 'DC1' and value != 'DC2':
                    rtdc1 = value
                    rtdc1 = rtdc1.strip()

            # RT DC2
            cell = 'H' + str(x)
            value = ws[cell].value 
            
            if value is not None and value != 'RT' and value != 'DC1' and value != 'DC2':
                    rtdc2 = value
                    rtdc2 = rtdc2.strip()
            
            # inner VDC to FW encap

            cell = 'I' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('encapsulation',str(value))) and not bool(re.search('Inner',str(value))):
                    invdcencap = value
            
            # OSPF DC1 
            cell = 'J' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('OSPF',str(value), re.IGNORECASE)) and not bool(re.search('DC',str(value), re.IGNORECASE)):
                    ospfdc1 = value
            
            # OSPF DC2 
            cell = 'K' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('OSPF',str(value), re.IGNORECASE)) and not bool(re.search('DC',str(value), re.IGNORECASE)):
                    ospfdc2 = value


            # encap - outer VDC to FW
            cell = 'L' + str(x)
            value = ws[cell].value 
            
            if value is not None and not bool(re.search('encapsulation',str(value), re.IGNORECASE)) and not bool(re.search('Inside',str(value), re.IGNORECASE)):
                    outvdcencap = value
                     
           # Push data into dictionary
           # Use Debug option to print data

            # If tenant and district exist, append attributes as a list
            if district in ws_definition_data:
                if tenant in ws_definition_data[district]:
                  ws_definition_data[district][tenant].append(
                                    {
                                     'vrfnumber'     : vrf,
                                     'subzone'       : szone,
                                     'dc1vrf'        : vrfnamedc1,
                                     'dc2vrf'        : vrfnamedc2,
                                     'rtdc1'         : rtdc1,
                                     'rtdc2'         : rtdc2,
                                     'innervdcencap' : invdcencap,
                                     'ospfdc1'       : ospfdc1,
                                     'ospfdc2'       : ospfdc2,
                                     'outervdcencap' : outvdcencap
                                   })
                # If district exists, but not tenant, add new key (tenant) and initial attributes
                else:
                    ws_definition_data[district][tenant] = [ {  
                                     
                                      'vrfnumber'     : vrf,
                                      'subzone'       : szone,
                                      'dc1vrf'        : vrfnamedc1,
                                      'dc2vrf'        : vrfnamedc2,
                                      'rtdc1'         : rtdc1,
                                      'rtdc2'         : rtdc2,
                                      'innervdcencap' : invdcencap,
                                      'ospfdc1'       : ospfdc1,
                                      'ospfdc2'       : ospfdc2,
                                      'outervdcencap' : outvdcencap
                                     } ]

            # Initial key/value assignment
            else:
                    ws_definition_data.update({  
                              district :  
                              { 
                                tenant :  [ 
                                     {
                                      'vrfnumber'     : vrf,
                                      'subzone'       : szone,
                                      'dc1vrf'        : vrfnamedc1,
                                      'dc2vrf'        : vrfnamedc2,
                                      'rtdc1'         : rtdc1,
                                      'rtdc2'         : rtdc2,
                                      'innervdcencap' : invdcencap,
                                      'ospfdc1'       : ospfdc1,
                                      'ospfdc2'       : ospfdc2,
                                      'outervdcencap' : outvdcencap
                                    }
                                ] 
                              } 
                         })
       
        if debug == True :
            print json.dumps(ws_definition_data)
            

    ##############################################################################################

    # process <district>_inner_to_pa worksheets - format is similar for all 3 so use 1 function
    
        ws_soe_inner_to_pa = "SOE-VRF P2P Inner-to-PA"
        ws_sde_inner_to_pa = "SDE VRF P2P Inner-to-PA"
        ws_gis_inner_to_pa = "GIS VRF P2P Inner-to-PA"
  
        inner_ws = [ ws_soe_inner_to_pa,ws_sde_inner_to_pa,ws_gis_inner_to_pa]
        final_all_inner_data = {}

        for inner in inner_ws: 

            try:
                    wb.active = worksheets.index(inner)
            except:
                    print "Worksheet %s not found" % inner
                    sys.exit(9)
            else:
                    district = inner[:3]
                    inner_data = get_inner_to_pa(wb,district)
                    final_all_inner_data.update(inner_data)
        
        if debug == True :
           print json.dumps(final_all_inner_data)
            
    
    ############################################################################################## 

    # Load BGP ASN numbers - this is hardcoded as it will not change
    bgp_asn = {}

    bgp_asn =  { 'SOE' : { 'Outer' : { 'dc1' :  65500, 'dc2' : 65510 } } }
    bgp_asn['SOE'].update({'Inner' : { 'dc1' :  65501, 'dc2' : 65511 } } )
    
    bgp_asn.update({'GIS' : { 'Inner' : { 'dc1' :  65502, 'dc2' : 65512 } } } )
    bgp_asn['GIS'].update(  { 'Outer' : { 'dc1' :  65500, 'dc2' : 65510 } } )

    bgp_asn.update({'SDE' : { 'Inner' : { 'dc1' :  65506, 'dc2' : 65516 } } } )
    bgp_asn['SDE'].update(  { 'Outer' : { 'dc1' :  65505, 'dc2' : 65515 } } )
   
    if debug == True:
        print json.dumps(bgp_asn)

    #############################################################################################

    # Load Loopback assignments
    ws_loopback = "Loopback Assignment"
    loopback_data = {}

    try:
            wb.active = worksheets.index(ws_loopback)
    except:
            print "Worksheet %s not found" % ws_loopback
            sys.exit(9)
    else:
            loopback_data = get_loopback(wb)
      
    if debug == True :
       print json.dumps(loopback_data)


    ##############################################################################################
    # Load outer VDC to PA FW
    ws_outer_to_pa = "P2P-Outer-VDC"
    outer_to_pa_data = {}

    try:
            wb.active = worksheets.index(ws_outer_to_pa)
    except:
            print "Worksheet %s not found" % ws_outer_to_pa
            sys.exit(9)
    else:
            outer_to_pa_data = get_outer_to_pa(wb)
      
    if debug == True :
       print json.dumps(outer_to_pa_data)

    ##############################################################################################
    
    # Interfaces to PA Firewall - hardcoded - this will most likely not change
    
    n7k_fw_int = {}

    n7k_fw_int =  { 'SOE' : { 'Outer' : { 'N7K-A' :  { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-B' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-C' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'}  }, 'N7K-D' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} } }  } }
    n7k_fw_int['SOE'].update({'Inner' : { 'N7K-A' :  { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-B' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-C' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'}  }, 'N7K-D' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} } }   })
    
    n7k_fw_int.update({'GIS' : { 'Outer' : { 'N7K-A' :  { 'dc1'  : { 'int1' : 'E2/22', 'int2' : 'E2/30' }, 'dc2' : { 'int1' : 'E2/22', 'int2' : 'E2/30'} }, 'N7K-B' : { 'dc1'  : { 'int1' : 'E2/22', 'int2' : 'E2/30' }, 'dc2' : { 'int1' : 'E2/22', 'int2' : 'E2/30'} }, 'N7K-C' : { 'dc1'  : { 'int1' : 'E2/22', 'int2' : 'E2/30' }, 'dc2' : { 'int1' : 'E2/22', 'int2' : 'E2/30'}  }, 'N7K-D' : { 'dc1'  : { 'int1' : 'E2/22', 'int2' : 'E2/30' }, 'dc2' : { 'int1' : 'E2/22', 'int2' : 'E2/30'} } }  } })
    n7k_fw_int['GIS'].update({'Inner' : { 'N7K-A' :  { 'dc1'  : { 'int1' : 'E2/13', 'int2' : 'E2/14' }, 'dc2' : { 'int1' : 'E2/13', 'int2' : 'E2/14'} }, 'N7K-B' : { 'dc1'  : { 'int1' : 'E2/13', 'int2' : 'E2/14' }, 'dc2' : { 'int1' : 'E2/13', 'int2' : 'E2/14'} }, 'N7K-C' : { 'dc1'  : { 'int1' : 'E2/13', 'int2' : 'E2/14' }, 'dc2' : { 'int1' : 'E2/13', 'int2' : 'E2/14'}  }, 'N7K-D' : { 'dc1'  : { 'int1' : 'E2/13', 'int2' : 'E2/14' }, 'dc2' : { 'int1' : 'E2/13', 'int2' : 'E2/14'} } }   })
    
    n7k_fw_int.update({'SDE' : { 'Outer' : { 'N7K-E' :  { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }, 'N7K-F' : { 'dc1'  : { 'int1' : 'E2/21', 'int2' : 'E2/29' }, 'dc2' : { 'int1' : 'E2/21', 'int2' : 'E2/29'} }  } } })
    n7k_fw_int['SDE'].update({'Inner' : { 'N7K-E' :  { 'dc1'  : { 'int1' : 'E2/5', 'int2' : 'E2/13' }, 'dc2' : { 'int1' : 'E2/5', 'int2' : 'E2/13'} }, 'N7K-F' : { 'dc1'  : { 'int1' : 'E2/5', 'int2' : 'E2/13' }, 'dc2' : { 'int1' : 'E2/5', 'int2' : 'E2/13'} } }   })
    
    
    if debug == True:
        print json.dumps(n7k_fw_int)
    
    
    
    
    ##############################################################################################
    
    return ws_definition_data,final_all_inner_data,bgp_asn,outer_to_pa_data,n7k_fw_int,loopback_data

def main(argv):

    debug  = False
    configure = False

    if len(argv) == 0:
        print "Usage: " +  sys.argv[0] + " -f|--file <excel file name> -d|--debug.  No arguments given"
        print "-d|--debug:  Prints excel data in JSON format (no switch changes made)"
        print "-e|--execute <n7k list>: Invokes N7K API to configure N7K. <n7k list> file must be in this format:"
        print "        <SOE|GIS|SDE>,<DC1|DC2>,<N7K-A|N7K-B|N7K-C|N7K-D>,<IP>"
        sys.exit(1)

    try:
        opts,args = getopt.getopt(argv,"f:hde:",["file=","help","debug","n7k="])
    except getopt.GetoptError as err:
        print str(err)
        sys.exit(2)
    else:
        for opt,arg in opts:
            if opt in ("-d","--debug"):
                debug = True
            if opt in ("-e", "--execute"):
                configure = True
                vdcfile = arg
                if not os.path.isfile(vdcfile):
                    print sys.argv[0] + " VDC File list %s NOT found" % vdcfile
                    sys.exit(1)
                
        if debug is True and configure is True:
            print "Cannot use option d and option e together.  Use option e or option d"
            sys.exit(9)
         
         
        for opt,arg in opts:
            if opt == '-h':
                print sys.argv[0] + " -f|--file <excel file name> -d|--debug"
                print "-d|--debug:  Prints excel data in JSON format (no switch changes made)"
                print "-e|--execute <n7k list>: Invokes N7K API to configure N7K. <n7k list> file must be in this format:"
                print "        <SOE|GIS|SDE>,<DC1|DC2>,<N7K-A|N7K-B|N7K-C|N7K-D>,<IP>"
        
                sys.exit(1)
            elif opt in ( "-f", "--file"):
                filename = arg
                if not os.path.isfile(filename):
                    print sys.argv[0] + " excel file %s NOT found" % filename
                    sys.exit(1)
                else:
                    try:
                        open_workbook(filename)
                    except XLRDError:
                        print sys.argv[0] + " file %s is not an Excel file" % filename
                    else:
                        if magic.from_file(filename) == 'Microsoft Excel 2007+':
                            (ws_definition_data,final_all_inner_data,bgp_asn,outer_to_pa_data,n7k_fw_int,loopback_data) = process_xlsx(filename,debug)
                            if debug is False:
                                inner_vdc_config(ws_definition_data,final_all_inner_data,bgp_asn,outer_to_pa_data,n7k_fw_int,loopback_data,configure)
                        else:
                            print "File must be in .xlsx format"
                            sys.exit(10)
                            
             
                 

if __name__ == '__main__':
    main(sys.argv[1:])
